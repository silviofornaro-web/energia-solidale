from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
EON_ROOT = Path(
    "/Users/silviofornaro/Library/CloudStorage/"
    "GoogleDrive-bolletteillumia.banco@gmail.com/Il mio Drive/E.ON/Maggio 2026"
)
OUTPUT = ROOT / "estrazioni_tariffe" / "eon_tariffe_2026-05.xlsx"


@dataclass(frozen=True)
class PdfSource:
    path: Path
    segment: str
    commodity: str
    offer_name: str
    offer_type: str
    fallback_valid_from: str
    fallback_valid_to: str


SOURCES = [
    PdfSource(EON_ROOT / "residenziali/E.ON Flex Gas Compensato_6R_D_FLX_CLSB.pdf", "RESIDENZIALE", "GAS", "E.ON Flex Gas", "VARIABILE", "2026-05-08", "2026-05-21"),
    PdfSource(EON_ROOT / "residenziali/E.ON Flex Gas Compensato_6R_D_P12_CLSE.pdf", "RESIDENZIALE", "GAS", "E.ON Flex Gas Protetta", "VARIABILE", "2026-05-08", "2026-05-21"),
    PdfSource(EON_ROOT / "residenziali/E.ON Gas Verde Compensato_6R_D_CLSA.pdf", "RESIDENZIALE", "GAS", "E.ON Gas Tua", "FISSA", "2026-05-08", "2026-05-21"),
    PdfSource(EON_ROOT / "residenziali/E.ON ValoreMercato Luce_6R_D_FLX_CLSC.pdf", "RESIDENZIALE", "EE", "E.ON Flex Luce", "VARIABILE", "2026-05-08", "2026-05-21"),
    PdfSource(EON_ROOT / "residenziali/E.ON ValoreMercato Luce_6R_D_ACR_CLSE.pdf", "RESIDENZIALE", "EE", "E.ON Flex Luce Casa", "VARIABILE", "2026-05-08", "2026-05-21"),
    PdfSource(EON_ROOT / "residenziali/E.ON Luce Comfort_6R_D_RI_CLSD.pdf", "RESIDENZIALE", "EE", "E.ON Flex Luce Per24", "VARIABILE", "2026-05-08", "2026-05-21"),
    PdfSource(EON_ROOT / "residenziali/E.ON LuceVerde_6R_D_CLSA.pdf", "RESIDENZIALE", "EE", "E.ON Luce Tua", "FISSA", "2026-05-08", "2026-05-21"),
    PdfSource(EON_ROOT / "residenziali/E.ON Luce Relax_6R_D_CLSB.pdf", "RESIDENZIALE", "EE", "E.ON Luce Tua Per 24", "FISSA", "2026-05-08", "2026-05-21"),
    PdfSource(EON_ROOT / "residenziali/E.ON LuceVerde_6R_D_ACR_CLSE.pdf", "RESIDENZIALE", "EE", "E.ON LuceVerde Casa", "FISSA", "2026-05-08", "2026-05-21"),
    PdfSource(EON_ROOT / "Micro Business/E.ON Gas Impresa_6Q_M_CLSC.pdf", "MICROBUSINESS", "GAS", "E.ON Gas Impresa CLSC", "VARIABILE", "2026-04-23", "2026-05-20"),
    PdfSource(EON_ROOT / "Micro Business/E.ON Gas Impresa_6Q_M_CLSE.pdf", "MICROBUSINESS", "GAS", "E.ON Gas Impresa CLSE", "VARIABILE", "2026-04-23", "2026-05-20"),
    PdfSource(EON_ROOT / "Micro Business/E.ON ChiaraGas Rinnovo_6Q_M_CLSC.pdf", "MICROBUSINESS", "GAS", "E.ON ChiaraGas Rinnovo CLSC", "FISSA", "2026-04-23", "2026-05-20"),
    PdfSource(EON_ROOT / "Micro Business/E.ON ChiaraGas Rinnovo_6Q_M_CLSE.pdf", "MICROBUSINESS", "GAS", "E.ON ChiaraGas Rinnovo CLSE", "FISSA", "2026-04-23", "2026-05-20"),
    PdfSource(EON_ROOT / "Micro Business/E.ON LuceDinamica ECO_6Q_M_CLSC.pdf", "MICROBUSINESS", "EE", "E.ON LuceDinamica ECO CLSC", "VARIABILE", "2026-04-23", "2026-05-20"),
    PdfSource(EON_ROOT / "Micro Business/E.ON LuceDinamica ECO_6Q_M_CLSE.pdf", "MICROBUSINESS", "EE", "E.ON LuceDinamica ECO CLSE", "VARIABILE", "2026-04-23", "2026-05-20"),
    PdfSource(EON_ROOT / "Micro Business/E.ON EnergiaChiara - ECO_6Q_M_CLSC.pdf", "MICROBUSINESS", "EE", "E.ON EnergiaChiara ECO CLSC", "FISSA", "2026-04-23", "2026-05-20"),
    PdfSource(EON_ROOT / "Micro Business/E.ON EnergiaChiara - ECO_6Q.2_M_CLSE.pdf", "MICROBUSINESS", "EE", "E.ON EnergiaChiara ECO CLSE", "FISSA", "2026-04-23", "2026-05-20"),
    PdfSource(EON_ROOT / "Business/E.ON Profilo Dinamico Gas_6Q_P.pdf", "BUSINESS", "GAS", "E.ON Profilo Dinamico Gas P", "VARIABILE", "2026-05-05", "2026-05-12"),
    PdfSource(EON_ROOT / "Business/E.ON Profilo Dinamico Gas_6Q_R.pdf", "BUSINESS", "GAS", "E.ON Profilo Dinamico Gas R", "VARIABILE", "2026-05-05", "2026-05-12"),
    PdfSource(EON_ROOT / "Business/E.ON_ProfiloSicuro_6Q_B.pdf", "BUSINESS", "EE", "E.ON Profilo Sicuro B", "FISSA", "2026-05-05", "2026-05-12"),
    PdfSource(EON_ROOT / "Business/E.ON_ProfiloSicuro_6Q_S.pdf", "BUSINESS", "EE", "E.ON Profilo Sicuro S", "FISSA", "2026-05-05", "2026-05-12"),
    PdfSource(EON_ROOT / "Business/E.ON_ProfiloSicuro_6Q_T.pdf", "BUSINESS", "EE", "E.ON Profilo Sicuro T", "FISSA", "2026-05-05", "2026-05-12"),
]


def read_pdf_text(path: Path) -> str:
    reader = PdfReader(str(path))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def compact(text: str) -> str:
    return re.sub(r"\s+", " ", text)


def number(value: str | None) -> float | None:
    if value is None:
        return None
    return float(value.replace(".", "").replace(",", "."))


def date_it(value: str | None) -> str:
    if not value:
        return ""
    return datetime.strptime(value, "%d/%m/%Y").date().isoformat()


def first(patterns: list[str], text: str, flags: int = re.IGNORECASE | re.DOTALL) -> str | None:
    for pattern in patterns:
        match = re.search(pattern, text, flags)
        if match:
            return match.group(1)
    return None


def first_pair(patterns: list[str], text: str, flags: int = re.IGNORECASE | re.DOTALL) -> tuple[float | None, float | None]:
    for pattern in patterns:
        match = re.search(pattern, text, flags)
        if match:
            return number(match.group(1)), number(match.group(2))
    return None, None


def extract_common(src: PdfSource, raw_text: str, text: str) -> dict[str, Any]:
    codes = re.findall(r"000362[A-Z0-9]+", raw_text)
    valid_to = date_it(first([r"Scadenza:?\s*(\d{2}/\d{2}/\d{4})"], text)) or src.fallback_valid_to
    valid_dates = [date_it(v) for v in re.findall(r"alla data(?: del)?\s+(\d{2}/\d{2}/\d{4})", text, re.IGNORECASE)]
    valid_from = next((d for d in valid_dates if d and d != valid_to), src.fallback_valid_from)
    bonus = None
    if not re.search(r"non (?:sono previsti|prevede) sconti", text, re.IGNORECASE):
        bonus = number(first([r"Bonus(?:\s+di)?\s+([0-9]+,[0-9]+)\s*€"], text))
    return {
        "codice_offerta": codes[0] if codes else "",
        "valid_from": valid_from,
        "valid_to": valid_to,
        "bonus": bonus,
    }


def fixed_fee_total(values: dict[str, Any]) -> float | None:
    parts = [values.get("ccv_quota_fissa"), values.get("gestione_energetica_fissa")]
    total = sum(v for v in parts if v is not None)
    return total or None


def extract_energy_values(src: PdfSource, text: str) -> dict[str, Any]:
    values: dict[str, Any] = {}
    values["ccv_quota_fissa"] = number(
        first(
            [
                r"Corrispettivo Annuo\s+([0-9]+,[0-9]+)\s*€/P[ODRdrpP]+/anno",
                r"Corrispettivo commercializzazione al dettaglio quota fissa\s+([0-9]+,[0-9]+)\s*€/P[ODRdrpP]+/anno",
                r"Corrispettivo di commercializzazione al dettaglio\s+([0-9]+,[0-9]+)\s*€/P[ODRdrpP]+/anno",
                r"Corrispettivo commercializzazione e vendita\s+([0-9]+,[0-9]+)\s*€/P[ODRdrpP]+/anno",
                r"Commercializzazione e vendita solo BT\s+([0-9]+,[0-9]+)\s*€/POD/anno",
                r"commercializzazione e vendita,\s+pari a\s+([0-9]+,[0-9]+)\s*€/P[ODRdrpP]+/anno",
                r"Commercializzazione vendita in €/POD/anno\s+([0-9]+,[0-9]+)",
            ],
            text,
        )
    )
    values["ccv_quota_variabile"] = number(
        first([r"quota variabile\s+([0-9]+,[0-9]+)\s*€/Smc"], text)
    )
    values["gestione_energetica_fissa"] = number(
        first([r"gestione energetica\s+([0-9]+(?:,[0-9]+)?)\s*€/P[ODRdrpP]+/anno"], text)
    )

    if src.commodity == "GAS":
        values["omega"] = number(
            first(
                [
                    r"Corrispettivo per il consumo\s+PSV DA\s*\+\s*Ω:\s*([0-9]+,[0-9]+)\s*€/Smc",
                    r"Corrispettivo per il consumo\s+PSV DA\s*\+\s*([0-9]+,[0-9]+)\s*€/Smc",
                    r"corrispettivo variabile\s*(?:Ω|OMEGA).*?pari a\s+([0-9]+,[0-9]+)\s*€/Smc",
                ],
                text,
            )
        )
        values["prezzo_fisso"] = number(
            first(
                [
                    r"Corrispettivo per il consumo\s+([0-9]+,[0-9]+)\s*€/Smc",
                    r"Materia Prima Gas.*?pari\s+a\s+([0-9]+,[0-9]+)\s*€/Smc",
                ],
                text,
            )
        )
        return values

    values["omega"] = number(
        first(
            [
                r"Corrispettivo per il consumo\s+PUN Index GMEFascia\s*\+\s*Ω:?\s*([0-9]+,[0-9]+)\s*€/kWh",
                r"PUN GME\s*\+\s*([0-9]+,[0-9]+)\s*€/kWh",
                r"corrispettivo variabile.*?pari a\s+([0-9]+,[0-9]+)\s*€/kWh",
            ],
            text,
        )
    )
    values["prezzo_mono"] = number(
        first(
            [
                r"Corrispettivo per il consumo\s+Monoraria:\s*([0-9]+,[0-9]+)\s*€/kWh",
                r"Monoraria:\s*([0-9]+,[0-9]+)\s*€/kWh",
                r"Prezzo Fascia F0\s+pari\s+a\s+([0-9]+,[0-9]+)\s*€/kWh",
            ],
            text,
        )
    )
    values["prezzo_f1"], values["prezzo_f23"] = first_pair(
        [r"Opzione 3 Fasce.*?F12\s+([0-9]+,[0-9]+)\s*€/kWh\s*-\s*F3\s+([0-9]+,[0-9]+)\s*€/kWh"],
        text,
    )
    return values


def row(src: PdfSource, common: dict[str, Any], component: str, uom: str, value: float | None, notes: str) -> dict[str, Any]:
    return {
        "provider": "E.ON",
        "segmento": src.segment,
        "commodity": src.commodity,
        "offer_type": src.offer_type,
        "offer_name": src.offer_name,
        "valid_from": common["valid_from"],
        "valid_to": common["valid_to"],
        "component": component,
        "uom": uom,
        "value": value,
        "index_name": "PUN" if src.commodity == "EE" and src.offer_type == "VARIABILE" else "PSV" if src.commodity == "GAS" and src.offer_type == "VARIABILE" else "",
        "notes": notes,
        "codice_offerta": common["codice_offerta"],
        "source_pdf": src.path.name,
    }


def build_tariff_rows() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    tariff_rows: list[dict[str, Any]] = []
    summary_rows: list[dict[str, Any]] = []

    for src in SOURCES:
        if not src.path.exists():
            raise FileNotFoundError(src.path)
        raw_text = read_pdf_text(src.path)
        text = compact(raw_text)
        common = extract_common(src, raw_text, text)
        values = extract_energy_values(src, text)
        fixed_total = fixed_fee_total(values)

        if src.commodity == "GAS" and src.offer_type == "VARIABILE":
            tariff_rows.append(row(src, common, "fee_energia", "€/Smc", values.get("omega"), "Omega su PSV DA."))
            tariff_rows.append(row(src, common, "ccv_quota_fissa", "€/anno", fixed_total, "Quota fissa annua; per microbusiness include anche gestione energetica fissa."))
            if values.get("ccv_quota_variabile") is not None:
                tariff_rows.append(row(src, common, "ccv_quota_variabile", "€/Smc", values.get("ccv_quota_variabile"), "Commercializzazione al dettaglio quota variabile."))
        elif src.commodity == "GAS":
            tariff_rows.append(row(src, common, "prezzo_fisso", "€/Smc", values.get("prezzo_fisso"), "Materia Prima Gas fissa."))
            tariff_rows.append(row(src, common, "ccv_quota_fissa", "€/anno", fixed_total, "Quota fissa annua; per microbusiness include anche gestione energetica fissa."))
            if values.get("ccv_quota_variabile") is not None:
                tariff_rows.append(row(src, common, "ccv_quota_variabile", "€/Smc", values.get("ccv_quota_variabile"), "Commercializzazione al dettaglio quota variabile."))
        elif src.commodity == "EE" and src.offer_type == "VARIABILE":
            tariff_rows.append(row(src, common, "fee_energia", "€/kWh", values.get("omega"), "Omega su PUN. PUN da considerare con perdite di rete 10%."))
            tariff_rows.append(row(src, common, "perdite_rete", "%", 0.10, "Perdite rete bassa tensione: fattore 1,100."))
            tariff_rows.append(row(src, common, "ccv_quota_fissa", "€/anno", fixed_total, "Quota fissa annua; per microbusiness include anche gestione energetica fissa."))
        else:
            tariff_rows.append(row(src, common, "prezzo_mono", "€/kWh", values.get("prezzo_mono"), "Prezzo monorario."))
            if values.get("prezzo_f1") is not None:
                tariff_rows.append(row(src, common, "prezzo_f1", "€/kWh", values.get("prezzo_f1"), "Prezzo fascia F1/F2."))
            if values.get("prezzo_f23") is not None:
                tariff_rows.append(row(src, common, "prezzo_f23", "€/kWh", values.get("prezzo_f23"), "Prezzo fascia F3."))
            tariff_rows.append(row(src, common, "ccv_quota_fissa", "€/anno", fixed_total, "Quota fissa annua; per microbusiness include anche gestione energetica fissa."))

        if common.get("bonus") is not None:
            tariff_rows.append(row(src, common, "sconto_bonus", "€ una tantum", -abs(common["bonus"]), "Bonus commerciale E.ON, distinto dal Bonus Sociale statale."))

        summary_rows.append(
            {
                "provider": "E.ON",
                "segmento": src.segment,
                "commodity": src.commodity,
                "offer_type": src.offer_type,
                "offer_name": src.offer_name,
                "valid_from": common["valid_from"],
                "valid_to": common["valid_to"],
                "codice_offerta": common["codice_offerta"],
                "omega": values.get("omega") if src.offer_type == "VARIABILE" else None,
                "prezzo_fisso": values.get("prezzo_fisso") if src.commodity == "GAS" and src.offer_type == "FISSA" else None,
                "prezzo_mono": values.get("prezzo_mono"),
                "prezzo_f1": values.get("prezzo_f1"),
                "prezzo_f23": values.get("prezzo_f23"),
                "ccv_quota_fissa_totale_annua": fixed_total,
                "ccv_quota_variabile": values.get("ccv_quota_variabile"),
                "gestione_energetica_fissa": values.get("gestione_energetica_fissa"),
                "bonus_commerciale": -abs(common["bonus"]) if common.get("bonus") is not None else None,
                "source_pdf": src.path.name,
            }
        )
    return tariff_rows, summary_rows


def write_sheet(ws, rows: list[dict[str, Any]], headers: list[str]) -> None:
    ws.append(headers)
    for item in rows:
        ws.append([item.get(h, "") for h in headers])
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, header in enumerate(headers, start=1):
        width = max(len(str(header)) + 2, 12)
        for value in [ws.cell(row=r, column=idx).value for r in range(2, min(ws.max_row, 80) + 1)]:
            width = min(max(width, len(str(value)) + 2 if value is not None else width), 48)
        ws.column_dimensions[get_column_letter(idx)].width = width


def main() -> None:
    tariff_rows, summary_rows = build_tariff_rows()
    wb = Workbook()
    ws_tariffe = wb.active
    ws_tariffe.title = "tariffe"
    tariff_headers = [
        "provider",
        "segmento",
        "commodity",
        "offer_type",
        "offer_name",
        "valid_from",
        "valid_to",
        "component",
        "uom",
        "value",
        "index_name",
        "notes",
        "codice_offerta",
        "source_pdf",
    ]
    write_sheet(ws_tariffe, tariff_rows, tariff_headers)

    ws_summary = wb.create_sheet("riepilogo_offerte")
    summary_headers = [
        "provider",
        "segmento",
        "commodity",
        "offer_type",
        "offer_name",
        "valid_from",
        "valid_to",
        "codice_offerta",
        "omega",
        "prezzo_fisso",
        "prezzo_mono",
        "prezzo_f1",
        "prezzo_f23",
        "ccv_quota_fissa_totale_annua",
        "ccv_quota_variabile",
        "gestione_energetica_fissa",
        "bonus_commerciale",
        "source_pdf",
    ]
    write_sheet(ws_summary, summary_rows, summary_headers)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Creato: {OUTPUT}")
    print(f"Offerte: {len(summary_rows)}")
    print(f"Righe tariffarie: {len(tariff_rows)}")


if __name__ == "__main__":
    main()
