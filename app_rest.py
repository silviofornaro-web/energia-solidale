import io
import os
import re
import base64
import calendar
import json
import hashlib
import hmac
import html
from copy import copy
from pathlib import Path
from datetime import datetime, date
from zoneinfo import ZoneInfo

import streamlit as st
import streamlit.components.v1 as components
import openpyxl
from openpyxl.styles import Font

st.set_page_config(page_title="Confronto bollette", layout="wide")

BASE_DIR = Path(__file__).parent
LOGO_JPG = BASE_DIR / "assets" / "logo_energia_solidale.jpg"
LOGO_PNG = BASE_DIR / "assets" / "logo_energia_solidale.png"
TEMPLATE_XLSX = BASE_DIR / "esempio_confronto_corretto.xlsx"
TARIFFE_BASE = BASE_DIR / "tariffe"
EON_TARIFFE_BASE = BASE_DIR / "estrazioni_tariffe"
INDICI_XLSX = BASE_DIR / "indici_pun_psv_2025_2026.xlsx"
STATIC_DIR = BASE_DIR / "static"
STATIC_DOWNLOADS_DIR = STATIC_DIR / "downloads"
APP_STATE_VERSION = "2026-05-16-dashboard-shot-1"
APP_TZ = ZoneInfo("Europe/Rome")
DEFAULT_RENDER_AUTH_USERS = {
    "silviofornaro@gmail.com": {
        "name": "silvio fornaro",
        "password_hash": "pbkdf2_sha256$260000$7c986bc474e411dd6b0f2381629044ac$H7u9e7yi8DsLqYA6GvJVHJtaKd0a291Fq2zjGRivfMM=",
    },
    "ioilre58@gmail.com": {
        "name": "roberto marcon",
        "password_hash": "pbkdf2_sha256$260000$badb8d5e6e1b280be351452d51a3fdf9$CfhFQhA1NXPOJXm+KVB6dGpEEHc6b/2Y7yf0ZZj85Y0=",
    },
}
EE_EXCISE_RATE = 0.0227
EE_RESIDENTIAL_EXEMPT_KWH_PER_MONTH = 150.0
ACCESSORY_SERVICES_VAT_RATES = {
    "22%": 0.22,
    "10%": 0.10,
    "Esente": 0.0,
}
ACCESSORY_SERVICES_VAT_OPTIONS = tuple(ACCESSORY_SERVICES_VAT_RATES)
GAS_VAT_REDUCED_ANNUAL_THRESHOLD = 480.0
GAS_EXCISE_BRACKETS = (
    (120.0, 0.044),
    (480.0, 0.175),
    (1560.0, 0.170),
    (None, 0.186),
)
GAS_REGIONAL_ADDITIONAL_MIN_RATES = {
    "Abruzzo": 0.02,
    "Basilicata": 0.02,
    "Calabria": 0.02,
    "Campania": 0.02,
    "Emilia-Romagna": 0.02,
    "Friuli-Venezia Giulia": 0.0,
    "Lazio": 0.02,
    "Liguria": 0.02,
    "Lombardia": 0.0,
    "Marche": 0.02,
    "Molise": 0.02,
    "Piemonte": 0.022,
    "Puglia": 0.02,
    "Sardegna": 0.0,
    "Sicilia": 0.0,
    "Toscana": 0.02,
    "Trentino-Alto Adige": 0.0,
    "Umbria": 0.02,
    "Valle d'Aosta": 0.0,
    "Veneto": 0.007747,
}
REGION_OPTIONS = tuple(GAS_REGIONAL_ADDITIONAL_MIN_RATES)
SEGMENT_LABEL_TO_VALUE = {
    "Residenziale": "RESIDENZIALE",
    "Business": "BUSINESS",
}
SEGMENT_VALUE_TO_LABEL = {value: label for label, value in SEGMENT_LABEL_TO_VALUE.items()}
SUPPLY_LABEL_TO_VALUE = {
    "Gas": "GAS",
    "Luce": "EE",
}
SUPPLY_VALUE_TO_LABEL = {value: label for label, value in SUPPLY_LABEL_TO_VALUE.items()}
SEGMENT_OPTIONS = tuple(SEGMENT_LABEL_TO_VALUE)
SUPPLY_OPTIONS = tuple(SUPPLY_LABEL_TO_VALUE)
BILL_OFFER_TYPE_OPTIONS = ("Fissa", "Variabile")
SUPPLIER_LABEL_TO_CODE = {
    "Illumia": "ILLUMIA",
    "E-ON": "EON",
}
SUPPLIER_CODE_TO_LABEL = {value: label for label, value in SUPPLIER_LABEL_TO_CODE.items()}
SUPPLIER_OPTIONS = tuple(SUPPLIER_LABEL_TO_CODE)
ITALIAN_MONTHS = (
    "",
    "Gennaio",
    "Febbraio",
    "Marzo",
    "Aprile",
    "Maggio",
    "Giugno",
    "Luglio",
    "Agosto",
    "Settembre",
    "Ottobre",
    "Novembre",
    "Dicembre",
)

if str(st.query_params.get("debug", "")).lower() in {"1", "true", "si", "yes"}:
    st.title("Energia Solidale")
    st.success("Modalita test: Streamlit sta renderizzando la pagina.")
    st.write(
        "Se vedi questo messaggio, il browser riesce a mostrare Streamlit. "
        "Il problema e in un componente della pagina principale."
    )
    print("APP_DEBUG_RENDER", flush=True)
    st.stop()

# Legacy fallback (solo se non hai ancora tariffe/)
LEGACY_RES_FILES = [
    BASE_DIR / "tariffe_illumia_template.xlsx",
    BASE_DIR / "tariffe_illumia_templatemarzo.xlsx",
]
LEGACY_RES_FOUND = [p for p in LEGACY_RES_FILES if p.exists()]

if not TEMPLATE_XLSX.exists():
    st.error("❌ Manca esempio_confronto_corretto.xlsx nella cartella del progetto.")
    st.stop()

KEYS = [
    "vendita_consumo",
    "vendita_fissa",
    "rete_consumi",
    "rete_fissa",
    "quota_potenza",
    "sconti",
    "ricalcoli",
    "bonus_sociale",
    "arrotondamenti",
    "servizi_accessori",
    "accise_iva",
]

def ss(k, v):
    if k not in st.session_state:
        st.session_state[k] = v

def reset_state_on_app_update():
    if str(st.query_params.get("reset", "")).lower() in {"1", "true", "si", "yes"}:
        st.session_state.clear()
        st.query_params.clear()
        st.rerun()
    if st.session_state.get("_app_state_version") == APP_STATE_VERSION:
        return
    keep = {
        "auth_logged_in": st.session_state.get("auth_logged_in", False),
        "auth_email": st.session_state.get("auth_email", ""),
        "auth_name": st.session_state.get("auth_name", ""),
    }
    st.session_state.clear()
    for key, value in keep.items():
        if value:
            st.session_state[key] = value
    st.session_state["_app_state_version"] = APP_STATE_VERSION

reset_state_on_app_update()

# Stato base
ss("segmento", "RESIDENZIALE")  # RESIDENZIALE / BUSINESS
ss("nome_cliente", "")
ss("nome_cliente_input", st.session_state["nome_cliente"] or "Cliente")
ss("commodity", "GAS")          # GAS / EE
if st.session_state["commodity"] not in ("GAS", "EE"):
    st.session_state["commodity"] = "GAS"
ss("comparison_supplier_select", "Illumia")
if st.session_state["comparison_supplier_select"] not in SUPPLIER_LABEL_TO_CODE:
    st.session_state["comparison_supplier_select"] = "Illumia"
ss("comparison_supplier", SUPPLIER_LABEL_TO_CODE[st.session_state["comparison_supplier_select"]])
ss("prev_comparison_supplier", st.session_state["comparison_supplier"])
ss("segmento_select", SEGMENT_VALUE_TO_LABEL.get(st.session_state["segmento"], "Residenziale"))
ss("commodity_select", SUPPLY_VALUE_TO_LABEL.get(st.session_state["commodity"], "Gas"))
ss("prev_segmento", st.session_state["segmento"])
ss("prev_commodity", st.session_state["commodity"])
ss("consumo", 0.0)

# Periodo bolletta (solo informativo/controllo; NON è validità offerta)
ss("bill_start", date.today())
ss("bill_end", date.today())
ss("bill_start_day", st.session_state["bill_start"].day)
ss("bill_start_month", st.session_state["bill_start"].month)
ss("bill_start_year", st.session_state["bill_start"].year)
ss("bill_end_day", st.session_state["bill_end"].day)
ss("bill_end_month", st.session_state["bill_end"].month)
ss("bill_end_year", st.session_state["bill_end"].year)

# Periodicità (per quota fissa scontrino)
ss("billing_divisor", 12)
ss("billing_months", 1)
ss("assume_fixed_is_monthly", True)

BILLING_PERIODS = {
    1: "MENSILE",
    2: "BIMESTRALE",
    3: "TRIMESTRALE",
}

# Bolletta B
for k in KEYS:
    ss(f"b_{k}", 0.0)

# Indici
ss("pun_override", 0.0)
ss("psv_override", 0.0)
ss("indice_month", "")
ss("indice_source", "")
ss("indice_file_path", "")
ss("include_dispbt", True)

# Output fornitore confronto
ss("c_vendita_consumo", 0.0)
ss("c_vendita_fissa", 0.0)
ss("d_vendita_consumo", 0.0)
ss("d_vendita_fissa", 0.0)
ss("c_sconti", 0.0)
ss("d_sconti", 0.0)
ss("illumia_calculated", False)

# Sconti default fornitore confronto
ss("ill_sconto_var", -3.0)
ss("ill_sconto_fix", -3.0)

# Tariffe: upload override
ss("tariffe_uploaded_bytes", None)
ss("offers_loaded", False)

# Offerte selezionate automaticamente dall'app
ss("offer_var_name", "")
ss("offer_fix_name", "")

# Validità offerta (presa dal file tariffe più recente)
ss("offer_valid_from", None)
ss("offer_valid_to", None)
ss("offer_file_path", "")

# UI
ss("export_mode", "ENTRAMBE")
ss("excel_ready", False)
ss("comparison_generated_at", "")
ss("tax_primary_home", "Sì")
ss("tax_power_kw", 3.0)
ss("tax_annual_consumption", 0.0)
ss("tax_region", "Veneto")
ss("servizi_accessori_iva", "22%")
if st.session_state["servizi_accessori_iva"] not in ACCESSORY_SERVICES_VAT_RATES:
    st.session_state["servizi_accessori_iva"] = "22%"
ss("bill_offer_type", "Fissa")
if st.session_state["bill_offer_type"] not in BILL_OFFER_TYPE_OPTIONS:
    st.session_state["bill_offer_type"] = "Fissa"


# -----------------------------
# Utils
# -----------------------------
def norm(s: str) -> str:
    return " ".join(str(s).strip().lower().split())

def parse_number(x) -> float:
    if x is None:
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace("€", "").replace("\u20ac", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0

def parse_date_any(x):
    if x is None:
        return None
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x
    s = str(x).strip()
    # accetta più formati perché nel tuo file si vedono date tipo 03/11/2026, 04/10/2026 ecc. [1](https://onedrive.live.com/personal/8d36b8086e9d2af7/_layouts/15/doc.aspx?resid=ae62e5e9-da89-4f80-8edb-c1ae21b28205&cid=8d36b8086e9d2af7)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None

def date_from_parts(prefix):
    try:
        year = int(st.session_state[f"{prefix}_year"])
        month = int(st.session_state[f"{prefix}_month"])
        day = int(st.session_state[f"{prefix}_day"])
    except Exception:
        return date.today()

    year = min(max(year, 2000), 2100)
    month = min(max(month, 1), 12)
    day = min(max(day, 1), calendar.monthrange(year, month)[1])
    return date(year, month, day)

def date_from_month_year(prefix, end_of_month=False):
    try:
        year = int(st.session_state[f"{prefix}_year"])
        month = int(st.session_state[f"{prefix}_month"])
    except Exception:
        today = date.today()
        year = today.year
        month = today.month

    year = min(max(year, 2000), 2100)
    month = min(max(month, 1), 12)
    day = calendar.monthrange(year, month)[1] if end_of_month else 1
    return date(year, month, day)

def truthy(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("true", "1", "yes", "y", "si", "sì"):
        return True
    if s in ("false", "0", "no", "n"):
        return False
    return None

def current_comparison_timestamp():
    return datetime.now(APP_TZ).isoformat()

def format_comparison_timestamp(value=None):
    raw_value = value if value is not None else st.session_state.get("comparison_generated_at", "")
    if isinstance(raw_value, datetime):
        generated_at = raw_value
    else:
        try:
            generated_at = datetime.fromisoformat(str(raw_value))
        except Exception:
            return "N.D."
    return generated_at.strftime("%d/%m/%Y %H:%M:%S")

def format_month_year_label(d):
    if not isinstance(d, date):
        return "N.D."
    return f"{ITALIAN_MONTHS[d.month]} {d.year}"

def format_bill_period_label():
    start = st.session_state.get("bill_start")
    end = st.session_state.get("bill_end")
    return f"{format_month_year_label(start)} - {format_month_year_label(end)}"

def format_offer_expiration_label():
    offer_valid_to = st.session_state.get("offer_valid_to")
    if isinstance(offer_valid_to, datetime):
        offer_valid_to = offer_valid_to.date()
    if isinstance(offer_valid_to, date):
        return offer_valid_to.strftime("%d/%m/%Y")
    return "N.D."

def selected_supplier_label():
    label = st.session_state.get("comparison_supplier_select", "Illumia")
    return label if label in SUPPLIER_LABEL_TO_CODE else "Illumia"

def selected_supplier_code():
    return SUPPLIER_LABEL_TO_CODE[selected_supplier_label()]

def supplier_name_for_code(code: str):
    return SUPPLIER_CODE_TO_LABEL.get(str(code).upper(), "Illumia")

def supplier_column_labels():
    supplier = selected_supplier_label()
    return f"{supplier} Variabile", f"{supplier} Fissa"

def supplier_slug():
    return selected_supplier_label().lower().replace("-", "").replace(".", "").replace(" ", "_")

def comparison_summary_fields():
    return (
        (
            ("Cliente", st.session_state.get("nome_cliente", "Cliente")),
            ("Tipo tariffa bolletta", st.session_state.get("bill_offer_type", "Fissa")),
            ("Periodo", format_bill_period_label()),
            ("Indice usato", st.session_state.get("indice_month") or "N.D."),
            ("Offerta variabile", st.session_state.get("offer_var_name") or "N.D."),
        ),
        (
            ("Confronto eseguito", format_comparison_timestamp()),
            ("Periodicità", billing_label_from_months(st.session_state.get("billing_months", 1))),
            ("Mesi fatturati", str(st.session_state.get("billing_months", 1))),
            ("Fornitore confronto", selected_supplier_label()),
            ("Offerta fissa", st.session_state.get("offer_fix_name") or "N.D."),
            ("Scadenza offerta", format_offer_expiration_label()),
        ),
    )

def get_auth_config():
    try:
        auth = st.secrets.get("auth", {})
        if auth:
            return auth
    except Exception:
        pass

    render_hostname = os.environ.get("RENDER_EXTERNAL_HOSTNAME", "").strip()
    env_auth_enabled = truthy(os.environ.get("AUTH_ENABLED"))
    if env_auth_enabled is not True and not render_hostname:
        return {}

    users = {}
    names = {}
    if render_hostname or env_auth_enabled is True:
        for email, data in DEFAULT_RENDER_AUTH_USERS.items():
            users[email] = data["password_hash"]
            names[email] = data["name"]

    email = os.environ.get("AUTH_USER_EMAIL", "").strip().lower()
    password_hash = os.environ.get("AUTH_USER_PASSWORD_HASH", "").strip()
    name = os.environ.get("AUTH_USER_NAME", email).strip()
    if email and password_hash:
        users[email] = password_hash
        names[email] = name or email

    if not users:
        return {"enabled": True, "users": {}, "names": {}}

    return {
        "enabled": True,
        "users": users,
        "names": names,
    }

def auth_enabled():
    return truthy(get_auth_config().get("enabled", False)) is True

def verify_password(password, stored_hash):
    try:
        algo, iterations, salt, expected = str(stored_hash).split("$", 3)
        if algo != "pbkdf2_sha256":
            return False
        iterations = int(iterations)
        digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), iterations)
        actual = base64.b64encode(digest).decode("ascii")
        return hmac.compare_digest(actual, expected)
    except Exception:
        return False

def logout():
    st.session_state["auth_logged_in"] = False
    st.session_state["auth_email"] = ""
    st.session_state["auth_name"] = ""

def require_authentication():
    if not auth_enabled():
        return

    ss("auth_logged_in", False)
    ss("auth_email", "")
    ss("auth_name", "")
    ss("auth_error", "")

    if st.session_state["auth_logged_in"]:
        with st.sidebar:
            st.caption(f"Accesso: {st.session_state['auth_name'] or st.session_state['auth_email']}")
            if st.button("Esci", key="auth_logout"):
                logout()
                st.rerun()
        return

    auth = get_auth_config()
    users = auth.get("users", {})
    names = auth.get("names", {})
    if not users:
        st.error("Accesso non configurato: aggiungi gli utenti nei Secrets di Streamlit Cloud.")
        st.stop()

    logo_path = LOGO_JPG if LOGO_JPG.exists() else LOGO_PNG
    if logo_path.exists():
        render_logo(logo_path)
    st.subheader("Accedi")
    with st.form("login_form"):
        email = st.text_input("Email").strip().lower()
        password = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Accedi")

    if submitted:
        stored_hash = users.get(email, "")
        if stored_hash and verify_password(password, stored_hash):
            st.session_state["auth_logged_in"] = True
            st.session_state["auth_email"] = email
            st.session_state["auth_name"] = names.get(email, email) if names else email
            st.session_state["auth_error"] = ""
            st.rerun()
        else:
            st.session_state["auth_error"] = "Email o password non valide."

    if st.session_state["auth_error"]:
        st.error(st.session_state["auth_error"])
    st.stop()

def clean_text(v) -> str:
    if v is None:
        return ""
    return str(v).strip()

def billing_months_from_dates(start: date, end: date) -> int:
    if not isinstance(start, date) or not isinstance(end, date):
        return 1
    if end < start:
        start, end = end, start
    months = (end.year - start.year) * 12 + (end.month - start.month) + 1
    return max(1, months)

def billing_label_from_months(months) -> str:
    try:
        months = int(months)
    except Exception:
        months = 1
    return BILLING_PERIODS.get(months, f"{months} MESI")

def billing_divisor_from_months(months) -> float:
    try:
        months = int(months)
    except Exception:
        months = 1
    months = max(1, months)
    return 12.0 / float(months)

def bill_fixed_multiplier() -> int:
    try:
        return max(1, int(st.session_state.get("billing_months", 1)))
    except Exception:
        return 1

def bill_fixed_period_amount(value) -> float:
    return float(value) * float(bill_fixed_multiplier())

def month_key_from_date(d: date) -> str:
    return f"{d.year:04d}-{d.month:02d}"

def parse_index_month(x):
    if x is None:
        return ""
    if isinstance(x, datetime):
        return month_key_from_date(x.date())
    if isinstance(x, date):
        return month_key_from_date(x)
    s = str(x).strip()
    m = re.search(r"(\d{4})[-/](\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    parsed = parse_date_any(s)
    return month_key_from_date(parsed) if parsed else ""

def load_indici_rows(path: Path):
    if not path.exists():
        return []
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if "indici" not in wb.sheetnames:
            return []
        ws = wb["indici"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [clean_text(c).lower() for c in rows[0]]
        if not {"mese", "pun", "psv"}.issubset(set(header)):
            return []
        i_mese = header.index("mese")
        i_pun = header.index("pun")
        i_psv = header.index("psv")
        out = []
        for r in rows[1:]:
            if not r or not any(r):
                continue
            mese = parse_index_month(r[i_mese] if i_mese < len(r) else None)
            if not mese:
                continue
            out.append(
                {
                    "mese": mese,
                    "pun": parse_number(r[i_pun] if i_pun < len(r) else None),
                    "psv": parse_number(r[i_psv] if i_psv < len(r) else None),
                }
            )
        return sorted(out, key=lambda row: row["mese"])
    except Exception:
        return []

def select_indice_for_bill_period(rows, bill_start: date, bill_end: date):
    if not rows:
        return None, "missing"
    if bill_end < bill_start:
        bill_start, bill_end = bill_end, bill_start
    start_month = month_key_from_date(bill_start)
    end_month = month_key_from_date(bill_end)

    in_period = [r for r in rows if start_month <= r["mese"] <= end_month]
    if in_period:
        return in_period[-1], "period"

    before_end = [r for r in rows if r["mese"] <= end_month]
    if before_end:
        return before_end[-1], "before_end"

    return rows[-1], "latest_available"

def commodity_label(comm: str) -> str:
    return "GAS" if comm == "GAS" else "LUCE"

def safe_nome_cognome(nome_cliente: str) -> str:
    if not nome_cliente:
        return "senza_nome"
    s = re.sub(r'[\\/:*?"<>|]', "", nome_cliente.strip())
    parts = s.split()
    if len(parts) >= 2:
        return f"{parts[0]}_{parts[1]}"
    return parts[0]

def safe_download_filename(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", str(name).strip())
    return cleaned or "confronto_bolletta.xlsx"

def write_static_download(data: bytes, file_name: str):
    STATIC_DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)
    safe_name = safe_download_filename(file_name)
    stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    stored_name = f"{stamp}_{safe_name}"
    path = STATIC_DOWNLOADS_DIR / stored_name
    path.write_bytes(data)
    return f"app/static/downloads/{stored_name}", safe_name

def badge_ok():
    st.markdown(
        "<div style='background:#1f7a1f;color:white;padding:8px 12px;border-radius:10px;display:inline-block;font-weight:700;'>Dati completi ✅</div>",
        unsafe_allow_html=True,
    )

def badge_missing(msg: str):
    st.markdown(
        "<div style='background:#b3261e;color:white;padding:8px 12px;border-radius:10px;display:inline-block;font-weight:700;'>Dati mancanti ❌</div>"
        f"<div style='margin-top:6px;color:#b3261e;font-weight:600;'>{msg}</div>",
        unsafe_allow_html=True,
    )

# -----------------------------
# Tariffe: trova cartella più recente + carica template fisso
# -----------------------------
def find_latest_offer_folder(base: Path):
    if not base.exists():
        return None

    candidates = []

    # supporta: tariffe/2026-04
    for p in base.glob("*"):
        if p.is_dir() and re.match(r"^\d{4}-\d{2}$", p.name):
            candidates.append((p.name, p))

    # supporta: tariffe/2026/04
    for y in base.glob("*"):
        if y.is_dir() and re.match(r"^\d{4}$", y.name):
            for m in y.glob("*"):
                if m.is_dir() and re.match(r"^\d{2}$", m.name):
                    candidates.append((f"{y.name}-{m.name}", m))

    if not candidates:
        return None

    candidates.sort(key=lambda x: x[0])
    return candidates[-1][1]

def tariff_month_key(path: Path) -> str:
    try:
        parts = path.relative_to(TARIFFE_BASE).parts
    except ValueError:
        parts = path.parts

    for i, part in enumerate(parts):
        found = re.search(r"\d{4}-\d{2}", part)
        if found:
            return found.group(0)
        if re.match(r"^\d{4}-\d{2}$", part):
            return part
        if re.match(r"^\d{4}$", part) and i + 1 < len(parts) and re.match(r"^\d{2}$", parts[i + 1]):
            return f"{part}-{parts[i + 1]}"
    return ""

def normalized_supplier_code(value) -> str:
    text = clean_text(value).upper().replace(".", "").replace("-", "").replace(" ", "")
    if text in {"EON", "E"}:
        return "EON"
    if text == "ILLUMIA":
        return "ILLUMIA"
    return text

def supplier_matches(row, supplier_code: str) -> bool:
    provider = normalized_supplier_code(row.get("provider"))
    return not provider or provider == supplier_code

def segment_matches(row, segmento: str) -> bool:
    row_segment = clean_text(row.get("segmento")).upper()
    if not row_segment:
        return True
    if segmento == "BUSINESS":
        return row_segment in {"BUSINESS", "MICROBUSINESS"}
    return row_segment == segmento

def tariff_matches_segment(path: Path, segmento: str) -> bool:
    name = path.name.lower()
    parts = {p.lower() for p in path.parts}
    if path.suffix.lower() != ".xlsx" or name.startswith("~$"):
        return False
    if segmento == "BUSINESS":
        return "business" in parts or "business" in name
    return ("residenziale" in parts or "template" in name) and "business" not in name

def find_latest_eon_tariffe_file():
    if not EON_TARIFFE_BASE.exists():
        return None
    candidates = [
        p for p in EON_TARIFFE_BASE.glob("eon_tariffe_*.xlsx")
        if p.is_file() and not p.name.startswith("~$")
    ]
    if not candidates:
        return None
    candidates.sort(key=lambda p: (tariff_month_key(p), p.stat().st_mtime, p.name))
    return candidates[-1]

def load_tariffe_file_for_segment(segmento: str, supplier_code: str = None):
    supplier_code = supplier_code or selected_supplier_code()
    if supplier_code == "EON":
        selected = find_latest_eon_tariffe_file()
        if selected:
            st.session_state["offer_file_path"] = str(selected)
            return selected
        st.session_state["offer_file_path"] = ""
        return None

    if TARIFFE_BASE.exists():
        candidates = [
            p for p in TARIFFE_BASE.rglob("*.xlsx")
            if tariff_matches_segment(p, segmento)
        ]
        if candidates:
            candidates.sort(key=lambda p: (tariff_month_key(p), p.stat().st_mtime, p.name))
            selected = candidates[-1]
            st.session_state["offer_file_path"] = str(selected)
            return selected

    # fallback legacy SOLO residenziale
    if segmento == "RESIDENZIALE" and LEGACY_RES_FOUND:
        st.session_state["offer_file_path"] = str(LEGACY_RES_FOUND[0])
        return LEGACY_RES_FOUND[0]

    return None

def filter_tariffe_rows(rows, segmento: str, supplier_code: str):
    return [
        r for r in rows
        if supplier_matches(r, supplier_code) and segment_matches(r, segmento)
    ]

def get_file_valid_range(xlsx_path: Path):
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        if "tariffe" not in wb.sheetnames:
            return None, None
        ws = wb["tariffe"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None, None
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        if "valid_from" not in header or "valid_to" not in header:
            return None, None
        i_from = header.index("valid_from")
        i_to = header.index("valid_to")

        dfs, dts = [], []
        for r in rows[1:]:
            if not r:
                continue
            vf = parse_date_any(r[i_from] if i_from < len(r) else None)
            vt = parse_date_any(r[i_to] if i_to < len(r) else None)
            if vf:
                dfs.append(vf)
            if vt:
                dts.append(vt)
        if not dfs or not dts:
            return None, None
        return min(dfs), max(dts)
    except Exception:
        return None, None

def load_tariffe_from_path(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "tariffe" not in wb.sheetnames:
        return []
    ws = wb["tariffe"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    out = []
    for r in rows[1:]:
        if not any(r):
            continue
        d = {}
        for i, key in enumerate(header):
            d[key] = r[i] if i < len(r) else None
        d["provider"] = normalized_supplier_code(d.get("provider"))
        d["segmento"] = clean_text(d.get("segmento")).upper()
        d["commodity"] = clean_text(d.get("commodity")).upper()
        d["offer_type"] = clean_text(d.get("offer_type")).upper()
        d["offer_name"] = clean_text(d.get("offer_name"))
        d["component"] = clean_text(d.get("component")).lower()
        d["value_num"] = parse_number(d.get("value"))
        d["is_sellable"] = d.get("is_sellable")
        out.append(d)
    return out

def load_tariffe_from_bytes(xlsx_bytes: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    if "tariffe" not in wb.sheetnames:
        return []
    ws = wb["tariffe"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    out = []
    for r in rows[1:]:
        if not any(r):
            continue
        d = {}
        for i, key in enumerate(header):
            d[key] = r[i] if i < len(r) else None
        d["provider"] = normalized_supplier_code(d.get("provider"))
        d["segmento"] = clean_text(d.get("segmento")).upper()
        d["commodity"] = clean_text(d.get("commodity")).upper()
        d["offer_type"] = clean_text(d.get("offer_type")).upper()
        d["offer_name"] = clean_text(d.get("offer_name"))
        d["component"] = clean_text(d.get("component")).lower()
        d["value_num"] = parse_number(d.get("value"))
        d["is_sellable"] = d.get("is_sellable")
        out.append(d)
    return out


# -----------------------------
# Offerta scelta dall'app
# -----------------------------
def select_offer_name(rows, commodity, offer_type, segmento, supplier_code=None):
    c = commodity.upper()
    o = offer_type.upper()
    supplier_code = supplier_code or selected_supplier_code()
    subset = [
        r for r in rows
        if r.get("commodity") == c
        and r.get("offer_type") == o
        and segment_matches(r, segmento)
        and supplier_matches(r, supplier_code)
    ]

    if any(r.get("is_sellable") is not None for r in subset):
        subset = [r for r in subset if truthy(r.get("is_sellable")) is True]

    # BUSINESS: se manca is_sellable, per prudenza non vendiamo FISSA
    if supplier_code == "ILLUMIA" and segmento == "BUSINESS" and not any(r.get("is_sellable") is not None for r in rows):
        if o == "FISSA":
            return ""

    if not subset:
        return ""

    counts = {}
    for r in subset:
        name = clean_text(r.get("offer_name"))
        if name:
            counts[name] = counts.get(name, 0) + 1
    if not counts:
        return ""

    return sorted(counts.items(), key=lambda x: (-x[1], x[0]))[0][0]

def filter_rows_by_offer(rows, commodity, offer_type, offer_name, segmento=None, supplier_code=None):
    c = commodity.upper()
    o = offer_type.upper()
    segmento = segmento or st.session_state.get("segmento", "RESIDENZIALE")
    supplier_code = supplier_code or selected_supplier_code()
    return [
        r for r in rows
        if r.get("commodity") == c
        and r.get("offer_type") == o
        and clean_text(r.get("offer_name")) == offer_name
        and segment_matches(r, segmento)
        and supplier_matches(r, supplier_code)
    ]


# -----------------------------
# Calcolo fornitore
# -----------------------------
def comp_sum(rows, commodity, offer, component):
    c = commodity.upper()
    o = offer.upper()
    comp = component.lower()
    return float(sum(r.get("value_num", 0.0) for r in rows if r.get("commodity") == c and r.get("offer_type") == o and r.get("component") == comp))

def comp_first(rows, commodity, offer, component):
    c = commodity.upper()
    o = offer.upper()
    comp = component.lower()
    for r in rows:
        if r.get("commodity") == c and r.get("offer_type") == o and r.get("component") == comp:
            return float(r.get("value_num", 0.0))
    return 0.0

def supplier_discount(rows_offer, commodity, offer, fallback):
    tariff_discount = comp_sum(rows_offer, commodity, offer, "sconto_bonus")
    if tariff_discount != 0.0:
        return tariff_discount
    if selected_supplier_code() == "ILLUMIA":
        return float(fallback)
    return 0.0

def calc_supplier_vendite(rows_offer, commodity, offer, consumo, pun, psv, billing_divisor):
    comm = commodity.upper()
    off = offer.upper()

    def fixed_annua(c, o):
        base = comp_sum(rows_offer, c, o, "ccv_quota_fissa")
        disp = comp_sum(rows_offer, c, o, "dispbt") if st.session_state.get("include_dispbt", True) else 0.0
        return base + disp

    if comm == "EE":
        if off == "VARIABILE":
            fee = comp_sum(rows_offer, "EE", "VARIABILE", "fee_energia")
            ccv_var = comp_sum(rows_offer, "EE", "VARIABILE", "ccv_quota_variabile")
            sbil = comp_sum(rows_offer, "EE", "VARIABILE", "sbilanciamento")
            perdite = comp_sum(rows_offer, "EE", "VARIABILE", "perdite_rete")
            if perdite == 0.0:
                perdite = 0.10
            base = pun * (1.0 + perdite)
            prezzo = base + fee + ccv_var + sbil
            vend_cons = consumo * prezzo
            vend_fix = fixed_annua("EE", "VARIABILE") / float(billing_divisor)
            return float(vend_cons), float(vend_fix)

        prezzo = comp_first(rows_offer, "EE", "FISSA", "prezzo_mono")
        if prezzo == 0.0:
            p1 = comp_first(rows_offer, "EE", "FISSA", "prezzo_f1")
            p23 = comp_first(rows_offer, "EE", "FISSA", "prezzo_f23")
            prezzo = max(p1, p23)
        vend_cons = consumo * prezzo
        vend_fix = fixed_annua("EE", "FISSA") / float(billing_divisor)
        return float(vend_cons), float(vend_fix)

    if off == "VARIABILE":
        fee = comp_sum(rows_offer, "GAS", "VARIABILE", "fee_energia")
        ccv_var = comp_sum(rows_offer, "GAS", "VARIABILE", "ccv_quota_variabile")
        bil = comp_sum(rows_offer, "GAS", "VARIABILE", "bilanciamento")
        prezzo = psv + fee + ccv_var + bil
        vend_cons = consumo * prezzo
        vend_fix = fixed_annua("GAS", "VARIABILE") / float(billing_divisor)
        return float(vend_cons), float(vend_fix)

    prezzo = comp_first(rows_offer, "GAS", "FISSA", "prezzo_fisso")
    vend_cons = consumo * prezzo
    vend_fix = fixed_annua("GAS", "FISSA") / float(billing_divisor)
    return float(vend_cons), float(vend_fix)


# -----------------------------
# Export Excel (template + accise conformi al tuo esempio)
# -----------------------------
def find_row_map(ws):
    rm = {}
    for r in range(1, ws.max_row + 1):
        v = ws[f"A{r}"].value
        if not isinstance(v, str):
            continue
        t = norm(v)
        if t == "vendita consumo":
            rm["vendita_consumo"] = r
        elif t.startswith("vendita fissa") and "luce" in t:
            rm["vendita_fissa_luce"] = r
        elif t.startswith("vendita fissa") and "gas" in t:
            rm["vendita_fissa_gas"] = r
        elif "rete" in t and "consumi" in t:
            rm["rete_consumi"] = r
        elif "rete" in t and "fissa" in t:
            rm["rete_fissa"] = r
        elif "quota potenza" in t:
            rm["quota_potenza"] = r
        elif t == "sconti":
            rm["sconti"] = r
        elif "ricalc" in t:
            rm["ricalcoli"] = r
        elif "bonus social" in t:
            rm["bonus_sociale"] = r
        elif "arrotond" in t:
            rm["arrotondamenti"] = r
        elif "servizi" in t and "access" in t:
            rm["servizi_accessori"] = r
        elif "accise" in t:
            rm["accise_iva"] = r
        elif t == "totale":
            rm["totale"] = r

    if "arrotondamenti" not in rm and "bonus_sociale" not in rm and "accise_iva" in rm:
        candidate = rm["accise_iva"] - 1
        if candidate > 0:
            rm["arrotondamenti"] = candidate

    return rm

def copy_row_format(ws, source_row: int, target_row: int):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        source = ws.cell(row=source_row, column=col)
        target = ws.cell(row=target_row, column=col)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)

def ensure_export_rows(ws):
    rm = find_row_map(ws)
    missing_count = 0
    if "arrotondamenti" not in rm:
        missing_count += 1
    if "servizi_accessori" not in rm:
        missing_count += 1
    if missing_count == 0:
        return

    insert_at = rm.get("accise_iva", rm.get("totale", 13))
    ws.insert_rows(insert_at, amount=missing_count)
    source_row = insert_at + missing_count
    for row in range(insert_at, insert_at + missing_count):
        copy_row_format(ws, source_row, row)

def ensure_export_summary_rows(ws):
    if norm(ws["A7"].value) == "voce":
        return
    ws.insert_rows(3, amount=4)

def write_summary_pair(ws, label_cell, value_cell, label, value):
    ws[label_cell] = f"{label}:"
    ws[value_cell] = value
    if label == "Scadenza offerta":
        ws[label_cell].font = Font(bold=True, color="B3261E")
        ws[value_cell].font = Font(color="B3261E")
    else:
        ws[label_cell].font = Font(bold=True)

def write_comparison_summary(ws):
    left_fields, right_fields = comparison_summary_fields()
    for idx, (label, value) in enumerate(left_fields, start=1):
        write_summary_pair(ws, f"A{idx}", f"B{idx}", label, value)
    for idx, (label, value) in enumerate(right_fields, start=1):
        write_summary_pair(ws, f"E{idx}", f"F{idx}", label, value)

    for cell_ref in ("A7", "B7", "C7", "D7"):
        ws[cell_ref].font = Font(bold=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 38

def apply_export_labels(ws, nome_cliente: str):
    ensure_export_summary_rows(ws)
    ensure_export_rows(ws)
    labels = {
        6: None,
        7: "VOCE",
        8: "Vendita Consumo",
        9: "Rete e oneri di sistema Consumi",
        10: "Vendita Fissa Luce",
        11: "Vendita Fissa  Gas",
        12: "Rete e oneri di sistema Fissa",
        13: "Quota Potenza",
        14: "Sconti",
        15: "Ricalcoli/Partite pregresse",
        16: "Bonus Sociale",
        17: "Arrotondamenti",
        18: "Servizi accessori",
        19: "Accise e Iva",
        20: "Totale",
    }
    for row, value in labels.items():
        ws[f"A{row}"] = value
    ws["B7"] = "Bolletta"
    var_label, fix_label = supplier_column_labels()
    ws["C7"] = var_label
    ws["D7"] = fix_label

def write_export_metadata(ws):
    write_comparison_summary(ws)
    ws["H1"] = "Validità offerta"
    if st.session_state.get("offer_valid_from") and st.session_state.get("offer_valid_to"):
        ws["I1"] = f"{st.session_state['offer_valid_from']} → {st.session_state['offer_valid_to']}"
    else:
        ws["I1"] = "N.D. (upload manuale)"
    ws["H2"] = "File tariffe"
    ws["I2"] = st.session_state.get("offer_file_path", "")
    ws["H3"] = "Parametri Accise/IVA"
    ws["I3"] = format_fiscal_parameters()
    ws["H1"].font = Font(bold=True)
    ws["H2"].font = Font(bold=True)
    ws["H3"].font = Font(bold=True)

def validate_row_map(rm):
    required = [
        "vendita_consumo",
        "vendita_fissa_luce",
        "vendita_fissa_gas",
        "rete_consumi",
        "rete_fissa",
        "quota_potenza",
        "sconti",
        "ricalcoli",
        "bonus_sociale",
        "arrotondamenti",
        "servizi_accessori",
        "accise_iva",
        "totale",
    ]
    missing = [key for key in required if key not in rm]
    if missing:
        raise ValueError("Template Excel incompleto: mancano le righe " + ", ".join(missing))

def apply_total_formula(ws, rm, col_letter):
    acc = rm["accise_iva"]
    start = rm["vendita_consumo"]
    end = acc - 1
    ws[f"{col_letter}{rm['totale']}"] = f"=SUM({col_letter}{start}:{col_letter}{end})+{col_letter}{acc}"

def write_column(ws, rm, col, vals, commodity):
    ws[f"{col}{rm['vendita_consumo']}"] = float(vals["vendita_consumo"])
    if commodity == "GAS":
        ws[f"{col}{rm['vendita_fissa_gas']}"] = float(vals["vendita_fissa"])
        ws[f"{col}{rm['vendita_fissa_luce']}"] = "N.A."
        ws[f"{col}{rm['quota_potenza']}"] = "N.A."
    else:
        ws[f"{col}{rm['vendita_fissa_luce']}"] = float(vals["vendita_fissa"])
        ws[f"{col}{rm['vendita_fissa_gas']}"] = "N.A."
        ws[f"{col}{rm['quota_potenza']}"] = float(vals["quota_potenza"])

    ws[f"{col}{rm['rete_consumi']}"] = float(vals["rete_consumi"])
    ws[f"{col}{rm['rete_fissa']}"] = float(vals["rete_fissa"])
    ws[f"{col}{rm['sconti']}"] = float(vals["sconti"])
    ws[f"{col}{rm['ricalcoli']}"] = float(vals["ricalcoli"])
    if "bonus_sociale" in rm:
        ws[f"{col}{rm['bonus_sociale']}"] = float(vals.get("bonus_sociale", 0.0))
    if "arrotondamenti" in rm:
        ws[f"{col}{rm['arrotondamenti']}"] = float(vals["arrotondamenti"])
    if "servizi_accessori" in rm:
        ws[f"{col}{rm['servizi_accessori']}"] = float(vals.get("servizi_accessori", 0.0))

def fill_column_text(ws, rm, col, text):
    for key, r in rm.items():
        if key == "totale" and text == "N.A.":
            continue
        ws[f"{col}{r}"] = text

def fill_column_na(ws, rm, col):
    fill_column_text(ws, rm, col, "N.A.")

def fill_column_nd(ws, rm, col):
    fill_column_text(ws, rm, col, "N.D.")

def reset_illumia_results():
    st.session_state["illumia_calculated"] = False
    st.session_state["c_vendita_consumo"] = 0.0
    st.session_state["c_vendita_fissa"] = 0.0
    st.session_state["d_vendita_consumo"] = 0.0
    st.session_state["d_vendita_fissa"] = 0.0
    st.session_state["c_sconti"] = 0.0
    st.session_state["d_sconti"] = 0.0

def reset_bill_data():
    st.session_state["consumo"] = 0.0
    for key in KEYS:
        st.session_state[f"b_{key}"] = 0.0
    st.session_state["servizi_accessori_iva"] = "22%"
    st.session_state["bill_offer_type"] = "Fissa"
    st.session_state["excel_ready"] = False
    st.session_state["comparison_generated_at"] = ""
    reset_illumia_results()

def progressive_tax_amount(consumption, brackets):
    consumption = max(0.0, float(consumption))
    total = 0.0
    lower = 0.0
    for upper, rate in brackets:
        if upper is None:
            taxable = max(0.0, consumption - lower)
        else:
            taxable = max(0.0, min(consumption, upper) - lower)
        total += taxable * float(rate)
        if upper is None or consumption <= upper:
            break
        lower = float(upper)
    return total

def annual_progressive_tax_for_period(period_consumption, annual_consumption, brackets):
    period_consumption = max(0.0, float(period_consumption))
    annual_consumption = max(0.0, float(annual_consumption))
    if period_consumption <= 0 or annual_consumption <= 0:
        return 0.0
    annual_tax = progressive_tax_amount(annual_consumption, brackets)
    return annual_tax * (period_consumption / annual_consumption)

def fiscal_annual_consumption():
    return max(0.0, float(st.session_state.get("tax_annual_consumption", 0.0)))

def fiscal_primary_home():
    return str(st.session_state.get("tax_primary_home", "Sì")).strip().lower() in {"sì", "si", "yes", "true", "1"}

def fiscal_region_rate():
    region = st.session_state.get("tax_region", "Veneto")
    return float(GAS_REGIONAL_ADDITIONAL_MIN_RATES.get(region, 0.0))

def accessory_services_vat_label():
    label = st.session_state.get("servizi_accessori_iva", "22%")
    return label if label in ACCESSORY_SERVICES_VAT_RATES else "22%"

def accessory_services_vat_rate():
    return float(ACCESSORY_SERVICES_VAT_RATES[accessory_services_vat_label()])

def format_fiscal_parameters():
    unit = "Smc" if st.session_state["commodity"] == "GAS" else "kWh"
    return (
        f"Prima casa/residente: {st.session_state.get('tax_primary_home', 'Sì')} | "
        f"Potenza: {float(st.session_state.get('tax_power_kw', 0.0)):g} kW | "
        f"Consumo annuo: {fiscal_annual_consumption():g} {unit}/anno | "
        f"Regione: {st.session_state.get('tax_region', 'Veneto')} | "
        f"IVA servizi accessori: {accessory_services_vat_label()}"
    )

def calculate_accise_iva(vals, commodity):
    period_consumption = max(0.0, float(st.session_state.get("consumo", 0.0)))
    accessory_services = float(vals.get("servizi_accessori", 0.0))
    social_bonus = float(vals.get("bonus_sociale", 0.0))
    taxable_supply_subtotal = comparison_subtotal(vals, commodity) - accessory_services - social_bonus
    accessory_services_vat = accessory_services * accessory_services_vat_rate()

    if commodity == "EE":
        power_kw = max(0.0, float(st.session_state.get("tax_power_kw", 0.0)))
        residential = st.session_state.get("segmento") == "RESIDENZIALE"
        primary_home = fiscal_primary_home()
        exempt_kwh = EE_RESIDENTIAL_EXEMPT_KWH_PER_MONTH * float(st.session_state.get("billing_months", 1))
        if residential and primary_home and power_kw <= 3.0:
            taxable_kwh = max(0.0, period_consumption - exempt_kwh)
        else:
            taxable_kwh = period_consumption
        excise = taxable_kwh * EE_EXCISE_RATE
        vat_rate = 0.10 if residential and primary_home else 0.22
        vat = ((taxable_supply_subtotal + excise) * vat_rate) + accessory_services_vat
        return excise + vat

    annual_consumption = fiscal_annual_consumption()
    excise = annual_progressive_tax_for_period(period_consumption, annual_consumption, GAS_EXCISE_BRACKETS)
    regional = period_consumption * fiscal_region_rate()

    variable_base = (
        float(vals["vendita_consumo"])
        + float(vals["rete_consumi"])
        + float(vals["sconti"])
        + excise
        + regional
    )
    fixed_base = (
        float(vals["vendita_fissa"])
        + float(vals["rete_fissa"])
        + float(vals["ricalcoli"])
        + float(vals["arrotondamenti"])
    )
    reduced_ratio = min(annual_consumption, GAS_VAT_REDUCED_ANNUAL_THRESHOLD) / annual_consumption if annual_consumption else 0.0
    reduced_ratio = min(max(reduced_ratio, 0.0), 1.0)
    vat = (
        (variable_base * reduced_ratio * 0.10)
        + (variable_base * (1.0 - reduced_ratio) * 0.22)
        + (fixed_base * 0.22)
        + accessory_services_vat
    )
    return excise + regional + vat

def comparison_subtotal(vals, commodity):
    subtotal = (
        float(vals["vendita_consumo"])
        + float(vals["rete_consumi"])
        + float(vals["vendita_fissa"])
        + float(vals["rete_fissa"])
        + float(vals["sconti"])
        + float(vals["ricalcoli"])
        + float(vals.get("bonus_sociale", 0.0))
        + float(vals["arrotondamenti"])
        + float(vals.get("servizi_accessori", 0.0))
    )
    if commodity == "EE":
        subtotal += float(vals["quota_potenza"])
    return subtotal

def comparison_total(vals, commodity):
    return comparison_subtotal(vals, commodity) + float(vals["accise_iva"])

def build_comparison_values():
    comm = st.session_state["commodity"]

    b_vals = {k: float(st.session_state[f"b_{k}"]) for k in KEYS}
    b_vals["bonus_sociale"] = -abs(float(b_vals.get("bonus_sociale", 0.0)))
    b_vals["vendita_fissa"] = bill_fixed_period_amount(b_vals["vendita_fissa"])
    b_vals["rete_fissa"] = bill_fixed_period_amount(b_vals["rete_fissa"])

    c_vals = b_vals.copy()
    d_vals = b_vals.copy()

    c_vals["vendita_consumo"] = float(st.session_state["c_vendita_consumo"])
    c_vals["vendita_fissa"] = float(st.session_state["c_vendita_fissa"])
    c_vals["sconti"] = float(st.session_state.get("c_sconti", st.session_state["ill_sconto_var"]))
    c_vals["ricalcoli"] = 0.0
    c_vals["arrotondamenti"] = 0.0

    d_vals["vendita_consumo"] = float(st.session_state["d_vendita_consumo"])
    d_vals["vendita_fissa"] = float(st.session_state["d_vendita_fissa"])
    d_vals["sconti"] = float(st.session_state.get("d_sconti", st.session_state["ill_sconto_fix"]))
    d_vals["ricalcoli"] = 0.0
    d_vals["arrotondamenti"] = 0.0

    c_vals["accise_iva"] = calculate_accise_iva(c_vals, comm)
    d_vals["accise_iva"] = calculate_accise_iva(d_vals, comm)

    return {
        "commodity": comm,
        "bolletta": b_vals,
        "variabile": c_vals,
        "fissa": d_vals,
        "has_offer_var": bool(st.session_state.get("offer_var_name")),
        "has_offer_fix": bool(st.session_state.get("offer_fix_name")),
    }

def comparison_value(vals, key, commodity):
    if key == "vendita_fissa_luce":
        return float(vals["vendita_fissa"]) if commodity == "EE" else "N.A."
    if key == "vendita_fissa_gas":
        return float(vals["vendita_fissa"]) if commodity == "GAS" else "N.A."
    if key == "quota_potenza":
        return float(vals["quota_potenza"]) if commodity == "EE" else "N.A."
    if key == "totale":
        return comparison_total(vals, commodity)
    return float(vals.get(key, 0.0))

def format_eur(value):
    if isinstance(value, str):
        return value
    amount = float(value)
    sign = "-" if amount < 0 else ""
    formatted = f"{abs(amount):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{sign}€ {formatted}"

def format_index_value(value):
    try:
        return f"{float(value):.4f}".replace(".", ",")
    except Exception:
        return "0,0000"

def render_logo(path):
    suffix = path.suffix.lower()
    mime = "image/png" if suffix == ".png" else "image/jpeg"
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    st.markdown(
        f"""
        <div style="display:flex;justify-content:center;margin:0.25rem 0 1.5rem;">
            <img src="data:{mime};base64,{encoded}" style="width:300px;max-width:50%;height:auto;" alt="Energia Solidale">
        </div>
        """,
        unsafe_allow_html=True,
    )

def render_month_year_parts(title, prefix, disabled=False):
    st.markdown(f"**{title}**")
    m_col, y_col = st.columns([1, 1.2])
    with m_col:
        st.number_input("Mese", min_value=1, max_value=12, step=1, key=f"{prefix}_month", disabled=disabled)
    with y_col:
        st.number_input("Anno", min_value=2000, max_value=2100, step=1, key=f"{prefix}_year", disabled=disabled)

def query_value(name):
    value = st.query_params.get(name, "")
    if isinstance(value, list):
        value = value[0] if value else ""
    return str(value).strip()

def query_float(name, default):
    raw = query_value(name)
    return parse_number(raw) if raw else float(default)

def build_comparison_table_rows(values=None):
    values = values or build_comparison_values()
    comm = values["commodity"]
    var_label, fix_label = supplier_column_labels()
    rows_config = [
        ("vendita_consumo", "Vendita Consumo"),
        ("rete_consumi", "Rete e oneri di sistema Consumi"),
        ("vendita_fissa_luce", "Vendita Fissa Luce"),
        ("vendita_fissa_gas", "Vendita Fissa  Gas"),
        ("rete_fissa", "Rete e oneri di sistema Fissa"),
        ("quota_potenza", "Quota Potenza"),
        ("sconti", "Sconti"),
        ("ricalcoli", "Ricalcoli/Partite pregresse"),
        ("bonus_sociale", "Bonus Sociale"),
        ("arrotondamenti", "Arrotondamenti"),
        ("servizi_accessori", "Servizi accessori"),
        ("accise_iva", "Accise e Iva"),
        ("totale", "Totale"),
    ]
    rows = []
    for key, label in rows_config:
        var_value = (
            comparison_value(values["variabile"], key, comm)
            if values["has_offer_var"]
            else "N.D."
        )
        fix_value = (
            comparison_value(values["fissa"], key, comm)
            if values["has_offer_fix"]
            else "N.D."
        )
        rows.append(
            {
                "VOCE": label,
                "Bolletta": format_eur(comparison_value(values["bolletta"], key, comm)),
                var_label: format_eur(var_value),
                fix_label: format_eur(fix_value),
            }
        )
    return rows

def build_markdown_table(rows):
    var_label, fix_label = supplier_column_labels()
    lines = [f"| VOCE | Bolletta | {var_label} | {fix_label} |", "|---|---:|---:|---:|"]
    for row in rows:
        lines.append(
            "| "
            + str(row["VOCE"]).replace("|", "/")
            + " | "
            + str(row["Bolletta"]).replace("|", "/")
            + " | "
            + str(row[var_label]).replace("|", "/")
            + " | "
            + str(row[fix_label]).replace("|", "/")
            + " |"
        )
    return "\n".join(lines)

def html_escape(value):
    return html.escape(str(value), quote=True)

def markdown_escape(value):
    text = str(value)
    text = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    for char in ("\\", "*", "_", "`", "[", "]", "(", ")", "#", "+", "-", ".", "!"):
        text = text.replace(char, f"\\{char}")
    return text

def render_comparison_summary():
    left_fields, right_fields = comparison_summary_fields()

    def render_lines(fields):
        lines = []
        for label, value in fields:
            line = f"**{label}:** {markdown_escape(value)}"
            if label == "Scadenza offerta":
                line = f":red[{line}]"
            lines.append(line)
        return "  \n".join(lines)

    left_col, right_col = st.columns(2, gap="large")
    left_col.markdown(render_lines(left_fields))
    right_col.markdown(render_lines(right_fields))

def render_comparison_rows(rows):
    var_label, fix_label = supplier_column_labels()

    table_rows = []
    for row in rows:
        is_total = str(row["VOCE"]).strip().lower() == "totale"
        tr_class = ' class="es-total-row"' if is_total else ""
        table_rows.append(
            f"<tr{tr_class}>"
            f"<td>{html_escape(row['VOCE'])}</td>"
            f"<td>{html_escape(row['Bolletta'])}</td>"
            f"<td>{html_escape(row[var_label])}</td>"
            f"<td>{html_escape(row[fix_label])}</td>"
            "</tr>"
        )

    st.markdown(
        f"""
        <style>
        .es-comparison-table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 0.25rem;
            font-size: 1.02rem;
            color: #1f2937;
        }}
        .es-comparison-table th {{
            background: #eef5fb;
            padding: 0.55rem 0.7rem;
            font-weight: 800;
            text-align: right;
            border-bottom: 1px solid #dce4ea;
            white-space: nowrap;
        }}
        .es-comparison-table th:first-child {{
            text-align: left;
        }}
        .es-comparison-table td {{
            padding: 0.48rem 0.7rem;
            border-bottom: 1px solid #e5e7eb;
            text-align: right;
            white-space: nowrap;
        }}
        .es-comparison-table td:first-child {{
            text-align: left;
            white-space: normal;
        }}
        .es-total-row td {{
            font-weight: 800;
        }}
        </style>
        <table class="es-comparison-table">
            <thead>
                <tr>
                    <th>VOCE</th>
                    <th>Bolletta</th>
                    <th>{html_escape(var_label)}</th>
                    <th>{html_escape(fix_label)}</th>
                </tr>
            </thead>
            <tbody>
                {"".join(table_rows)}
            </tbody>
        </table>
        """,
        unsafe_allow_html=True,
    )

def render_blob_download_button(data: bytes, file_name: str):
    encoded = base64.b64encode(data).decode("ascii")
    safe_name = safe_download_filename(file_name)
    html_doc = f"""
    <button id="download-xlsx" style="
        width:100%;
        padding:0.65rem 1rem;
        border:0;
        border-radius:0.5rem;
        background:#1597d3;
        color:white;
        font-weight:700;
        cursor:pointer;
        font-family:Arial, sans-serif;
        font-size:16px;
    ">Scarica Excel</button>
    <script>
    const data = {json.dumps(encoded)};
    const fileName = {json.dumps(safe_name)};
    function base64ToBytes(base64) {{
        const binary = atob(base64);
        const bytes = new Uint8Array(binary.length);
        for (let i = 0; i < binary.length; i++) {{
            bytes[i] = binary.charCodeAt(i);
        }}
        return bytes;
    }}
    document.getElementById("download-xlsx").addEventListener("click", function () {{
        const bytes = base64ToBytes(data);
        const blob = new Blob([bytes], {{
            type: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }});
        const url = URL.createObjectURL(blob);
        const link = document.createElement("a");
        link.href = url;
        link.download = fileName;
        document.body.appendChild(link);
        link.click();
        link.remove();
        setTimeout(() => URL.revokeObjectURL(url), 1000);
    }});
    </script>
    """
    components.html(html_doc, height=56)

def bolletta_is_valid():
    v1 = float(st.session_state["b_vendita_consumo"])
    v2 = float(st.session_state["b_vendita_fissa"])
    v3 = float(st.session_state["b_rete_consumi"])
    v4 = float(st.session_state["b_rete_fissa"])
    return not (v1 == 0.0 and v2 == 0.0 and v3 == 0.0 and v4 == 0.0)

require_authentication()

# -----------------------------
# UI
# -----------------------------
logo_path = LOGO_JPG if LOGO_JPG.exists() else LOGO_PNG
if logo_path.exists():
    render_logo(logo_path)

left, center, right = st.columns([0.15, 4.7, 0.15])
with center:
    st.title(f"Confronto bollette vs {selected_supplier_label()}")

    top_name, top_segment, top_supply, top_supplier = st.columns([2, 1, 1, 1])
    with top_name:
        cliente_param = query_value("cliente")
        if cliente_param:
            st.session_state["nome_cliente"] = cliente_param
            st.session_state["nome_cliente_input"] = cliente_param
        elif not st.session_state["nome_cliente"].strip():
            st.session_state["nome_cliente"] = "Cliente"
            st.session_state["nome_cliente_input"] = "Cliente"
        st.caption("Cliente")
        st.text_input("Nome e Cognome", key="nome_cliente_input", disabled=st.session_state["offers_loaded"])
        st.session_state["nome_cliente"] = st.session_state["nome_cliente_input"].strip() or "Cliente"

    with top_segment:
        st.selectbox("Segmento", SEGMENT_OPTIONS, key="segmento_select", disabled=st.session_state["offers_loaded"])
        st.session_state["segmento"] = SEGMENT_LABEL_TO_VALUE[st.session_state["segmento_select"]]

    with top_supply:
        st.selectbox("Fornitura", SUPPLY_OPTIONS, key="commodity_select", disabled=st.session_state["offers_loaded"])
        st.session_state["commodity"] = SUPPLY_LABEL_TO_VALUE[st.session_state["commodity_select"]]
        if st.session_state["commodity"] == "GAS":
            st.session_state["tax_power_kw"] = 0.0

    with top_supplier:
        st.selectbox("Fornitore confronto", SUPPLIER_OPTIONS, key="comparison_supplier_select")
        st.session_state["comparison_supplier"] = SUPPLIER_LABEL_TO_CODE[st.session_state["comparison_supplier_select"]]

    if (
        st.session_state["segmento"] != st.session_state["prev_segmento"]
        or st.session_state["commodity"] != st.session_state["prev_commodity"]
        or st.session_state["comparison_supplier"] != st.session_state["prev_comparison_supplier"]
    ):
        reset_illumia_results()
        st.session_state["offers_loaded"] = False
        st.session_state["prev_segmento"] = st.session_state["segmento"]
        st.session_state["prev_commodity"] = st.session_state["commodity"]
        st.session_state["prev_comparison_supplier"] = st.session_state["comparison_supplier"]

    st.caption(
        f"Selezione attiva: {st.session_state['segmento_select']} | "
        f"{st.session_state['commodity_select']} | "
        f"Confronto: {selected_supplier_label()}"
    )
    if st.session_state["offers_loaded"]:
        st.info("Tariffe e indici caricati. Puoi modificare i dati bolletta: il confronto si aggiorna automaticamente.")
    print(
        "APP_RERUN "
        f"segmento={st.session_state['segmento']} "
        f"commodity={st.session_state['commodity']} "
        f"supplier={st.session_state['comparison_supplier']}",
        flush=True,
    )
    nome_ok = bool(st.session_state["nome_cliente"].strip())

    st.divider()

    # 1) Periodicità
    st.header("1️⃣ Periodicità")
    p1, p2 = st.columns(2)
    with p1:
        render_month_year_parts("Dal (bolletta)", "bill_start", disabled=st.session_state["offers_loaded"])
    with p2:
        render_month_year_parts("Al (bolletta)", "bill_end", disabled=st.session_state["offers_loaded"])

    st.session_state["bill_start"] = date_from_month_year("bill_start")
    st.session_state["bill_end"] = date_from_month_year("bill_end", end_of_month=True)

    billing_months = billing_months_from_dates(st.session_state["bill_start"], st.session_state["bill_end"])
    st.session_state["billing_months"] = billing_months
    st.session_state["billing_divisor"] = billing_divisor_from_months(billing_months)

    f1, f2 = st.columns([2, 3])
    with f1:
        st.markdown(f"**Periodicità:** {billing_label_from_months(billing_months)}")
        st.caption(
            f"Periodo: {st.session_state['bill_start'].strftime('%m/%Y')} - "
            f"{st.session_state['bill_end'].strftime('%m/%Y')}"
        )
    with f2:
        st.session_state["assume_fixed_is_monthly"] = True
        st.caption(
            "Quote fisse inserite come importo mensile: vendita fissa luce/gas "
            "e rete/oneri fissa sono sempre moltiplicate per i mesi fatturati."
        )

    st.caption(
        f"Mesi fatturati = {billing_months} | "
        f"Divisore periodo = {float(st.session_state['billing_divisor']):g} | "
        f"Moltiplicatore quote fisse bolletta = {bill_fixed_multiplier()}"
    )
    print("APP_SECTION periodicita_ok", flush=True)
    st.divider()

    # Dati fiscali per Accise + IVA
    st.subheader("Dati fiscali Accise/IVA")
    tax_unit = "Smc" if st.session_state["commodity"] == "GAS" else "kWh"
    f_tax_1, f_tax_2 = st.columns(2)
    with f_tax_1:
        st.selectbox("Prima casa / residente", ("Sì", "No"), key="tax_primary_home")
        st.number_input(
            "Potenza impegnata (kW)",
            min_value=0.0,
            step=0.1,
            key="tax_power_kw",
            disabled=st.session_state["commodity"] == "GAS",
        )
    with f_tax_2:
        st.number_input(
            f"Consumo annuo stimato ({tax_unit}/anno)",
            min_value=0.0,
            step=1.0,
            key="tax_annual_consumption",
        )
        st.selectbox("Regione", REGION_OPTIONS, key="tax_region")
    st.caption("Per l'addizionale regionale gas viene usato il valore minimo della tabella disponibile.")
    print("APP_SECTION fiscal_inputs_ok", flush=True)
    st.divider()

    # 2) Bolletta
    st.subheader("2️⃣ Bolletta (scontrino dell'energia)")

    unit = "Smc" if st.session_state["commodity"] == "GAS" else "kWh"

    if st.session_state["offers_loaded"]:
        fixed_multiplier = bill_fixed_multiplier()
        vendita_fissa_periodo = bill_fixed_period_amount(st.session_state["b_vendita_fissa"])
        rete_fissa_periodo = bill_fixed_period_amount(st.session_state["b_rete_fissa"])
        consumo_fmt = f"{float(st.session_state['consumo']):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        st.text(
            "\n".join(
                [
                    f"Consumo ({unit}): {consumo_fmt}",
                    f"Vendita consumo: {format_eur(st.session_state['b_vendita_consumo'])}",
                    f"Rete/oneri consumi: {format_eur(st.session_state['b_rete_consumi'])}",
                    f"Vendita fissa totale periodo ({fixed_multiplier} mesi): {format_eur(vendita_fissa_periodo)}",
                    f"Rete/oneri fissa totale periodo ({fixed_multiplier} mesi): {format_eur(rete_fissa_periodo)}",
                    f"Quota potenza: {format_eur(st.session_state['b_quota_potenza'])}",
                    f"Sconti: {format_eur(st.session_state['b_sconti'])}",
                    f"Ricalcoli: {format_eur(st.session_state['b_ricalcoli'])}",
                    f"Bonus Sociale: {format_eur(-abs(float(st.session_state['b_bonus_sociale'])))}",
                    f"Arrotondamenti: {format_eur(st.session_state['b_arrotondamenti'])}",
                    f"Servizi accessori: {format_eur(st.session_state['b_servizi_accessori'])} | IVA: {accessory_services_vat_label()}",
                    f"Accise + IVA: {format_eur(st.session_state['b_accise_iva'])}",
                ]
            )
        )

    st.selectbox("Tipo tariffa bolletta", BILL_OFFER_TYPE_OPTIONS, key="bill_offer_type")

    reset_col, _ = st.columns([1, 2])
    with reset_col:
        if st.button("Reset dati bolletta", use_container_width=True):
            reset_bill_data()
            st.rerun()

    b1, b2 = st.columns(2)
    with b1:
        st.number_input(f"Consumo ({unit})", min_value=0.0, step=0.01, key="consumo")
        st.number_input("Vendita consumo", step=0.01, key="b_vendita_consumo")
        st.number_input("Rete/oneri consumi", step=0.01, key="b_rete_consumi")
        st.number_input("Vendita fissa mensile (scontrino)", step=0.01, key="b_vendita_fissa")
        st.number_input("Rete/oneri fissa mensile", step=0.01, key="b_rete_fissa")
    with b2:
        st.number_input("Quota potenza (solo luce)", step=0.01, key="b_quota_potenza", disabled=st.session_state["commodity"] == "GAS")
        st.number_input("Sconti", step=0.01, key="b_sconti")
        st.number_input("Ricalcoli", step=0.01, key="b_ricalcoli")
        st.number_input("Bonus Sociale (negativo)", step=0.01, key="b_bonus_sociale")
        st.number_input("Arrotondamenti", step=0.01, key="b_arrotondamenti")
        st.number_input("Servizi accessori (imponibile)", step=0.01, key="b_servizi_accessori")
        st.selectbox(
            "IVA servizi accessori",
            ACCESSORY_SERVICES_VAT_OPTIONS,
            key="servizi_accessori_iva",
        )
        st.number_input("Accise + IVA", step=0.01, key="b_accise_iva")

    bol_ok = bolletta_is_valid()
    if not bol_ok:
        st.warning("⚠️ Bolletta incompleta: inserisci almeno una voce tra Vendita/Rete.")
    print("APP_SECTION bolletta_ok", flush=True)
    st.divider()

    # 3) Offerta più recente (validità bloccata) + indici + selezione offerta + calcolo
    supplier_label = selected_supplier_label()
    supplier_code = selected_supplier_code()
    st.header(f"3️⃣ Offerta {supplier_label} (più recente) + Indici + Offerta automatica + Calcolo")
    if not st.session_state["offers_loaded"]:
        st.info("Carica tariffe e indici quando hai completato i dati della bolletta.")
        if st.button("Carica tariffe e indici", use_container_width=True):
            st.session_state["offers_loaded"] = True
            st.session_state["excel_ready"] = False
            st.rerun()
        print("APP_SECTION waiting_offer_load", flush=True)
        st.stop()

    print("APP_SECTION loading_offers_start", flush=True)

    tariffe_rows = []
    offer_vf, offer_vt = None, None

    try:
        print("APP_SECTION select_offer_file_start", flush=True)
        offer_file = load_tariffe_file_for_segment(st.session_state["segmento"], supplier_code)
        print(f"APP_SECTION select_offer_file_done file={offer_file}", flush=True)
        if offer_file is None:
            st.error("❌ Nessun file tariffe trovato (né tariffe/ né legacy).")
        else:
            print("APP_SECTION load_tariffe_start", flush=True)
            tariffe_rows = load_tariffe_from_path(offer_file)
            tariffe_rows = filter_tariffe_rows(tariffe_rows, st.session_state["segmento"], supplier_code)
            print(f"APP_SECTION load_tariffe_done rows={len(tariffe_rows)}", flush=True)
            print("APP_SECTION valid_range_start", flush=True)
            offer_vf, offer_vt = get_file_valid_range(offer_file)
            print(f"APP_SECTION valid_range_done from={offer_vf} to={offer_vt}", flush=True)
            st.session_state["offer_valid_from"] = offer_vf
            st.session_state["offer_valid_to"] = offer_vt
    except Exception as exc:
        print(f"APP_ERROR tariffe {type(exc).__name__}: {exc}", flush=True)
        st.error(f"Errore nel caricamento tariffe: {exc}")
        tariffe_rows = []
        offer_vf, offer_vt = None, None

    # MOSTRA VALIDITÀ OFFERTA SOLO COME TESTO (NON MODIFICABILE)
    if offer_vf and offer_vt:
        st.markdown(
            f"**Validità offerta {supplier_label} (da file più recente):**  \n"
            f"Decorrenza: {offer_vf.strftime('%d/%m/%Y')}  \n"
            f"<span style='color:#b3261e;font-weight:700;'>"
            f"Scadenza offerta: {offer_vt.strftime('%d/%m/%Y')}"
            f"</span>",
            unsafe_allow_html=True,
        )
        st.caption(f"Fonte: {st.session_state.get('offer_file_path','')}")
        # warning se bolletta fuori validità, ma confronto comunque con offerta più recente
        if not (offer_vf <= st.session_state["bill_start"] <= offer_vt and offer_vf <= st.session_state["bill_end"] <= offer_vt):
            st.warning("⚠️ Il periodo bolletta NON rientra nella validità dell’ultima offerta. Il confronto usa comunque l’offerta più recente.")
    else:
        st.info("Validità offerta: N.D. (override upload)")

    # Indici PUN/PSV automatici dal file indici, sul mese più recente del periodo bolletta.
    try:
        print("APP_SECTION load_indici_start", flush=True)
        indici_rows = load_indici_rows(INDICI_XLSX)
        print(f"APP_SECTION load_indici_done rows={len(indici_rows)}", flush=True)
        selected_indice, indice_source = select_indice_for_bill_period(
            indici_rows,
            st.session_state["bill_start"],
            st.session_state["bill_end"],
        )
        print(f"APP_SECTION select_indice_done month={selected_indice['mese'] if selected_indice else None}", flush=True)
    except Exception as exc:
        print(f"APP_ERROR indici {type(exc).__name__}: {exc}", flush=True)
        st.error(f"Errore nel caricamento indici PUN/PSV: {exc}")
        indici_rows = []
        selected_indice, indice_source = None, "missing"

    if selected_indice:
        st.session_state["pun_override"] = float(selected_indice["pun"])
        st.session_state["psv_override"] = float(selected_indice["psv"])
        st.session_state["indice_month"] = selected_indice["mese"]
        st.session_state["indice_source"] = indice_source
        st.session_state["indice_file_path"] = str(INDICI_XLSX)

        st.markdown(
            f"**Indice PUN/PSV usato:** {selected_indice['mese']}  \n"
            f"Fonte: `{INDICI_XLSX.name}`"
        )
        if indice_source == "before_end":
            st.warning(
                "⚠️ Nessun indice disponibile dentro il periodo bolletta: uso il mese più recente disponibile "
                "non successivo alla fine periodo."
            )
        elif indice_source == "latest_available":
            st.warning("⚠️ Il periodo bolletta è precedente agli indici disponibili: uso l’ultimo mese presente nel file.")
    else:
        st.session_state["pun_override"] = 0.0
        st.session_state["psv_override"] = 0.0
        st.session_state["indice_month"] = ""
        st.session_state["indice_source"] = "missing"
        st.session_state["indice_file_path"] = str(INDICI_XLSX)
        st.error("❌ Nessun indice PUN/PSV trovato nel file indici_pun_psv_2025_2026.xlsx.")

    if st.session_state["commodity"] == "EE":
        st.caption("EE variabile: perdite rete da tariffa; se assenti, 10% su PUN applicate automaticamente")
        st.markdown(f"**PUN (€/kWh) da file indici:** {format_index_value(st.session_state['pun_override'])}")
        indice_ok = float(st.session_state["pun_override"]) > 0
    else:
        st.markdown(f"**PSV (€/Smc) da file indici:** {format_index_value(st.session_state['psv_override'])}")
        indice_ok = float(st.session_state["psv_override"]) > 0

    st.session_state["ill_sconto_var"] = query_float("sconto_var", st.session_state["ill_sconto_var"])
    st.session_state["ill_sconto_fix"] = query_float("sconto_fissa", st.session_state["ill_sconto_fix"])
    s1, s2 = st.columns(2)
    if supplier_code == "ILLUMIA":
        with s1:
            st.markdown(f"**Sconto {supplier_label} Variabile:** {format_eur(st.session_state['ill_sconto_var'])}")
        with s2:
            st.markdown(f"**Sconto {supplier_label} Fissa:** {format_eur(st.session_state['ill_sconto_fix'])}")
    else:
        with s1:
            st.markdown(f"**Sconto {supplier_label} Variabile:** da file tariffe, se presente")
        with s2:
            st.markdown(f"**Sconto {supplier_label} Fissa:** da file tariffe, se presente")
    print("APP_SECTION sconti_ok", flush=True)

    # Offerta automatica dall’app
    offer_var = select_offer_name(tariffe_rows, st.session_state["commodity"], "VARIABILE", st.session_state["segmento"], supplier_code) if tariffe_rows else ""
    offer_fix = select_offer_name(tariffe_rows, st.session_state["commodity"], "FISSA", st.session_state["segmento"], supplier_code) if tariffe_rows else ""

    st.session_state["offer_var_name"] = offer_var
    st.session_state["offer_fix_name"] = offer_fix

    if offer_var:
        st.success(f"✅ Offerta VARIABILE {supplier_label} selezionata dall’app: {offer_var}")
    else:
        st.info("ℹ️ Offerta VARIABILE non selezionata (assente/non vendibile).")

    if offer_fix:
        st.success(f"✅ Offerta FISSA {supplier_label} selezionata dall’app: {offer_fix}")
    else:
        st.info("ℹ️ Offerta FISSA non selezionata (assente/non vendibile).")

    print("APP_SECTION offers_loaded_ok", flush=True)

    missing = []
    if not nome_ok:
        missing.append("Nome cliente")
    if not bol_ok:
        missing.append("Bolletta completa")
    if not indice_ok:
        missing.append("PUN/PSV")
    if not tariffe_rows:
        missing.append("Tariffe")
    if not offer_var and not offer_fix:
        missing.append(f"Offerta {supplier_label}")
    if st.session_state["commodity"] == "GAS" and fiscal_annual_consumption() <= 0:
        missing.append("Consumo annuo stimato")
    if st.session_state["commodity"] == "EE" and float(st.session_state.get("tax_power_kw", 0.0)) <= 0:
        missing.append("Potenza impegnata")

    if missing:
        badge_missing("Mancano: " + ", ".join(missing))
        can_calc = False
    else:
        badge_ok()
        can_calc = True

    if can_calc:
        print("APP_SECTION calc_supplier_start", flush=True)
        comm = st.session_state["commodity"]
        consumo = float(st.session_state["consumo"])
        div = float(st.session_state["billing_divisor"])
        pun = float(st.session_state["pun_override"])
        psv = float(st.session_state["psv_override"])

        if offer_var:
            rows_var = filter_rows_by_offer(tariffe_rows, comm, "VARIABILE", offer_var, st.session_state["segmento"], supplier_code)
            v_cons, v_fix = calc_supplier_vendite(rows_var, comm, "VARIABILE", consumo, pun, psv, div)
            v_sconto = supplier_discount(rows_var, comm, "VARIABILE", st.session_state["ill_sconto_var"])
        else:
            v_cons, v_fix = 0.0, 0.0
            v_sconto = 0.0

        if offer_fix:
            rows_fix = filter_rows_by_offer(tariffe_rows, comm, "FISSA", offer_fix, st.session_state["segmento"], supplier_code)
            f_cons, f_fix = calc_supplier_vendite(rows_fix, comm, "FISSA", consumo, pun, psv, div)
            f_sconto = supplier_discount(rows_fix, comm, "FISSA", st.session_state["ill_sconto_fix"])
        else:
            f_cons, f_fix = 0.0, 0.0
            f_sconto = 0.0

        st.session_state["c_vendita_consumo"] = v_cons
        st.session_state["c_vendita_fissa"] = v_fix
        st.session_state["d_vendita_consumo"] = f_cons
        st.session_state["d_vendita_fissa"] = f_fix
        st.session_state["c_sconti"] = v_sconto
        st.session_state["d_sconti"] = f_sconto
        st.session_state["illumia_calculated"] = True
        st.session_state["comparison_generated_at"] = current_comparison_timestamp()
        print("APP_SECTION calc_supplier_done", flush=True)
        st.success(f"✅ Calcolo {supplier_label} aggiornato automaticamente.")
    else:
        st.session_state["illumia_calculated"] = False

    st.divider()

    # 4) Confronto in dashboard
    st.header("4️⃣ Confronto")
    if st.session_state["illumia_calculated"]:
        print("APP_SECTION comparison_render_start", flush=True)
        try:
            comparison_values = build_comparison_values()
            print("APP_SECTION comparison_values_ok", flush=True)
            comparison_rows = build_comparison_table_rows(comparison_values)
            print(f"APP_SECTION comparison_rows_ok rows={len(comparison_rows)}", flush=True)
            render_comparison_summary()
            render_comparison_rows(comparison_rows)
            st.caption(f"Parametri Accise/IVA: {format_fiscal_parameters()}")
            print("APP_SECTION comparison_render_done", flush=True)
        except Exception as exc:
            print(f"APP_ERROR comparison_render {type(exc).__name__}: {exc}", flush=True)
            st.error(f"Errore nella visualizzazione del confronto: {exc}")
            st.stop()
    else:
        st.info("Completa i dati richiesti per visualizzare il confronto.")

    st.divider()

    # 5) Export Excel
    st.header("5️⃣ Scarica Excel")
    st.session_state["export_mode"] = "ENTRAMBE"
    st.caption(f"Export: Bolletta + {selected_supplier_label()} variabile + {selected_supplier_label()} fissa")

    def build_excel_bytes():
        wb = openpyxl.load_workbook(TEMPLATE_XLSX)
        ws = wb["Confronto"]
        apply_export_labels(ws, st.session_state["nome_cliente"])
        rm = find_row_map(ws)
        validate_row_map(rm)

        write_export_metadata(ws)

        comparison_values = build_comparison_values()
        comm = comparison_values["commodity"]
        b_vals = comparison_values["bolletta"]
        c_vals = comparison_values["variabile"]
        d_vals = comparison_values["fissa"]

        # B
        write_column(ws, rm, "B", b_vals, comm)
        ws[f"B{rm['accise_iva']}"] = b_vals["accise_iva"]
        apply_total_formula(ws, rm, "B")

        mode = st.session_state["export_mode"]
        has_offer_var = comparison_values["has_offer_var"]
        has_offer_fix = comparison_values["has_offer_fix"]
        if mode == "ENTRAMBE":
            if has_offer_var:
                write_column(ws, rm, "C", c_vals, comm)
                ws[f"C{rm['accise_iva']}"] = c_vals["accise_iva"]
                apply_total_formula(ws, rm, "C")
            else:
                fill_column_nd(ws, rm, "C")
            if has_offer_fix:
                write_column(ws, rm, "D", d_vals, comm)
                ws[f"D{rm['accise_iva']}"] = d_vals["accise_iva"]
                apply_total_formula(ws, rm, "D")
            else:
                fill_column_nd(ws, rm, "D")
        elif mode == "VARIABILE":
            if has_offer_var:
                write_column(ws, rm, "C", c_vals, comm)
                ws[f"C{rm['accise_iva']}"] = c_vals["accise_iva"]
                apply_total_formula(ws, rm, "C")
            else:
                fill_column_nd(ws, rm, "C")
            if has_offer_fix:
                fill_column_na(ws, rm, "D")
            else:
                fill_column_nd(ws, rm, "D")
        else:
            if has_offer_var:
                fill_column_na(ws, rm, "C")
            else:
                fill_column_nd(ws, rm, "C")
            if has_offer_fix:
                write_column(ws, rm, "D", d_vals, comm)
                ws[f"D{rm['accise_iva']}"] = d_vals["accise_iva"]
                apply_total_formula(ws, rm, "D")
            else:
                fill_column_nd(ws, rm, "D")

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    if can_calc:
        if not st.session_state["excel_ready"]:
            st.info("Il confronto è pronto. Prepara l'Excel solo quando vuoi scaricarlo.")
            if st.button("Prepara Excel", use_container_width=True):
                st.session_state["excel_ready"] = True
            else:
                print("APP_SECTION excel_waiting_user_action", flush=True)
        if st.session_state["excel_ready"]:
            try:
                print("APP_SECTION excel_build_start", flush=True)
                data = build_excel_bytes()
                print("APP_SECTION excel_build_done", flush=True)
            except Exception as exc:
                print(f"APP_ERROR excel_build {type(exc).__name__}: {exc}", flush=True)
                st.error(f"❌ Errore durante la generazione Excel: {exc}")
                st.stop()
            nome_file = safe_nome_cognome(st.session_state["nome_cliente"])
            comm_lab = commodity_label(st.session_state["commodity"])
            mode = st.session_state["export_mode"]
            file_name = f"confronto_{supplier_slug()}_{nome_file}_{comm_lab}_{mode}.xlsx"

            href, download_name = write_static_download(data, file_name)
            render_blob_download_button(data, file_name)
            st.caption("Se il pulsante non scarica, copia il link diretto qui sotto in una nuova scheda.")
            st.caption(f"Link diretto file: {href}")
            print(f"APP_SECTION excel_blob_button_done href={href}", flush=True)
    else:
        st.info("Il file Excel sarà disponibile appena il confronto è completo.")

    st.divider()
    st.caption("Energia Solidale")
    st.caption("Associazione senza scopo di lucro - Chioggia (VE) - info@energiasolidale.org")
    print("APP_SECTION footer_render_done", flush=True)
