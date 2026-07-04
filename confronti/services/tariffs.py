import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

from .utils import (
    EON_TARIFFE_DIR,
    INDICI_XLSX,
    PROVIDERS,
    SEGMENTS,
    TARIFFE_BASE,
    bool_from_data,
    clean_text,
    month_key_from_date,
    normalize_provider,
    normalize_tariff_selection_mode,
    normalize_providers,
    parse_date_any,
    parse_number,
    provider_label,
)


def parse_index_month(x):
    if x is None:
        return ""
    if isinstance(x, datetime):
        return month_key_from_date(x.date())
    if isinstance(x, date):
        return month_key_from_date(x)
    s = str(x).strip()
    m = re.search(r"(\d{4})[-/](\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    parsed = parse_date_any(s)
    return month_key_from_date(parsed) if parsed else ""


def load_indici_rows(path=None):
    path = path or INDICI_XLSX
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "indici" not in wb.sheetnames:
        return []
    ws = wb["indici"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [clean_text(c).lower() for c in rows[0]]
    if not {"mese", "pun", "psv"}.issubset(set(header)):
        return []
    i_mese = header.index("mese")
    i_pun = header.index("pun")
    i_psv = header.index("psv")
    out = []
    for r in rows[1:]:
        if not r or not any(r):
            continue
        mese = parse_index_month(r[i_mese] if i_mese < len(r) else None)
        if not mese:
            continue
        out.append(
            {
                "mese": mese,
                "pun": parse_number(r[i_pun] if i_pun < len(r) else None),
                "psv": parse_number(r[i_psv] if i_psv < len(r) else None),
            }
        )
    return sorted(out, key=lambda row: row["mese"])


def select_indice_for_bill_period(rows, bill_start, bill_end):
    if not rows:
        return None, "missing"
    if bill_end < bill_start:
        bill_start, bill_end = bill_end, bill_start
    start_month = month_key_from_date(bill_start)
    end_month = month_key_from_date(bill_end)
    in_period = [r for r in rows if start_month <= r["mese"] <= end_month]
    if in_period:
        return in_period[-1], "period"
    before_end = [r for r in rows if r["mese"] <= end_month]
    if before_end:
        return before_end[-1], "before_end"
    return rows[-1], "latest_available"


def tariff_month_key(path):
    try:
        parts = path.relative_to(TARIFFE_BASE).parts
    except ValueError:
        parts = path.parts
    for i, part in enumerate(parts):
        if re.match(r"^\d{4}-\d{2}$", part):
            return part
        if re.match(r"^\d{4}$", part) and i + 1 < len(parts) and re.match(r"^\d{2}$", parts[i + 1]):
            return f"{part}-{parts[i + 1]}"
    match = re.search(r"(\d{4})-(\d{2})", path.name)
    if match:
        return f"{match.group(1)}-{match.group(2)}"
    return ""


def tariff_matches_segment(path, segmento):
    segment_norm = clean_text(segmento).upper()
    name = path.name.lower()
    parts = {p.lower() for p in path.parts}
    if path.suffix.lower() != ".xlsx" or name.startswith("~$"):
        return False
    if segment_norm == "BUSINESS":
        is_micro = "microbusiness" in parts or "microbusiness" in name or "micro business" in str(path).lower()
        return not is_micro and ("business" in parts or "business" in name)
    if segment_norm == "MICROBUSINESS":
        return "microbusiness" in parts or "microbusiness" in name or "micro business" in str(path).lower()
    return ("residenziale" in parts or "template" in name) and "business" not in name


def tariff_segment_fallbacks(segmento):
    segment_norm = clean_text(segmento).upper()
    if segment_norm == "MICROBUSINESS":
        return ("MICROBUSINESS", "BUSINESS")
    return (segment_norm,)


def _tariff_file_candidates_for_segment_exact(segmento, provider="ILLUMIA"):
    provider_norm = normalize_provider(provider)
    if provider_norm in {"EON", "CVE"}:
        if not EON_TARIFFE_DIR.exists():
            return []
        prefix = provider_norm.lower()
        candidates = [
            p for p in EON_TARIFFE_DIR.glob(f"{prefix}_tariffe_*.xlsx") if p.is_file() and not p.name.startswith("~$")
        ]
    else:
        if not TARIFFE_BASE.exists():
            return []
        candidates = [p for p in TARIFFE_BASE.rglob("*.xlsx") if tariff_matches_segment(p, segmento)]
    candidates.sort(key=lambda p: (tariff_month_key(p), p.stat().st_mtime, p.name))
    return candidates


def tariff_file_candidates_for_segment(segmento, provider="ILLUMIA"):
    for effective_segment in tariff_segment_fallbacks(segmento):
        candidates = _tariff_file_candidates_for_segment_exact(effective_segment, provider)
        if candidates:
            return candidates
    return []


def select_tariffe_file_from_candidates(candidates, target_month=""):
    candidates = list(candidates)
    if not candidates:
        return None
    target_month = clean_text(target_month)
    if not target_month:
        return candidates[-1]
    with_month = [p for p in candidates if tariff_month_key(p)]
    if not with_month:
        return candidates[-1]
    exact = [p for p in with_month if tariff_month_key(p) == target_month]
    return exact[-1] if exact else None


def load_latest_eon_tariffe_file():
    candidates = tariff_file_candidates_for_segment("RESIDENZIALE", "EON")
    return candidates[-1] if candidates else None


def load_tariffe_file_for_segment(segmento, provider="ILLUMIA", selection_mode="LATEST", target_month="", commodity=""):
    offer_file, _effective_segment = load_tariffe_file_for_segment_with_effective_segment(
        segmento, provider, selection_mode, target_month, commodity
    )
    return offer_file


def load_tariffe_file_for_segment_with_effective_segment(
    segmento, provider="ILLUMIA", selection_mode="LATEST", target_month="", commodity=""
):
    requested_segment = clean_text(segmento).upper()
    for effective_segment in tariff_segment_fallbacks(requested_segment):
        candidates = _tariff_file_candidates_for_segment_exact(effective_segment, provider)
        if not candidates:
            continue
        if normalize_tariff_selection_mode(selection_mode) == "PERIOD":
            selected = select_tariffe_file_from_candidates(candidates, target_month)
            if selected and tariff_file_has_context_options(selected, provider, effective_segment, commodity):
                return selected, effective_segment
            continue
        for selected in reversed(candidates):
            if tariff_file_has_context_options(selected, provider, effective_segment, commodity):
                return selected, effective_segment
        return candidates[-1], effective_segment
    return None, requested_segment


def tariff_file_has_context_options(path, provider, segmento, commodity=""):
    rows = load_tariffe_from_path(path)
    filtered_rows = filter_rows_by_context(rows, provider, segmento)
    return rows_have_tariff_options(filtered_rows, commodity)


def get_file_valid_range(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if "tariffe" not in wb.sheetnames:
        return None, None
    ws = wb["tariffe"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None, None
    header = [clean_text(c).lower() for c in rows[0]]
    if "valid_from" not in header or "valid_to" not in header:
        return None, None
    i_from = header.index("valid_from")
    i_to = header.index("valid_to")
    dfs, dts = [], []
    for r in rows[1:]:
        if not r:
            continue
        vf = parse_date_any(r[i_from] if i_from < len(r) else None)
        vt = parse_date_any(r[i_to] if i_to < len(r) else None)
        if vf:
            dfs.append(vf)
        if vt:
            dts.append(vt)
    if not dfs or not dts:
        return None, None
    return min(dfs), max(dts)


def valid_range_from_rows(rows):
    dfs, dts = [], []
    for r in rows:
        vf = parse_date_any(r.get("valid_from"))
        vt = parse_date_any(r.get("valid_to"))
        if vf:
            dfs.append(vf)
        if vt:
            dts.append(vt)
    return (min(dfs) if dfs else None, max(dts) if dts else None)


def load_tariffe_from_path(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "tariffe" not in wb.sheetnames:
        return []
    ws = wb["tariffe"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [clean_text(c).lower() for c in rows[0]]
    out = []
    for r in rows[1:]:
        if not any(r):
            continue
        d = {key: (r[i] if i < len(r) else None) for i, key in enumerate(header)}
        d["commodity"] = clean_text(d.get("commodity")).upper()
        d["offer_type"] = clean_text(d.get("offer_type")).upper()
        d["offer_name"] = clean_text(d.get("offer_name"))
        d["component"] = clean_text(d.get("component")).lower()
        d["provider"] = normalize_provider(d.get("provider"))
        d["segmento"] = clean_text(d.get("segmento")).upper()
        d["is_sellable"] = clean_text(d.get("is_sellable")).upper()
        d["value_num"] = parse_number(d.get("value"))
        out.append(d)
    return out


def filter_rows_by_context(rows, provider, segmento):
    provider_norm = normalize_provider(provider)
    segment_norm = clean_text(segmento).upper()
    if provider_norm == "ILLUMIA":
        return [r for r in rows if normalize_provider(r.get("provider")) == "ILLUMIA"]
    return [
        r
        for r in rows
        if normalize_provider(r.get("provider")) == provider_norm
        and clean_text(r.get("segmento")).upper() == segment_norm
    ]


def rows_have_tariff_options(rows, commodity=""):
    commodity_norm = clean_text(commodity).upper()
    if commodity_norm:
        return bool(offer_names(rows, commodity_norm, "VARIABILE") or offer_names(rows, commodity_norm, "FISSA"))
    return any(row_is_sellable(r) and clean_text(r.get("offer_name")) for r in rows)


def filter_rows_by_context_with_fallback(rows, provider, segmento, commodity=""):
    segment_norm = clean_text(segmento).upper()
    primary_rows = filter_rows_by_context(rows, provider, segment_norm)
    if segment_norm == "MICROBUSINESS" and not rows_have_tariff_options(primary_rows, commodity):
        business_rows = filter_rows_by_context(rows, provider, "BUSINESS")
        if rows_have_tariff_options(business_rows, commodity):
            return business_rows, "BUSINESS"
    return primary_rows, segment_norm


def cve_row_matches_annual_context(row, annual_consumption, over70):
    requires_over70 = bool_from_data(row.get("requires_over70"))
    if over70:
        return requires_over70
    if requires_over70:
        return False
    annual = max(0.0, float(annual_consumption or 0.0))
    annual_min = clean_text(row.get("annual_min"))
    annual_max = clean_text(row.get("annual_max"))
    min_value = parse_number(annual_min) if annual_min else None
    max_value = parse_number(annual_max) if annual_max else None
    if min_value is not None and annual <= min_value:
        return False
    if max_value is not None and annual > max_value:
        return False
    return True


def filter_cve_rows_by_annual_context(rows, commodity, annual_consumption, over70=False):
    commodity_norm = clean_text(commodity).upper()
    annual = max(0.0, float(annual_consumption or 0.0))
    return [
        row
        for row in rows
        if clean_text(row.get("commodity")).upper() == commodity_norm
        and cve_row_matches_annual_context(row, annual, bool_from_data(over70))
    ]


def filter_rows_by_offer(rows, commodity, offer_type, offer_name):
    c = commodity.upper()
    o = offer_type.upper()
    return [r for r in rows if r.get("commodity") == c and r.get("offer_type") == o and clean_text(r.get("offer_name")) == offer_name]


def row_is_sellable(row):
    marker = clean_text(row.get("is_sellable")).upper()
    return marker not in {"FALSE", "NO", "0"}


def offer_names(rows, commodity, offer_type):
    c = commodity.upper()
    o = offer_type.upper()
    subset = [r for r in rows if r.get("commodity") == c and r.get("offer_type") == o and row_is_sellable(r)]
    counts = {}
    for r in subset:
        name = clean_text(r.get("offer_name"))
        if name:
            counts[name] = counts.get(name, 0) + 1
    return [name for name, _count in sorted(counts.items(), key=lambda x: (-x[1], x[0]))]


def select_offer_name(rows, commodity, offer_type, segmento, preferred=""):
    names = offer_names(rows, commodity, offer_type)
    preferred = clean_text(preferred)
    if preferred and preferred in names:
        return preferred
    if segmento == "BUSINESS" and offer_type.upper() == "FISSA" and not names:
        return ""
    if not names:
        return ""
    return names[0]


_offer_options_cache = None


def offer_options_payload():
    global _offer_options_cache
    if _offer_options_cache is not None:
        return _offer_options_cache
    payload = {}
    for provider in PROVIDERS:
        for segmento in SEGMENTS:
            for commodity in ("GAS", "EE"):
                rows = []
                for effective_segment in tariff_segment_fallbacks(segmento):
                    segment_rows = []
                    for offer_file in _tariff_file_candidates_for_segment_exact(effective_segment, provider):
                        segment_rows.extend(
                            filter_rows_by_context(load_tariffe_from_path(offer_file), provider, effective_segment)
                        )
                    if rows_have_tariff_options(segment_rows, commodity):
                        rows = segment_rows
                        break
                key = f"{provider}|{segmento}|{commodity}"
                payload[key] = {
                    "VARIABILE": offer_names(rows, commodity, "VARIABILE"),
                    "FISSA": offer_names(rows, commodity, "FISSA"),
                }
    _offer_options_cache = payload
    return payload


def invalidate_offer_options_cache():
    global _offer_options_cache
    _offer_options_cache = None
