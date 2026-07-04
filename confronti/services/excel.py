import io
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .utils import (
    TEMPLATE_XLSX,
    clean_text,
    normalize_accessory_services_vat_label,
    accessory_services_vat_rate,
    safe_download_filename,
    _excel_decimal,
)


def find_row_map(ws):
    rm = {}
    for r in range(1, ws.max_row + 1):
        t = clean_text(ws[f"A{r}"].value).lower()
        if not t:
            continue
        if "vendita consumo" in t:
            rm["vendita_consumo"] = r
        elif "vendita fissa luce" in t:
            rm["vendita_fissa_luce"] = r
        elif "vendita fissa" in t and "gas" in t:
            rm["vendita_fissa_gas"] = r
        elif "rete" in t and "consumi" in t:
            rm["rete_consumi"] = r
        elif "rete" in t and "fissa" in t:
            rm["rete_fissa"] = r
        elif "quota potenza" in t:
            rm["quota_potenza"] = r
        elif "sconti" in t:
            rm["sconti"] = r
        elif "ricalcoli" in t or "partite" in t:
            rm["ricalcoli"] = r
        elif "bonus social" in t:
            rm["bonus_sociale"] = r
        elif "arrotondamenti" in t:
            rm["arrotondamenti"] = r
        elif "servizi" in t and "access" in t:
            rm["servizi_accessori"] = r
        elif "accise" in t and "iva" in t:
            rm["accise_iva"] = r
        elif t.startswith("accise"):
            rm["accise"] = r
        elif t == "iva":
            rm["iva"] = r
        elif t == "totale":
            rm["totale"] = r
    if "arrotondamenti" not in rm and "bonus_sociale" not in rm and "accise_iva" in rm:
        candidate = rm["accise_iva"] - 1
        if candidate > 0:
            rm["arrotondamenti"] = candidate
    return rm


def copy_row_format(ws, source_row, target_row):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        source = ws.cell(row=source_row, column=col)
        target = ws.cell(row=target_row, column=col)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def ensure_export_rows(ws):
    rm = find_row_map(ws)
    missing_count = 0
    if "arrotondamenti" not in rm:
        missing_count += 1
    if "servizi_accessori" not in rm:
        missing_count += 1
    if "accise" not in rm:
        missing_count += 1
    if "iva" not in rm:
        missing_count += 1
    if missing_count == 0:
        return

    insert_at = rm.get("accise_iva", rm.get("totale", ws.max_row))
    ws.insert_rows(insert_at, amount=missing_count)
    source_row = insert_at + missing_count
    for row in range(insert_at, insert_at + missing_count):
        copy_row_format(ws, source_row, row)


def validate_row_map(rm, commodity="EE"):
    required = [
        "vendita_consumo",
        "vendita_fissa_gas",
        "rete_consumi",
        "rete_fissa",
        "sconti",
        "ricalcoli",
        "bonus_sociale",
        "arrotondamenti",
        "servizi_accessori",
        "accise",
        "iva",
        "accise_iva",
        "totale",
    ]
    if commodity == "EE":
        required.extend(["vendita_fissa_luce", "quota_potenza"])
    missing = [key for key in required if key not in rm]
    if missing:
        raise ValueError("Template Excel incompleto: mancano le righe " + ", ".join(missing))


def remove_export_rows_for_commodity(ws, rm, commodity):
    if commodity != "GAS":
        return
    rows_to_delete = [rm[key] for key in ("quota_potenza", "vendita_fissa_luce") if key in rm]
    for row in sorted(set(rows_to_delete), reverse=True):
        ws.delete_rows(row, amount=1)


def apply_export_labels(ws, nome_cliente, servizi_accessori_iva_label="22%", commodity="EE"):
    accise_label = "Accise + Addizionale regionale" if commodity == "GAS" else "Accise"
    labels = {
        1: nome_cliente or "Cliente",
        3: "VOCE",
        4: "Vendita Consumo",
        5: "Rete e oneri di sistema Consumi",
        6: "Vendita Fissa Luce",
        7: "Vendita Fissa Gas",
        8: "Rete e oneri di sistema Fissa",
        9: "Quota Potenza",
        10: "Sconti",
        11: "Ricalcoli/Partite pregresse",
        12: "Bonus Sociale",
        13: "Arrotondamenti",
        14: f"Servizi accessori (IVA {servizi_accessori_iva_label})",
        15: accise_label,
        16: "IVA",
        17: "Accise e Iva",
        18: "Totale",
    }
    for row, value in labels.items():
        ws[f"A{row}"] = value


def write_export_metadata(ws, prepared, start_col="F"):
    calc = prepared["calc"]
    provider_lines = []
    file_lines = []
    for result in calc.get("provider_results", []):
        provider_lines.append(
            f"{result['provider_label']}: variabile {result['offer_var'] or 'N.D.'}; "
            f"fissa {result['offer_fix'] or 'N.D.'}"
        )
        file_lines.append(f"{result['provider_label']}: {result.get('offer_file') or 'N.D.'}")
    ws[f"{start_col}1"] = f"Fornitori confronto: {calc.get('providers_label', calc.get('provider_label', 'Illumia'))}"
    ws[f"{start_col}2"] = " | ".join(provider_lines) if provider_lines else "Offerte: N.D."
    ws[f"{start_col}3"] = f"Scadenza offerta bolletta: {calc.get('bill_offer_expiry_label', 'N.D.')}"
    ws[f"{start_col}3"].font = Font(bold=True, color="B3261E")
    ws[f"{start_col}4"] = f"File tariffe: {' | '.join(file_lines) if file_lines else 'N.D.'}"
    indice = calc["indice"] or {}
    from .utils import INDICI_XLSX
    ws[f"{start_col}5"] = f"Indice PUN/PSV: {indice.get('mese', 'N.D.')} ({INDICI_XLSX.name})"
    ws[f"{start_col}6"] = f"Tipo tariffa bolletta: {calc.get('bill_tariff_type_label', 'Variabile')}"
    ws[f"{start_col}7"] = f"Confronto eseguito: {calc.get('comparison_datetime_label', '')}"
    ws[f"{start_col}8"] = f"IVA servizi accessori: {calc.get('servizi_accessori_iva_label', '22%')}"
    ws[f"{start_col}9"] = f"Consumo annuo stimato: {calc.get('tax_annual_consumption_label', 'N.D.')}"
    ws[f"{start_col}10"] = f"Parametri Accise/IVA: {calc.get('fiscal_parameters_label', '')}"
    ws[f"{start_col}11"] = (
        f"Logica tariffe: {calc.get('tariff_selection_mode_label', 'Ultime tariffe disponibili')}"
    )
    ws[f"{start_col}12"] = f"Indirizzo fornitura: {calc.get('indirizzo_fornitura') or 'N.D.'}"
    ws[f"{start_col}13"] = f"Codice POD/PDR: {calc.get('pod_pdr') or 'N.D.'}"
    ws[f"{start_col}14"] = f"Fornitura: {'Gas' if calc.get('commodity') == 'GAS' else 'Luce'}"
    ws[f"{start_col}15"] = f"Periodo bolletta: {calc.get('period_label', 'N.D.')}"
    if calc.get("cve_selected"):
        ws[f"{start_col}16"] = f"CVE Over 70: {calc.get('cve_over70_label', 'No')}"


def apply_accise_formula_conforme(ws, rm, col_letter, servizi_accessori_iva_rate=0.0):
    acc = rm["accise_iva"]
    start = rm["vendita_consumo"]
    end = rm.get("accise", acc) - 1
    if "servizi_accessori" in rm:
        svc = rm["servizi_accessori"]
        rate = _excel_decimal(servizi_accessori_iva_rate)
        ws[f"{col_letter}{acc}"] = (
            f"=IFERROR((SUM({col_letter}{start}:{col_letter}{end})-{col_letter}{svc})"
            f"*(B{acc}-B{svc}*{rate})/(SUM(B{start}:B{end})-B{svc})"
            f"+{col_letter}{svc}*{rate},0)"
        )
    else:
        ws[f"{col_letter}{acc}"] = f"=SUM({col_letter}{start}:{col_letter}{end})*B{acc}/SUM(B{start}:B{end})"


def apply_total_formula(ws, rm, col_letter):
    acc = rm["accise_iva"]
    start = rm["vendita_consumo"]
    end = rm.get("accise", acc) - 1
    ws[f"{col_letter}{rm['totale']}"] = f"=SUM({col_letter}{start}:{col_letter}{end})+{col_letter}{acc}"


def apply_export_number_format(ws, rm, last_data_column_index):
    number_format = "#,##0.00"
    amount_rows = sorted(set(rm.values()))
    for column_index in range(2, last_data_column_index + 1):
        col_letter = get_column_letter(column_index)
        for row_index in amount_rows:
            cell = ws[f"{col_letter}{row_index}"]
            if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.number_format = number_format
    if isinstance(ws["C1"].value, (int, float)) or (isinstance(ws["C1"].value, str) and ws["C1"].value.startswith("=")):
        ws["C1"].number_format = number_format


def write_column(ws, rm, col, vals, commodity):
    ws[f"{col}{rm['vendita_consumo']}"] = float(vals["vendita_consumo"])
    if commodity == "GAS":
        ws[f"{col}{rm['vendita_fissa_gas']}"] = float(vals["vendita_fissa"])
        if "vendita_fissa_luce" in rm:
            ws[f"{col}{rm['vendita_fissa_luce']}"] = "N.A."
        if "quota_potenza" in rm:
            ws[f"{col}{rm['quota_potenza']}"] = "N.A."
    else:
        ws[f"{col}{rm['vendita_fissa_luce']}"] = float(vals["vendita_fissa"])
        ws[f"{col}{rm['vendita_fissa_gas']}"] = "N.A."
        ws[f"{col}{rm['quota_potenza']}"] = float(vals["quota_potenza"])
    ws[f"{col}{rm['rete_consumi']}"] = float(vals["rete_consumi"])
    ws[f"{col}{rm['rete_fissa']}"] = float(vals["rete_fissa"])
    ws[f"{col}{rm['sconti']}"] = float(vals["sconti"])
    ws[f"{col}{rm['ricalcoli']}"] = float(vals["ricalcoli"])
    if "bonus_sociale" in rm:
        ws[f"{col}{rm['bonus_sociale']}"] = float(vals.get("bonus_sociale", 0.0))
    if "arrotondamenti" in rm:
        ws[f"{col}{rm['arrotondamenti']}"] = float(vals["arrotondamenti"])
    if "servizi_accessori" in rm:
        ws[f"{col}{rm['servizi_accessori']}"] = float(vals.get("servizi_accessori", 0.0))
    for key in ("accise", "iva"):
        if key in rm:
            value = vals.get(key)
            ws[f"{col}{rm[key]}"] = "N.D." if value is None else float(value)


def fill_column_text(ws, rm, col, text):
    for key, r in rm.items():
        if key == "totale" and text == "N.A.":
            continue
        ws[f"{col}{r}"] = text


REPORT_SUMMARY_BASE_COLUMNS = [
    "File",
    "Cliente",
    "Indirizzo fornitura",
    "Codice POD/PDR",
    "Fornitura",
    "Periodo bolletta",
    "Consumo",
    "Fornitori confronto",
]


def _summary_clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def _metadata_value(ws, prefix):
    needle = str(prefix).lower()
    for row in ws.iter_rows(values_only=True):
        for value in row:
            if isinstance(value, str) and value.lower().startswith(needle):
                return value.split(":", 1)[1].strip() if ":" in value else ""
    return ""


def _formula_cell_value(ws, cell, seen=None):
    value = cell.value
    if not (isinstance(value, str) and value.startswith("=")):
        return _summary_clean(value)
    seen = seen or set()
    if cell.coordinate in seen:
        return ""
    seen.add(cell.coordinate)
    formula = value.replace(" ", "").upper()
    match = re.fullmatch(r"=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)\+([A-Z]+)(\d+)", formula)
    if match:
        start_col, start_row, end_col, end_row, extra_col, extra_row = match.groups()
        total = 0.0
        for row in range(int(start_row), int(end_row) + 1):
            for cells in ws[f"{start_col}{row}:{end_col}{row}"]:
                for item in cells:
                    item_value = _formula_cell_value(ws, item, seen.copy())
                    if isinstance(item_value, (int, float)):
                        total += float(item_value)
        extra_value = _formula_cell_value(ws, ws[f"{extra_col}{extra_row}"], seen.copy())
        if isinstance(extra_value, (int, float)):
            total += float(extra_value)
        return round(total, 6)
    match = re.fullmatch(r"=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)", formula)
    if match:
        start_col, start_row, end_col, end_row = match.groups()
        total = 0.0
        for row in range(int(start_row), int(end_row) + 1):
            for cells in ws[f"{start_col}{row}:{end_col}{row}"]:
                for item in cells:
                    item_value = _formula_cell_value(ws, item, seen.copy())
                    if isinstance(item_value, (int, float)):
                        total += float(item_value)
        return round(total, 6)
    return value


def _report_data_columns(ws):
    columns = []
    col = 2
    while col <= ws.max_column:
        label = clean_text(ws.cell(row=3, column=col).value)
        if not label:
            break
        columns.append((col, label))
        col += 1
    return columns


def summarize_report_workbook(uploaded_file):
    import re
    uploaded_file.seek(0)
    wb = openpyxl.load_workbook(uploaded_file, data_only=False)
    ws = wb["Confronto"] if "Confronto" in wb.sheetnames else wb.active
    file_name = Path(getattr(uploaded_file, "name", "report.xlsx")).name
    summary = {
        "File": file_name,
        "Cliente": clean_text(ws["A1"].value) or "N.D.",
        "Indirizzo fornitura": _metadata_value(ws, "Indirizzo fornitura") or "N.D.",
        "Codice POD/PDR": _metadata_value(ws, "Codice POD/PDR") or "N.D.",
        "Fornitura": _metadata_value(ws, "Fornitura") or "N.D.",
        "Periodo bolletta": _metadata_value(ws, "Periodo bolletta") or "N.D.",
        "Consumo": _summary_clean(ws["C1"].value),
        "Fornitori confronto": _metadata_value(ws, "Fornitori confronto") or "N.D.",
    }
    label_rows = []
    for row in range(4, ws.max_row + 1):
        label = clean_text(ws.cell(row=row, column=1).value)
        if label:
            label_rows.append((row, label))
        if label.lower() == "totale":
            break
    for col, header in _report_data_columns(ws):
        for row, label in label_rows:
            value = _formula_cell_value(ws, ws.cell(row=row, column=col))
            summary[f"{header} - {label}"] = value
    return summary


def build_reports_summary(uploaded_files):
    rows = []
    warnings = []
    for uploaded_file in uploaded_files:
        try:
            rows.append(summarize_report_workbook(uploaded_file))
        except Exception as exc:
            warnings.append(f"{Path(getattr(uploaded_file, 'name', 'report')).name}: file non leggibile ({exc}).")
    dynamic_columns = []
    for row in rows:
        for key in row:
            if key not in REPORT_SUMMARY_BASE_COLUMNS and key not in dynamic_columns:
                dynamic_columns.append(key)
    columns = REPORT_SUMMARY_BASE_COLUMNS + dynamic_columns
    return {"columns": columns, "rows": rows, "warnings": warnings, "count": len(rows)}


def _is_summary_visible_column(label):
    label = clean_text(label)
    if label in REPORT_SUMMARY_BASE_COLUMNS:
        return True
    return label.endswith(" - Totale")


def _is_number_for_summary(value):
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def build_reports_summary_excel(summary):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sunto Report"
    columns = summary.get("columns") or []
    rows = summary.get("rows") or []
    for col_index, label in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_index, value=label)
        cell.font = Font(bold=True)
    for row_index, row in enumerate(rows, start=2):
        for col_index, label in enumerate(columns, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=row.get(label, ""))
            if _is_number_for_summary(cell.value):
                cell.number_format = "#,##0.00"
    ws.freeze_panes = "A2"
    if columns and rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"
    for col_index, label in enumerate(columns, start=1):
        col_letter = get_column_letter(col_index)
        max_len = len(str(label))
        for row_index in range(2, len(rows) + 2):
            value = ws.cell(row=row_index, column=col_index).value
            max_len = max(max_len, len(str(value or "")))
        width = min(max(max_len + 2, 12), 42)
        if _is_summary_visible_column(label) and label not in REPORT_SUMMARY_BASE_COLUMNS:
            width = min(max(width, 18), 26)
        ws.column_dimensions[col_letter].width = width
        ws.column_dimensions[col_letter].hidden = not _is_summary_visible_column(label)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_excel_bytes(data, prepared=None):
    from .comparison import prepare_comparison
    prepared = prepared or prepare_comparison(data)
    wb = openpyxl.load_workbook(TEMPLATE_XLSX)
    ws = wb["Confronto"]
    ensure_export_rows(ws)
    servizi_accessori_iva_label = normalize_accessory_services_vat_label(data.get("servizi_accessori_iva"))
    servizi_accessori_iva_rate = accessory_services_vat_rate(servizi_accessori_iva_label)
    values = prepared["values"]
    comm = values["commodity"]
    apply_export_labels(ws, data.get("nome_cliente", "Cliente"), servizi_accessori_iva_label, comm)
    rm = find_row_map(ws)
    remove_export_rows_for_commodity(ws, rm, comm)
    rm = find_row_map(ws)
    validate_row_map(rm, comm)
    ws["B1"] = None
    ws["C1"] = float(data["consumo"])
    ws["B3"] = "Bolletta"

    b_vals = values["bolletta"]

    write_column(ws, rm, "B", b_vals, comm)
    ws[f"B{rm['accise_iva']}"] = b_vals["accise_iva"]
    apply_total_formula(ws, rm, "B")

    for offset, column in enumerate(values["offer_columns"], start=3):
        col_letter = get_column_letter(offset)
        ws[f"{col_letter}3"] = column["label"]
        if column["has_offer"]:
            write_column(ws, rm, col_letter, column["vals"], comm)
            ws[f"{col_letter}{rm['accise_iva']}"] = column["vals"]["accise_iva"]
            apply_total_formula(ws, rm, col_letter)
        else:
            fill_column_text(ws, rm, col_letter, "N.D.")

    apply_export_number_format(ws, rm, 2 + len(values["offer_columns"]))

    metadata_column = get_column_letter(3 + len(values["offer_columns"]) + 1)
    write_export_metadata(ws, prepared, metadata_column)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
