import os
from datetime import date
from io import BytesIO
from io import StringIO
from unittest.mock import patch

from django import forms
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.test import Client, SimpleTestCase, TestCase
from openpyxl import load_workbook

from .bill_parser import ParsedBill, parse_bill_text
from .forms import ConfrontoForm, CustomerInviteForm
from .models import InviteCode
from .roles import ILLUMIA_OPERATOR_GROUP
from . import services


def valid_payload(**overrides):
    data = {
        "nome_cliente": "Mario Rossi",
        "pod_pdr": "IT001E12345678",
        "segmento": "RESIDENZIALE",
        "commodity": "EE",
        "bill_tariff_type": "VARIABILE",
        "tariff_selection_mode": "LATEST",
        "provider": "ILLUMIA",
        "providers": ["ILLUMIA"],
        "tax_primary_home": "SI",
        "tax_power_kw": "3",
        "tax_annual_consumption": "1200",
        "tax_region": "Veneto",
        "servizi_accessori_iva": "22%",
        "offer_var_choice": "",
        "offer_fix_choice": "",
        "offer_var_choice_illumia": "",
        "offer_fix_choice_illumia": "",
        "offer_var_choice_eon": "",
        "offer_fix_choice_eon": "",
        "offer_var_choice_cve": "",
        "offer_fix_choice_cve": "",
        "cve_over70": "",
        "bill_start": "2026-01",
        "bill_end": "2026-03",
        "consumo": "100",
        "b_vendita_consumo": "39.86",
        "b_rete_consumi": "8.47",
        "b_vendita_fissa": "3.70",
        "b_rete_fissa": "0.3967",
        "b_quota_potenza": "12.64",
        "b_sconti": "0",
        "b_ricalcoli": "0",
        "b_bonus_sociale": "21.60",
        "b_arrotondamenti": "0",
        "b_servizi_accessori": "0",
        "bill_offer_expiry": "2026-12-31",
        "b_accise_iva": "12.12",
        "action": "calculate",
    }
    data.update(overrides)
    if "provider" in overrides and "providers" not in overrides:
        data["providers"] = [data["provider"]]
    selected_provider = services.normalize_provider(data["providers"][0]) if data.get("providers") else "ILLUMIA"
    if "offer_var_choice" in overrides:
        data[f"offer_var_choice_{selected_provider.lower()}"] = data["offer_var_choice"]
    if "offer_fix_choice" in overrides:
        data[f"offer_fix_choice_{selected_provider.lower()}"] = data["offer_fix_choice"]
    return data


def service_data(**overrides):
    form = ConfrontoForm(valid_payload(**overrides))
    assert form.is_valid(), form.errors.as_json()
    return form.service_data()


class ServiceUtilityTests(SimpleTestCase):
    def test_parse_number_handles_italian_currency(self):
        self.assertEqual(services.parse_number("€ 1.234,56"), 1234.56)
        self.assertEqual(services.parse_number("-€ 21,60"), -21.6)
        self.assertEqual(services.parse_number(""), 0.0)

    def test_cve_over70_field_has_compact_checkbox_style(self):
        css = (services.BASE_DIR / "confronti/static/confronti/style.css").read_text()
        self.assertIn(".checkbox-field", css)
        self.assertIn("width: fit-content;", css)
        self.assertIn(".checkbox-field[hidden]", css)
        self.assertIn(".checkbox-field input", css)
        self.assertIn("height: 16px;", css)

    def test_power_fields_have_hide_style(self):
        css = (services.BASE_DIR / "confronti/static/confronti/style.css").read_text()
        self.assertIn("[data-power-field][hidden]", css)
        self.assertIn("display: none !important;", css)

    def test_billing_months_and_labels(self):
        self.assertEqual(services.billing_months_from_dates(date(2026, 1, 1), date(2026, 3, 31)), 3)
        self.assertEqual(services.billing_label_from_months(1), "MENSILE")
        self.assertEqual(services.billing_label_from_months(3), "TRIMESTRALE")
        self.assertEqual(services.bill_period_label(date(2026, 1, 1), date(2026, 3, 31)), "Gennaio 2026 - Marzo 2026")
        self.assertTrue(
            services.bill_period_outside_offer_validity(
                date(2026, 1, 1),
                date(2026, 3, 31),
                date(2026, 4, 1),
                date(2026, 4, 30),
            )
        )
        self.assertFalse(
            services.bill_period_outside_offer_validity(
                date(2026, 4, 1),
                date(2026, 4, 30),
                date(2026, 4, 1),
                date(2026, 4, 30),
            )
        )

    def test_select_indice_uses_latest_month_inside_bill_period(self):
        rows = [
            {"mese": "2026-01", "pun": 1, "psv": 10},
            {"mese": "2026-02", "pun": 2, "psv": 20},
            {"mese": "2026-03", "pun": 3, "psv": 30},
            {"mese": "2026-04", "pun": 4, "psv": 40},
        ]
        row, reason = services.select_indice_for_bill_period(rows, date(2026, 1, 1), date(2026, 3, 31))
        self.assertEqual(reason, "period")
        self.assertEqual(row["mese"], "2026-03")

    def test_build_comparison_values_keeps_bill_fixed_items_as_entered_and_copies_bonus(self):
        data = {
            "commodity": "EE",
            "segmento": "RESIDENZIALE",
            "consumo": 100,
            "tax_primary_home": "SI",
            "tax_power_kw": 3,
            "tax_annual_consumption": 1200,
            "tax_region": "Veneto",
            "b_vendita_consumo": 40,
            "b_vendita_fissa": 10,
            "b_rete_consumi": 8,
            "b_rete_fissa": 1,
            "b_quota_potenza": 12,
            "b_sconti": 0,
            "b_ricalcoli": 0,
            "b_bonus_sociale": 21.6,
            "b_arrotondamenti": 0,
            "b_servizi_accessori": 0,
            "b_accise_iva": 12,
            "servizi_accessori_iva": "22%",
            "ill_sconto_var": -3,
            "ill_sconto_fix": -3,
        }
        calc = {
            "billing_months": 3,
            "v_cons": 35,
            "v_fix": 15,
            "f_cons": 30,
            "f_fix": 18,
            "offer_var": "Variabile",
            "offer_fix": "Fissa",
        }
        values = services.build_comparison_values(data, calc)
        self.assertEqual(values["bolletta"]["vendita_fissa"], 10)
        self.assertEqual(values["bolletta"]["rete_fissa"], 1)
        self.assertEqual(values["bolletta"]["bonus_sociale"], -21.6)
        self.assertEqual(values["variabile"]["bonus_sociale"], -21.6)
        self.assertEqual(values["fissa"]["bonus_sociale"], -21.6)
        self.assertAlmostEqual(values["fissa"]["accise"], 0.0, places=6)
        self.assertAlmostEqual(values["fissa"]["iva"], 6.6, places=6)
        self.assertAlmostEqual(values["fissa"]["accise_iva"], 6.6, places=6)

    def test_accessory_services_stay_only_on_bill_column(self):
        data = {
            "commodity": "EE",
            "segmento": "RESIDENZIALE",
            "consumo": 100,
            "tax_primary_home": "SI",
            "tax_power_kw": 3,
            "tax_annual_consumption": 1200,
            "tax_region": "Veneto",
            "b_vendita_consumo": 40,
            "b_vendita_fissa": 10,
            "b_rete_consumi": 8,
            "b_rete_fissa": 1,
            "b_quota_potenza": 12,
            "b_sconti": 0,
            "b_ricalcoli": 0,
            "b_bonus_sociale": 21.6,
            "b_arrotondamenti": 0,
            "b_servizi_accessori": 5,
            "b_accise_iva": 12,
            "servizi_accessori_iva": "10%",
            "ill_sconto_var": -3,
            "ill_sconto_fix": -3,
        }
        calc = {
            "billing_months": 3,
            "v_cons": 35,
            "v_fix": 15,
            "f_cons": 30,
            "f_fix": 18,
            "offer_var": "Variabile",
            "offer_fix": "Fissa",
        }
        values = services.build_comparison_values(data, calc)
        self.assertEqual(values["bolletta"]["servizi_accessori"], 5)
        self.assertEqual(values["variabile"]["servizi_accessori"], 0)
        self.assertEqual(values["fissa"]["servizi_accessori"], 0)
        self.assertEqual(values["servizi_accessori_iva_label"], "10%")
        self.assertAlmostEqual(values["fissa"]["accise_iva"], 6.6, places=6)
        rows = services.build_comparison_table_rows(values)
        servizi_row = next(row for row in rows if row["voce"] == "Servizi accessori (IVA 10%)")
        self.assertEqual(servizi_row["cells"], ["€ 5,00", "€ 0,00", "€ 0,00"])

    def test_gas_annual_consumption_changes_accise_iva(self):
        data = {
            "commodity": "GAS",
            "segmento": "RESIDENZIALE",
            "consumo": 100,
            "tax_primary_home": "SI",
            "tax_power_kw": 0,
            "tax_annual_consumption": 1200,
            "tax_region": "Veneto",
            "b_vendita_consumo": 40,
            "b_vendita_fissa": 10,
            "b_rete_consumi": 8,
            "b_rete_fissa": 1,
            "b_quota_potenza": 0,
            "b_sconti": 0,
            "b_ricalcoli": 0,
            "b_bonus_sociale": 0,
            "b_arrotondamenti": 0,
            "b_servizi_accessori": 0,
            "b_accise_iva": 12,
            "servizi_accessori_iva": "22%",
            "ill_sconto_var": -3,
            "ill_sconto_fix": -3,
        }
        calc = {
            "billing_months": 1,
            "v_cons": 45,
            "v_fix": 6,
            "f_cons": 42,
            "f_fix": 7,
            "offer_var": "Variabile",
            "offer_fix": "Fissa",
        }
        offer_vals = {k: float(data.get(f"b_{k}", 0.0)) for k in services.KEYS}
        offer_vals["bonus_sociale"] = -abs(float(offer_vals.get("bonus_sociale", 0.0)))
        offer_vals["vendita_consumo"] = calc["f_cons"]
        offer_vals["vendita_fissa"] = calc["f_fix"]
        offer_vals["sconti"] = data["ill_sconto_fix"]
        offer_vals["ricalcoli"] = 0.0
        offer_vals["arrotondamenti"] = 0.0
        offer_vals["servizi_accessori"] = 0.0

        high_annual = services.calculate_tax_breakdown(data, calc, offer_vals, "GAS")["accise_iva"]
        low_annual = services.calculate_tax_breakdown(
            {**data, "tax_annual_consumption": 300},
            calc,
            offer_vals,
            "GAS",
        )["accise_iva"]
        self.assertNotEqual(round(high_annual, 6), round(low_annual, 6))

    def test_offer_tax_is_capped_to_bill_percentage_incidence(self):
        data = {
            "commodity": "GAS",
            "segmento": "RESIDENZIALE",
            "consumo": 100,
            "tax_primary_home": "SI",
            "tax_power_kw": 0,
            "tax_annual_consumption": 1200,
            "tax_region": "Veneto",
            "b_vendita_consumo": 40,
            "b_vendita_fissa": 10,
            "b_rete_consumi": 8,
            "b_rete_fissa": 1,
            "b_quota_potenza": 0,
            "b_sconti": 0,
            "b_ricalcoli": 0,
            "b_bonus_sociale": 0,
            "b_arrotondamenti": 0,
            "b_servizi_accessori": 0,
            "b_accise_iva": 5,
            "servizi_accessori_iva": "22%",
            "ill_sconto_var": -3,
            "ill_sconto_fix": -3,
        }
        calc = {
            "billing_months": 1,
            "v_cons": 45,
            "v_fix": 6,
            "f_cons": 42,
            "f_fix": 7,
            "offer_var": "Variabile",
            "offer_fix": "Fissa",
        }

        bill_vals = {k: float(data.get(f"b_{k}", 0.0)) for k in services.KEYS}
        bill_vals["bonus_sociale"] = -abs(float(bill_vals.get("bonus_sociale", 0.0)))
        bill_ratio = services.bill_tax_incidence_ratio(bill_vals, "GAS")

        uncapped_offer_vals = bill_vals.copy()
        uncapped_offer_vals["vendita_consumo"] = calc["f_cons"]
        uncapped_offer_vals["vendita_fissa"] = calc["f_fix"]
        uncapped_offer_vals["sconti"] = data["ill_sconto_fix"]
        uncapped_offer_vals["ricalcoli"] = 0.0
        uncapped_offer_vals["arrotondamenti"] = 0.0
        uncapped_offer_vals["servizi_accessori"] = 0.0
        uncapped_tax = services.calculate_tax_breakdown(data, calc, uncapped_offer_vals, "GAS")

        values = services.build_comparison_values(data, calc)
        capped_vals = values["fissa"]
        capped_total = services.comparison_subtotal(capped_vals, "GAS") * bill_ratio
        scale = capped_total / uncapped_tax["accise_iva"]
        fixed_column = next(column for column in values["offer_columns"] if column["offer_type"] == "FISSA")

        self.assertGreater(uncapped_tax["accise_iva"], capped_total)
        self.assertAlmostEqual(capped_vals["accise_iva"], capped_total, places=6)
        self.assertAlmostEqual(capped_vals["accise"], uncapped_tax["accise"] * scale, places=6)
        self.assertAlmostEqual(capped_vals["iva"], uncapped_tax["iva"] * scale, places=6)
        self.assertAlmostEqual(values["bill_tax_ratio"], bill_ratio, places=6)
        self.assertEqual(values["bill_tax_ratio_label"], "8,47%")
        self.assertTrue(fixed_column["tax_cap_applied"])
        self.assertEqual(fixed_column["tax_cap_status_label"], "Si")
        self.assertAlmostEqual(fixed_column["raw_tax_ratio"], uncapped_tax["accise_iva"] / services.comparison_subtotal(uncapped_offer_vals, "GAS"), places=6)
        self.assertAlmostEqual(fixed_column["tax_ratio"], bill_ratio, places=6)
        self.assertEqual(fixed_column["tax_ratio_label"], "8,47%")

    def test_table_marks_missing_illumia_offers_as_nd(self):
        data = service_data()
        calc = {
            "billing_months": 3,
            "v_cons": 0,
            "v_fix": 0,
            "f_cons": 0,
            "f_fix": 0,
            "offer_var": "",
            "offer_fix": "",
        }
        rows = services.build_comparison_table_rows(services.build_comparison_values(data, calc))
        self.assertTrue(all(row["variabile"] == "N.D." for row in rows))
        self.assertTrue(all(row["fissa"] == "N.D." for row in rows))

    def test_gas_comparison_table_hides_power_row(self):
        data = service_data(commodity="GAS", tax_power_kw="0", b_quota_potenza="0")
        calc = {
            "billing_months": 1,
            "v_cons": 45,
            "v_fix": 6,
            "f_cons": 42,
            "f_fix": 7,
            "offer_var": "Variabile",
            "offer_fix": "Fissa",
        }
        rows = services.build_comparison_table_rows(services.build_comparison_values(data, calc))
        labels = [row["voce"] for row in rows]
        self.assertNotIn("Vendita Fissa Luce", labels)
        self.assertNotIn("Quota Potenza", labels)
        self.assertIn("Vendita Fissa Gas", labels)

    def test_eon_offer_options_include_residential_and_business(self):
        options = services.offer_options_payload()
        self.assertIn("E.ON Flex Gas", options["EON|RESIDENZIALE|GAS"]["VARIABILE"])
        self.assertIn("E.ON Gas Tua", options["EON|RESIDENZIALE|GAS"]["FISSA"])
        self.assertIn("E.ON Gas Impresa CLSC", options["EON|MICROBUSINESS|GAS"]["VARIABILE"])
        self.assertIn("E.ON LuceDinamica ECO CLSE", options["EON|MICROBUSINESS|EE"]["VARIABILE"])
        self.assertIn("E.ON Profilo Dinamico Gas P", options["EON|BUSINESS|GAS"]["VARIABILE"])
        self.assertIn("E.ON Profilo Sicuro T", options["EON|BUSINESS|EE"]["FISSA"])

    def test_eon_business_latest_uses_latest_file_with_business_rows(self):
        latest_gas = services.load_tariffe_file_for_segment("BUSINESS", "EON", "LATEST", commodity="GAS")
        latest_luce = services.load_tariffe_file_for_segment("BUSINESS", "EON", "LATEST", commodity="EE")
        self.assertIn("2026-06", str(latest_gas))
        self.assertIn("2026-05", str(latest_luce))

    def test_eon_microbusiness_latest_uses_commodity_specific_file(self):
        latest_gas = services.load_tariffe_file_for_segment("MICROBUSINESS", "EON", "LATEST", commodity="GAS")
        latest_luce = services.load_tariffe_file_for_segment("MICROBUSINESS", "EON", "LATEST", commodity="EE")
        self.assertIn("2026-05", str(latest_gas))
        self.assertIn("2026-06", str(latest_luce))

    def test_cve_offer_options_include_residential_variable_tariffs_only(self):
        options = services.offer_options_payload()
        self.assertIn("CVE 1Casa Small Luce", options["CVE|RESIDENZIALE|EE"]["VARIABILE"])
        self.assertIn("CVE 1Casa Smart Luce", options["CVE|RESIDENZIALE|EE"]["VARIABILE"])
        self.assertIn("CVE 1Casa Big Gas", options["CVE|RESIDENZIALE|GAS"]["VARIABILE"])
        self.assertIn("CVE 1Casa Over 70 Gas", options["CVE|RESIDENZIALE|GAS"]["VARIABILE"])
        self.assertEqual(options["CVE|RESIDENZIALE|EE"]["FISSA"], [])
        self.assertEqual(options["CVE|RESIDENZIALE|GAS"]["FISSA"], [])

    def test_cve_annual_consumption_selects_small_smart_big_and_over70(self):
        ee_small = services.prepare_comparison(
            service_data(provider="CVE", providers=["CVE"], tax_annual_consumption="500")
        )
        ee_smart = services.prepare_comparison(
            service_data(provider="CVE", providers=["CVE"], tax_annual_consumption="1200")
        )
        ee_big = services.prepare_comparison(
            service_data(provider="CVE", providers=["CVE"], tax_annual_consumption="5000")
        )
        ee_over70 = services.prepare_comparison(
            service_data(provider="CVE", providers=["CVE"], tax_annual_consumption="1200", cve_over70="on")
        )
        self.assertEqual(ee_small["calc"]["offer_var"], "CVE 1Casa Small Luce")
        self.assertEqual(ee_smart["calc"]["offer_var"], "CVE 1Casa Smart Luce")
        self.assertEqual(ee_big["calc"]["offer_var"], "CVE 1Casa Big Luce")
        self.assertEqual(ee_over70["calc"]["offer_var"], "CVE 1Casa Over 70 Luce")
        self.assertEqual(ee_over70["calc"]["offer_fix"], "")

        gas_small = services.prepare_comparison(
            service_data(provider="CVE", providers=["CVE"], commodity="GAS", tax_annual_consumption="300", consumo="50")
        )
        gas_smart = services.prepare_comparison(
            service_data(provider="CVE", providers=["CVE"], commodity="GAS", tax_annual_consumption="800", consumo="50")
        )
        gas_big = services.prepare_comparison(
            service_data(provider="CVE", providers=["CVE"], commodity="GAS", tax_annual_consumption="2000", consumo="50")
        )
        gas_over70 = services.prepare_comparison(
            service_data(
                provider="CVE",
                providers=["CVE"],
                commodity="GAS",
                tax_annual_consumption="800",
                consumo="50",
                cve_over70="on",
            )
        )
        self.assertEqual(gas_small["calc"]["offer_var"], "CVE 1Casa Small Gas")
        self.assertEqual(gas_smart["calc"]["offer_var"], "CVE 1Casa Smart Gas")
        self.assertEqual(gas_big["calc"]["offer_var"], "CVE 1Casa Big Gas")
        self.assertEqual(gas_over70["calc"]["offer_var"], "CVE 1Casa Over 70 Gas")

    def test_cve_period_tariffe_requires_exact_month_while_latest_uses_available_file(self):
        latest = services.load_tariffe_file_for_segment("RESIDENZIALE", "CVE", "LATEST", "2026-06", commodity="EE")
        period_march = services.load_tariffe_file_for_segment("RESIDENZIALE", "CVE", "PERIOD", "2026-03", commodity="EE")
        period_june = services.load_tariffe_file_for_segment("RESIDENZIALE", "CVE", "PERIOD", "2026-06", commodity="EE")
        self.assertIn("cve_tariffe_2026-03.xlsx", str(latest))
        self.assertIn("cve_tariffe_2026-03.xlsx", str(period_march))
        self.assertIsNone(period_june)

    def test_missing_microbusiness_tariffe_falls_back_to_business(self):
        latest = services.load_tariffe_file_for_segment("MICROBUSINESS", "ILLUMIA", "LATEST", "2026-06")
        self.assertIn("/business/", str(latest))

        options = services.offer_options_payload()
        self.assertIn("GAS BUSINESS PREMIUM FLEX", options["ILLUMIA|MICROBUSINESS|GAS"]["VARIABILE"])

        prepared = services.prepare_comparison(
            service_data(
                segmento="MICROBUSINESS",
                provider="ILLUMIA",
                providers=["ILLUMIA"],
                commodity="GAS",
                tax_power_kw="0",
                b_quota_potenza="0",
            )
        )
        self.assertIn("/business/", prepared["calc"]["offer_file"])
        self.assertEqual(prepared["calc"]["provider_results"][0]["tariff_segment"], "BUSINESS")
        self.assertEqual(prepared["calc"]["offer_var"], "GAS PREMIUM FLEX BUSINESS")

    def test_tariff_selection_can_use_bill_period_month(self):
        latest = services.load_tariffe_file_for_segment("RESIDENZIALE", "ILLUMIA", "LATEST", "2026-03")
        period = services.load_tariffe_file_for_segment("RESIDENZIALE", "ILLUMIA", "PERIOD", "2026-03")
        self.assertIn("2026-06", str(latest))
        self.assertIn("2026-03", str(period))
        self.assertIsNone(services.load_tariffe_file_for_segment("RESIDENZIALE", "EON", "PERIOD", "2026-03"))

    def test_prepare_comparison_uses_period_tariffe_when_selected(self):
        latest = services.prepare_comparison(service_data(tariff_selection_mode="LATEST"))
        period = services.prepare_comparison(service_data(tariff_selection_mode="PERIOD"))
        self.assertIn("2026-06", latest["calc"]["offer_file"])
        self.assertIn("2026-03", period["calc"]["offer_file"])
        self.assertEqual(period["calc"]["tariff_selection_mode_label"], "Tariffe del periodo bolletta")
        self.assertEqual(period["calc"]["tariff_target_month"], "2026-03")

    def test_prepare_comparison_can_compare_illumia_eon_and_cve_together(self):
        prepared = services.prepare_comparison(
            service_data(
                providers=["ILLUMIA", "EON", "CVE"],
                offer_var_choice_eon="E.ON Flex Luce Casa",
                offer_fix_choice_eon="E.ON Luce Tua",
                tax_annual_consumption="1200",
            )
        )
        self.assertEqual(prepared["calc"]["providers_label"], "Illumia + E.ON + CVE")
        self.assertEqual(len(prepared["columns"]), 7)
        self.assertIn("CVE Variabile", [column["label"] for column in prepared["columns"]])
        cve = next(result for result in prepared["calc"]["provider_results"] if result["provider"] == "CVE")
        self.assertEqual(cve["offer_var"], "CVE 1Casa Smart Luce")
        self.assertEqual(cve["offer_fix"], "")

    def test_prepare_comparison_uses_selected_eon_fixed_gas_offer(self):
        data = service_data(
            provider="EON",
            commodity="GAS",
            offer_var_choice="E.ON Flex Gas",
            offer_fix_choice="E.ON Gas Tua",
        )
        prepared = services.prepare_comparison(data)
        self.assertEqual(prepared["calc"]["provider_label"], "E.ON")
        self.assertEqual(prepared["calc"]["offer_var"], "E.ON Flex Gas")
        self.assertEqual(prepared["calc"]["offer_fix"], "E.ON Gas Tua")
        self.assertEqual(prepared["calc"]["offer_valid_to"], date(2026, 6, 25))
        self.assertEqual(prepared["calc"]["bill_offer_expiry"], date(2026, 12, 31))
        self.assertEqual(prepared["calc"]["bill_offer_expiry_label"], "31/12/2026")
        self.assertEqual(prepared["calc"]["offer_expiry_label"], "31/12/2026")
        self.assertEqual(prepared["values"]["variabile"]["sconti"], -10)
        self.assertEqual(prepared["values"]["fissa"]["sconti"], -10)
        self.assertAlmostEqual(prepared["values"]["fissa"]["vendita_consumo"], 57.8, places=4)


class BillParserTests(SimpleTestCase):
    def test_parses_gas_bill_with_textual_months(self):
        parsed = parse_bill_text(
            """
            GAS NATURALE
            MERCATO LIBERO
            MASSIMO PASSARELLA
            VIA SIGNORIA 79
            Periodo oggetto di fatturazione: 1 ottobre 2025 - 31 ottobre 2025
            Consumo totale fatturato: 21,41202 Smc
            Consumo da inizio contratto (mc): 564
            Scadenza condizioni economiche: 30/09/2027
            SCONTRINO DELL'ENERGIA
            Codice PDR: 15351410010036
            QUOTA PER CONSUMI
            21,412020 Smc 0,714552 €/Smc 15,30
            di cui spesa per vendita gas naturale 0,487110 €/Smc 10,43
            di cui spesa per la rete e gli oneri generali di sistema 0,227442 €/Smc 4,87
            QUOTA FISSA
            1 mesi 16,330000 €/mese 16,33
            di cui spesa per vendita gas naturale 9,000000 €/mese 9,00
            di cui spesa per la rete e gli oneri generali di sistema 7,330000 €/mese 7,33
            Accise e IVA 12,07
            """
        )
        self.assertEqual(parsed.values["commodity"], "GAS")
        self.assertEqual(parsed.values["pod_pdr"], "15351410010036")
        self.assertEqual(parsed.values["bill_start"], date(2025, 10, 1))
        self.assertEqual(parsed.values["bill_end"], date(2025, 10, 31))
        self.assertAlmostEqual(parsed.values["b_rete_fissa"], 7.33)

    def test_parser_keeps_fixed_bill_amounts_for_multi_month_periods(self):
        parsed = parse_bill_text(
            """
            GAS NATURALE
            MERCATO LIBERO
            MARIO ROSSI
            VIA ROMA 10
            Periodo oggetto di fatturazione: 1 gennaio 2026 - 28 febbraio 2026
            Consumo totale fatturato: 100 Smc
            Consumo da inizio contratto (mc): 1200
            Scadenza condizioni economiche: 31/12/2026
            SCONTRINO DELL'ENERGIA
            Codice PDR: 15351410010036
            QUOTA PER CONSUMI
            100 Smc 0,400000 €/Smc 40,00
            di cui spesa per vendita gas naturale 0,250000 €/Smc 25,00
            di cui spesa per la rete e gli oneri generali di sistema 0,150000 €/Smc 15,00
            QUOTA FISSA
            2 mesi 5,000000 €/mese 10,00
            di cui spesa per vendita gas naturale 6,000000
            di cui spesa per la rete e gli oneri generali di sistema 4,000000
            Accise e IVA 12,07
            """
        )
        self.assertEqual(parsed.values["bill_start"], date(2026, 1, 1))
        self.assertEqual(parsed.values["bill_end"], date(2026, 2, 28))
        self.assertAlmostEqual(parsed.values["b_vendita_fissa"], 6.0)
        self.assertAlmostEqual(parsed.values["b_rete_fissa"], 4.0)


class ConfrontoFormTests(SimpleTestCase):
    def test_dashboard_context_fields_are_required_and_not_prefilled(self):
        form = ConfrontoForm()
        self.assertIsNone(form.fields["nome_cliente"].initial)
        self.assertFalse(form.fields["pod_pdr"].required)
        self.assertEqual(form.fields["segmento"].choices[0], ("", "Seleziona segmento"))
        self.assertEqual(form.fields["commodity"].choices[0], ("", "Seleziona fornitura"))
        self.assertEqual(form.fields["bill_tariff_type"].choices[0], ("", "Seleziona tariffa"))
        self.assertEqual(form.fields["providers"].choices[0], ("ILLUMIA", "Illumia"))
        self.assertIn(("CVE", "CVE"), form.fields["providers"].choices)
        self.assertFalse(form.fields["cve_over70"].required)
        self.assertIsNone(form.fields["tariff_selection_mode"].initial)
        self.assertIsNone(form.fields["bill_start"].initial)
        self.assertIsNone(form.fields["bill_end"].initial)
        self.assertIsNone(form.fields["bill_offer_expiry"].initial)
        self.assertIsNone(form.fields["tax_annual_consumption"].initial)

        invalid = ConfrontoForm(
            valid_payload(
                nome_cliente="",
                segmento="",
                commodity="",
                bill_tariff_type="",
                tariff_selection_mode="",
                providers=[],
                tax_annual_consumption="",
                bill_start="",
                bill_end="",
                bill_offer_expiry="",
            )
        )
        self.assertFalse(invalid.is_valid())
        for field in [
            "nome_cliente",
            "segmento",
            "commodity",
            "bill_tariff_type",
            "tariff_selection_mode",
            "providers",
            "bill_start",
            "bill_offer_expiry",
            "tax_annual_consumption",
        ]:
            self.assertIn(field, invalid.errors)

    def test_month_fields_are_normalized_to_full_bill_period(self):
        form = ConfrontoForm(valid_payload())
        self.assertTrue(form.is_valid(), form.errors.as_data())
        self.assertEqual(form.cleaned_data["bill_start"], date(2026, 1, 1))
        self.assertEqual(form.cleaned_data["bill_end"], date(2026, 3, 31))
        self.assertEqual(form.cleaned_data["bill_offer_expiry"], date(2026, 12, 31))
        self.assertEqual(form.cleaned_data["tariff_selection_mode"], "LATEST")
        self.assertEqual(float(form.cleaned_data["b_bonus_sociale"]), -21.6)

    def test_missing_end_month_defaults_to_start_month(self):
        form = ConfrontoForm(valid_payload(bill_start="2026-02", bill_end=""))
        self.assertTrue(form.is_valid(), form.errors.as_data())
        self.assertEqual(form.cleaned_data["bill_start"], date(2026, 2, 1))
        self.assertEqual(form.cleaned_data["bill_end"], date(2026, 2, 28))

    def test_bonus_sociale_is_optional(self):
        form = ConfrontoForm(valid_payload(b_bonus_sociale=""))
        self.assertTrue(form.is_valid(), form.errors.as_data())
        self.assertEqual(form.cleaned_data["b_bonus_sociale"], 0)

    def test_customer_invite_form_normalizes_whatsapp_phone(self):
        form = CustomerInviteForm(
            {
                "customer_name": "Mario Rossi",
                "customer_phone": "333 123 4567",
                "whatsapp_ready": "on",
            }
        )
        self.assertTrue(form.is_valid(), form.errors.as_data())
        self.assertEqual(form.cleaned_data["customer_phone"], "393331234567")
        self.assertEqual(form.cleaned_data["customer_phone_display"], "+393331234567")

    def test_customer_invite_form_requires_whatsapp_confirmation(self):
        form = CustomerInviteForm({"customer_name": "Mario Rossi", "customer_phone": "333 123 4567"})
        self.assertFalse(form.is_valid())
        self.assertIn("whatsapp_ready", form.errors)

    def test_italian_decimal_commas_are_accepted(self):
        form = ConfrontoForm(
            valid_payload(
                b_vendita_consumo="85,520",
                b_rete_consumi="24,61",
                b_vendita_fissa="12,105",
                b_rete_fissa="1,9",
                b_accise_iva="23,97",
            )
        )
        self.assertTrue(form.is_valid(), form.errors.as_data())
        self.assertEqual(float(form.cleaned_data["b_vendita_consumo"]), 85.52)
        self.assertEqual(float(form.cleaned_data["b_rete_consumi"]), 24.61)

    def test_service_data_keeps_fixed_bill_amounts_as_entered(self):
        form = ConfrontoForm(
            valid_payload(
                bill_start="2026-01",
                bill_end="2026-02",
                b_vendita_fissa="10",
                b_rete_fissa="4",
            )
        )
        self.assertTrue(form.is_valid(), form.errors.as_data())
        data = form.service_data()
        self.assertEqual(data["b_vendita_fissa"], 10)
        self.assertEqual(data["b_rete_fissa"], 4)

    def test_gas_disables_and_zeros_power_fields(self):
        form = ConfrontoForm(
            valid_payload(
                commodity="GAS",
                tax_power_kw="9",
                b_quota_potenza="99",
            )
        )
        self.assertTrue(form.fields["tax_power_kw"].disabled)
        self.assertTrue(form.fields["b_quota_potenza"].disabled)
        self.assertTrue(form.is_valid(), form.errors.as_data())
        self.assertEqual(form.cleaned_data["tax_power_kw"], 0)
        self.assertEqual(form.cleaned_data["b_quota_potenza"], 0)

    def test_end_month_before_start_month_is_rejected(self):
        form = ConfrontoForm(valid_payload(bill_start="2026-03", bill_end="2026-01"))
        self.assertFalse(form.is_valid())
        self.assertIn("Il mese finale non può essere precedente al mese iniziale.", form.non_field_errors())

    def test_customer_mode_allows_blank_offer_expiry_and_keeps_contact_fields(self):
        form = ConfrontoForm(
            valid_payload(
                bill_offer_expiry="",
                email_cliente="mario@example.com",
                telefono_cliente="3331234567",
            ),
            customer_mode=True,
        )
        self.assertFalse(isinstance(form.fields["email_cliente"].widget, forms.HiddenInput))
        self.assertFalse(isinstance(form.fields["telefono_cliente"].widget, forms.HiddenInput))
        self.assertFalse(form.fields["bill_offer_expiry"].required)
        self.assertTrue(form.is_valid(), form.errors.as_data())
        self.assertIsNone(form.cleaned_data["bill_offer_expiry"])
        self.assertEqual(form.cleaned_data["email_cliente"], "mario@example.com")
        self.assertEqual(form.cleaned_data["telefono_cliente"], "3331234567")

    def test_customer_mode_forces_illumia_even_if_provider_is_tampered(self):
        form = ConfrontoForm(
            valid_payload(
                pod_pdr="IT001E99999999",
                providers=["CVE"],
                tariff_selection_mode="PERIOD",
                offer_var_choice_eon="Qualsiasi offerta",
                offer_fix_choice_eon="Qualsiasi offerta",
                offer_var_choice_cve="CVE 1Casa Smart Luce",
                offer_fix_choice_cve="Qualsiasi offerta",
                cve_over70="on",
            ),
            customer_mode=True,
        )
        self.assertTrue(form.is_valid(), form.errors.as_data())
        self.assertEqual(form.cleaned_data["pod_pdr"], "")
        self.assertEqual(form.cleaned_data["providers"], ["ILLUMIA"])
        self.assertEqual(form.cleaned_data["provider"], "ILLUMIA")
        self.assertEqual(form.cleaned_data["tariff_selection_mode"], "LATEST")
        self.assertEqual(form.cleaned_data["offer_var_choice_eon"], "")
        self.assertEqual(form.cleaned_data["offer_fix_choice_eon"], "")
        self.assertEqual(form.cleaned_data["offer_var_choice_cve"], "")
        self.assertEqual(form.cleaned_data["offer_fix_choice_cve"], "")
        self.assertFalse(form.cleaned_data["cve_over70"])


    def test_operator_mode_forces_illumia_latest_but_keeps_pod(self):
        form = ConfrontoForm(
            valid_payload(
                pod_pdr="IT001E99999999",
                providers=["EON", "CVE"],
                tariff_selection_mode="PERIOD",
                offer_var_choice_eon="E.ON Flex Luce Casa",
                offer_fix_choice_eon="E.ON Luce Tua",
                offer_var_choice_cve="CVE 1Casa Smart Luce",
                offer_fix_choice_cve="Qualsiasi offerta",
                cve_over70="on",
            ),
            operator_mode=True,
        )
        self.assertTrue(isinstance(form.fields["providers"].widget, forms.HiddenInput))
        self.assertTrue(isinstance(form.fields["tariff_selection_mode"].widget, forms.HiddenInput))
        self.assertTrue(form.is_valid(), form.errors.as_data())
        self.assertEqual(form.cleaned_data["pod_pdr"], "IT001E99999999")
        self.assertEqual(form.cleaned_data["providers"], ["ILLUMIA"])
        self.assertEqual(form.cleaned_data["provider"], "ILLUMIA")
        self.assertEqual(form.cleaned_data["tariff_selection_mode"], "LATEST")
        self.assertEqual(form.cleaned_data["offer_var_choice_eon"], "")
        self.assertEqual(form.cleaned_data["offer_fix_choice_eon"], "")
        self.assertEqual(form.cleaned_data["offer_var_choice_cve"], "")
        self.assertEqual(form.cleaned_data["offer_fix_choice_cve"], "")
        self.assertFalse(form.cleaned_data["cve_over70"])


class InviteCommandTests(TestCase):
    def test_crea_invito_creates_single_code_with_label(self):
        output = StringIO()
        call_command("crea_invito", label="Cliente Mario", stdout=output)

        invites = list(InviteCode.objects.all())
        self.assertEqual(len(invites), 1)
        self.assertEqual(invites[0].label, "Cliente Mario")
        self.assertIn(invites[0].code, output.getvalue())

    def test_crea_invito_can_create_multiple_codes(self):
        output = StringIO()
        call_command("crea_invito", label="Campagna", count=3, stdout=output)

        invites = list(InviteCode.objects.order_by("created_at"))
        self.assertEqual(len(invites), 3)
        self.assertEqual([invite.label for invite in invites], ["Campagna 1", "Campagna 2", "Campagna 3"])
        lines = [line.strip() for line in output.getvalue().splitlines() if line.strip()]
        self.assertEqual(len(lines), 3)


    @patch.dict(
        os.environ,
        {
            "DJANGO_OPERATOR_USERNAME": "operatore.illumia@example.com",
            "DJANGO_OPERATOR_EMAIL": "operatore.illumia@example.com",
            "DJANGO_OPERATOR_PASSWORD": "OperatorTestPassword123!",
            "DJANGO_OPERATOR_FIRST_NAME": "Operatore",
            "DJANGO_OPERATOR_LAST_NAME": "Illumia",
        },
        clear=False,
    )
    def test_create_initial_superuser_can_create_illumia_operator(self):
        output = StringIO()
        call_command("create_initial_superuser", stdout=output)

        User = get_user_model()
        user = User.objects.get(username="operatore.illumia@example.com")
        self.assertEqual(user.email, "operatore.illumia@example.com")
        self.assertEqual(user.first_name, "Operatore")
        self.assertEqual(user.last_name, "Illumia")
        self.assertTrue(user.is_staff)
        self.assertFalse(user.is_superuser)
        self.assertTrue(user.check_password("OperatorTestPassword123!"))
        self.assertTrue(user.groups.filter(name=ILLUMIA_OPERATOR_GROUP).exists())
        self.assertIn("Initial Illumia operator created: operatore.illumia@example.com", output.getvalue())


    @patch.dict(
        os.environ,
        {
            "DJANGO_OPERATOR_USERNAME": "operatore.illumia@example.com",
            "DJANGO_OPERATOR_EMAIL": "operatore.illumia@example.com",
            "DJANGO_OPERATOR_PASSWORD": "OperatorTestPassword123!",
            "DJANGO_OPERATOR_FIRST_NAME": "Operatore",
            "DJANGO_OPERATOR_LAST_NAME": "Illumia",
        },
        clear=False,
    )
    def test_create_initial_superuser_updates_existing_illumia_operator_password(self):
        User = get_user_model()
        existing = User.objects.create_user(
            username="operatore.illumia@example.com",
            email="vecchia@example.com",
            password="OldPassword123!",
        )

        output = StringIO()
        call_command("create_initial_superuser", stdout=output)

        existing.refresh_from_db()
        self.assertEqual(existing.email, "operatore.illumia@example.com")
        self.assertEqual(existing.first_name, "Operatore")
        self.assertEqual(existing.last_name, "Illumia")
        self.assertTrue(existing.is_staff)
        self.assertFalse(existing.is_superuser)
        self.assertFalse(existing.check_password("OldPassword123!"))
        self.assertTrue(existing.check_password("OperatorTestPassword123!"))
        self.assertTrue(existing.groups.filter(name=ILLUMIA_OPERATOR_GROUP).exists())
        self.assertIn("Initial Illumia operator already exists: operatore.illumia@example.com", output.getvalue())

    def test_stato_clienti_lists_customers_and_invite_status(self):
        customer = get_user_model().objects.create_user(
            username="cliente@example.com",
            email="cliente@example.com",
            password="secret",
            first_name="Cliente",
            last_name="Uno",
        )
        get_user_model().objects.create_superuser(
            username="admin@example.com",
            email="admin@example.com",
            password="secret",
        )
        available = InviteCode.objects.create(code="DISPONIB1", label="Disponibile")
        used = InviteCode.objects.create(code="USATO12345", label="Usato")
        used.mark_used(customer)

        output = StringIO()
        call_command("stato_clienti", stdout=output)
        rendered = output.getvalue()

        self.assertIn("=== CLIENTI REGISTRATI ===", rendered)
        self.assertIn("cliente@example.com", rendered)
        self.assertNotIn("admin@example.com", rendered)
        self.assertIn("Totale clienti registrati: 1", rendered)
        self.assertIn("=== CODICI DISPONIBILI ===", rendered)
        self.assertIn(available.code, rendered)
        self.assertIn("Totale codici disponibili: 1", rendered)
        self.assertIn("=== CODICI USATI ===", rendered)
        self.assertIn(used.code, rendered)
        self.assertIn("usato_da=cliente@example.com", rendered)
        self.assertIn("Totale codici usati: 1", rendered)

    def test_stato_clienti_handles_empty_state(self):
        output = StringIO()
        call_command("stato_clienti", stdout=output)
        rendered = output.getvalue()

        self.assertIn("Totale clienti registrati: 0", rendered)
        self.assertIn("Totale codici disponibili: 0", rendered)
        self.assertIn("Totale codici usati: 0", rendered)


class ConfrontoViewTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.staff_user = User.objects.create_user(
            username="staff@example.com",
            email="staff@example.com",
            password="secret",
            is_staff=True,
            is_superuser=True,
        )
        self.customer_user = User.objects.create_user(
            username="cliente-base@example.com",
            email="cliente-base@example.com",
            password="secret",
            first_name="Cliente",
            last_name="Base",
        )
        self.operator_user = User.objects.create_user(
            username="operatore@example.com",
            email="operatore@example.com",
            password="secret",
            is_staff=True,
        )
        operator_group = Group.objects.create(name=ILLUMIA_OPERATOR_GROUP)
        self.operator_user.groups.add(operator_group)
        self.client = Client()

    def login(self):
        self.assertTrue(self.client.login(username="staff@example.com", password="secret"))

    def login_customer(self):
        self.assertTrue(self.client.login(username="cliente-base@example.com", password="secret"))

    def login_operator(self):
        self.assertTrue(self.client.login(username="operatore@example.com", password="secret"))

    def test_root_shows_homepage_and_internal_login_for_anonymous_users(self):
        response = self.client.get("/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Energia Solidale")
        self.assertContains(response, "Piattaforma di confronto")
        self.assertContains(response, "Dashboard Amministrazione (Interna)")
        self.assertContains(response, "Confronto Bollette")
        self.assertContains(response, "Mostra Dashboard Clienti")
        self.assertContains(response, "Mostra Registrazione Clienti")
        self.assertContains(response, "Stato Clienti")
        self.assertContains(response, "Genera Codici")
        self.assertContains(response, "Confronto Bollette/Offerte")
        self.assertContains(response, 'href="/area-clienti/"')
        self.assertContains(response, 'href="/accounts/register/"')
        self.assertContains(response, 'href="/accounts/login/?next=%2Farea-clienti%2F"')
        self.assertContains(response, 'href="/accounts/login/?next=%2F%3Fpanel%3Dstatus-clienti"')
        self.assertContains(response, 'href="/accounts/login/?next=%2F%3Fpanel%3Dgenera-codici"')
        self.assertContains(response, 'href="/accounts/login/?next=%2F%3Fpanel%3Dconfronto"')

    def test_internal_status_tab_opens_status_section_after_login(self):
        self.login()
        response = self.client.get("/?panel=status-clienti")
        self.assertEqual(response.status_code, 200)
        self.assertIsNotNone(response.context["customer_status"])
        self.assertContains(response, 'id="stato-clienti"')

    def test_customer_access_page_shows_public_access_choices_for_anonymous_users(self):
        response = self.client.get("/area-clienti/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'href="/area-clienti/"')
        self.assertContains(response, 'class="topbar-brand"')
        self.assertContains(response, "Accedi")
        self.assertContains(response, "Registrati con codice")
        self.assertContains(response, 'href="/accounts/login/"')
        self.assertContains(response, 'href="/accounts/register/"')

    def test_root_redirects_customer_users_to_customer_area(self):
        self.login_customer()
        response = self.client.get("/")
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], "/area-clienti/")

    def test_authenticated_staff_can_open_registration_page(self):
        self.login()
        response = self.client.get("/accounts/register/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Crea le tue credenziali")
        self.assertContains(response, 'href="/"')

    def test_customer_dashboard_logo_points_to_customer_area(self):
        self.login_customer()
        response = self.client.get("/area-clienti/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'class="topbar-brand"')
        self.assertContains(response, 'href="/area-clienti/"')
        self.assertNotContains(response, 'href="/"')

    def test_authenticated_customer_is_redirected_from_registration_page(self):
        self.login_customer()
        response = self.client.get("/accounts/register/")
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], "/area-clienti/")

    def test_customer_area_requires_login(self):
        response = self.client.get("/area-clienti/confronto-illumia/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_staff_login_redirects_to_internal_dashboard_by_default(self):
        response = self.client.post(
            "/accounts/login/",
            {
                "username": "staff@example.com",
                "password": "secret",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], "/")

    def test_registration_creates_customer_and_redirects_to_reserved_area(self):
        invite = InviteCode.objects.create(code="INVITO1234", label="Cliente Luca")
        response = self.client.post(
            "/accounts/register/",
            {
                "invite_code": invite.code,
                "first_name": "Luca",
                "last_name": "Bianchi",
                "email": "luca@example.com",
                "password1": "PasswordSicura123!",
                "password2": "PasswordSicura123!",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], "/area-clienti/")
        user = get_user_model().objects.get(username="luca@example.com")
        self.assertEqual(user.email, "luca@example.com")
        self.assertEqual(int(self.client.session["_auth_user_id"]), user.pk)
        invite.refresh_from_db()
        self.assertFalse(invite.is_active)
        self.assertEqual(invite.used_by, user)
        self.assertIsNotNone(invite.used_at)

    def test_staff_can_create_customer_from_registration_page_without_losing_admin_session(self):
        self.login()
        invite = InviteCode.objects.create(code="INVITO9999", label="Cliente Admin")
        response = self.client.post(
            "/accounts/register/",
            {
                "invite_code": invite.code,
                "first_name": "Giulia",
                "last_name": "Verdi",
                "email": "giulia@example.com",
                "password1": "PasswordSicura123!",
                "password2": "PasswordSicura123!",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Account cliente creato per giulia@example.com")
        created_user = get_user_model().objects.get(username="giulia@example.com")
        self.assertEqual(int(self.client.session["_auth_user_id"]), self.staff_user.pk)
        invite.refresh_from_db()
        self.assertEqual(invite.used_by, created_user)
        self.assertFalse(invite.is_active)

    def test_registration_requires_valid_invite_code(self):
        response = self.client.post(
            "/accounts/register/",
            {
                "invite_code": "NONVALIDO",
                "first_name": "Luca",
                "last_name": "Bianchi",
                "email": "luca@example.com",
                "password1": "PasswordSicura123!",
                "password2": "PasswordSicura123!",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Codice invito non valido.")
        self.assertFalse(get_user_model().objects.filter(username="luca@example.com").exists())

    def test_registration_rejects_already_used_invite_code(self):
        used_invite = InviteCode.objects.create(code="USATO1234", label="Invito usato")
        first_user = get_user_model().objects.create_user(username="usato@example.com", password="secret")
        used_invite.mark_used(first_user)
        response = self.client.post(
            "/accounts/register/",
            {
                "invite_code": used_invite.code,
                "first_name": "Nuovo",
                "last_name": "Cliente",
                "email": "nuovo@example.com",
                "password1": "PasswordSicura123!",
                "password2": "PasswordSicura123!",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Questo codice invito e gia stato utilizzato.")
        self.assertFalse(get_user_model().objects.filter(username="nuovo@example.com").exists())

    def test_form_renders_month_selectors_and_reset_button(self):
        self.login()
        response = self.client.get("/")
        self.assertNotContains(response, 'type="month"')
        self.assertNotContains(response, "Build locale CAP-FISCALE 2026-06-06")
        self.assertContains(response, "Esegui confronto")
        self.assertContains(response, "Genera Codice invito")
        self.assertContains(response, "Stato clienti")
        self.assertContains(response, "Apri dashboard cliente")
        self.assertContains(response, "Apri registrazione cliente")
        self.assertContains(response, 'href="/?panel=confronto#confronto-bollette-offerte"')
        self.assertContains(response, 'href="/?panel=genera-codici#genera-codici"')
        self.assertContains(response, 'href="/?panel=status-clienti#stato-clienti"')
        self.assertContains(response, 'href="/area-clienti/"')
        self.assertContains(response, 'href="/accounts/register/"')
        self.assertContains(response, "Vendita fissa (totale bolletta)")
        self.assertContains(response, "Rete/oneri fissa (totale bolletta)")
        self.assertContains(response, 'data-month-field="bill_start"')
        self.assertContains(response, 'data-month-field="bill_end"')
        self.assertContains(response, 'data-month-part="year"')
        self.assertContains(response, 'type="date"')
        self.assertContains(response, "Nuova bolletta")
        self.assertContains(response, "Fornitori confronto")
        self.assertContains(response, 'data-power-field')
        self.assertContains(response, 'togglePowerFields')
        self.assertContains(response, 'commodity.value === "EE"')
        self.assertContains(response, 'data-cve-over70-field hidden')
        self.assertContains(response, "Tariffa CVE Over 70")
        self.assertContains(response, 'activeProviders.includes("CVE")')
        self.assertContains(response, "CVE - Offerta variabile")
        self.assertContains(response, "Importa bolletta PDF")

    def test_internal_invite_tab_shows_whatsapp_tools(self):
        self.login()
        response = self.client.get("/?panel=genera-codici")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Genera Codice invito")
        self.assertContains(response, "Apri/Riapri WhatsApp Web")
        self.assertContains(response, "Il codice non verra creato finche non lo confermi")
        self.assertContains(response, "Genera codice e apri WhatsApp Web")
        self.assertContains(response, "3271044102")
        self.assertNotContains(response, "Importa bolletta PDF")


    def test_operator_dashboard_is_limited_to_illumia_latest_tariffs(self):
        self.login_operator()
        response = self.client.get("/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Confronto bollette vs Illumia")
        self.assertContains(response, "Fornitore abilitato:</strong> Illumia")
        self.assertContains(response, "Illumia - Offerta variabile")
        self.assertContains(response, "Illumia - Offerta fissa")
        self.assertNotContains(response, "Fornitori confronto")
        self.assertNotContains(response, "Logica tariffe confronto")
        self.assertNotContains(response, "Tariffe del periodo bolletta")
        self.assertNotContains(response, "E.ON - Offerta variabile")
        self.assertNotContains(response, "CVE - Offerta variabile")
        self.assertNotContains(response, "Tariffa CVE Over 70")
        self.assertNotContains(response, "Genera Codice invito")
        self.assertNotContains(response, "Stato clienti")
        self.assertNotContains(response, "Apri dashboard cliente")

    def test_operator_post_is_forced_to_illumia_latest_even_if_tampered(self):
        self.login_operator()
        response = self.client.post(
            "/",
            valid_payload(
                providers=["EON", "CVE"],
                tariff_selection_mode="PERIOD",
                offer_var_choice_eon="E.ON Flex Luce Casa",
                offer_fix_choice_eon="E.ON Luce Tua",
                offer_var_choice_cve="CVE 1Casa Smart Luce",
                cve_over70="on",
            ),
        )
        self.assertEqual(response.status_code, 200)
        html = response.content.decode()
        self.assertIn("Fornitori confronto:</strong> Illumia", html)
        self.assertIn("Illumia Variabile", html)
        self.assertIn("Illumia Fissa", html)
        self.assertNotIn("E.ON Variabile", html)
        self.assertNotIn("CVE Variabile", html)
        self.assertNotIn("Logica tariffe:</strong>", html)
        self.assertEqual(self.client.session["last_confronto"]["providers"], ["ILLUMIA"])
        self.assertEqual(self.client.session["last_confronto"]["provider"], "ILLUMIA")
        self.assertEqual(self.client.session["last_confronto"]["tariff_selection_mode"], "LATEST")

    def test_customer_page_hides_provider_picker_and_shows_illumia_lock(self):
        self.login()
        response = self.client.get("/area-clienti/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Confronto bolletta con Illumia")
        self.assertContains(response, "1. Carica la bolletta")
        self.assertContains(response, "2. Controlla i dati letti")
        self.assertContains(response, "3. Solo se manca in bolletta")
        self.assertContains(response, "4. Calcola confronto")
        self.assertContains(response, "Obbligatorio")
        self.assertNotContains(response, "Build locale CAP-FISCALE 2026-06-06")
        self.assertContains(response, "Email")
        self.assertContains(response, "Telefono")
        self.assertNotContains(response, "Fornitore confronto")
        self.assertNotContains(response, "Codice POD/PDR")
        self.assertNotContains(response, "Logica tariffe confronto")
        self.assertNotContains(response, "E.ON - Offerta variabile")
        self.assertNotContains(response, "CVE - Offerta variabile")
        self.assertNotContains(response, "Fornitori confronto")
        self.assertNotContains(response, "Strumenti clienti")
        self.assertNotContains(response, "Stato clienti")
        self.assertNotContains(response, "Genera codice e apri WhatsApp Web")

    def test_legacy_customer_dashboard_url_redirects_to_customer_area(self):
        self.login_customer()
        response = self.client.get("/area-clienti/confronto-illumia/")
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], "/area-clienti/")

    def test_internal_dashboard_generates_whatsapp_web_invite(self):
        self.login()
        response = self.client.post(
            "/",
            {
                "action": "send_customer_invite",
                "customer_name": "Mario Rossi",
                "customer_phone": "3331234567",
                "whatsapp_ready": "on",
            },
        )
        self.assertEqual(response.status_code, 200)
        invite = InviteCode.objects.get()
        self.assertEqual(invite.label, "Mario Rossi")
        self.assertIn("+393331234567", invite.note)
        self.assertContains(response, "Codice creato:")
        self.assertContains(response, invite.code)
        self.assertContains(response, "https://web.whatsapp.com/send?phone=393331234567")
        self.assertContains(response, f"https://energia-solidale.onrender.com/accounts/register/?invite_code={invite.code}")

    def test_internal_dashboard_does_not_generate_invite_without_whatsapp_confirmation(self):
        self.login()
        response = self.client.post(
            "/",
            {
                "action": "send_customer_invite",
                "customer_name": "Mario Rossi",
                "customer_phone": "3331234567",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(InviteCode.objects.count(), 0)
        self.assertContains(response, "Riapri WhatsApp Web e conferma prima di generare il codice.")

    def test_internal_dashboard_can_show_customer_status(self):
        self.login()
        customer = get_user_model().objects.create_user(
            username="cliente@example.com",
            email="cliente@example.com",
            password="secret",
            first_name="Cliente",
            last_name="Uno",
        )
        available_invite = InviteCode.objects.create(code="DISPONIB1", label="Disponibile")
        used_invite = InviteCode.objects.create(code="USATO12345", label="Usato")
        used_invite.mark_used(customer)

        response = self.client.post("/", {"action": "show_customer_status"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Clienti registrati")
        self.assertContains(response, "Codici disponibili")
        self.assertContains(response, "Codici usati")
        self.assertContains(response, "cliente@example.com")
        self.assertContains(response, available_invite.code)
        self.assertContains(response, used_invite.code)
        self.assertEqual(response.context["customer_status"]["user_count"], 2)
        self.assertEqual(response.context["customer_status"]["available_count"], 1)
        self.assertEqual(response.context["customer_status"]["used_count"], 1)

    def test_registration_page_prefills_invite_code_from_query_string(self):
        response = self.client.get("/accounts/register/?invite_code=invito-1234")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["form"]["invite_code"].value(), "INVITO1234")

    def test_invalid_form_renders_field_errors(self):
        self.login()
        response = self.client.post("/", valid_payload(tax_power_kw="0"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Il confronto non è partito")
        self.assertContains(response, "Indica la potenza impegnata")
        self.assertNotContains(response, "Scarica Excel")

    def test_calculate_comparison_stores_session_and_renders_period(self):
        self.login()
        response = self.client.post("/", valid_payload())
        self.assertEqual(response.status_code, 200)
        html = response.content.decode()
        self.assertIn("Gennaio 2026 - Marzo 2026", html)
        self.assertIn("Cliente:</strong> Mario Rossi", html)
        self.assertIn("Codice POD/PDR:</strong> IT001E12345678", html)
        self.assertIn("Confronto eseguito:</strong>", html)
        self.assertIn("Tipo tariffa bolletta:</strong> Variabile", html)
        self.assertIn("Consumo annuo stimato:</strong> 1200 kWh/anno", html)
        self.assertIn("Scadenza offerta bolletta:</strong> 31/12/2026", html)
        self.assertIn("Logica tariffe:</strong> Ultime tariffe disponibili", html)
        self.assertNotIn("periodo bolletta NON rientra", html)
        self.assertIn("TRIMESTRALE", html)
        self.assertIn('data-download-excel', html)
        self.assertIn("Scarica Excel", html)
        self.assertIn("last_confronto", self.client.session)
        self.assertEqual(self.client.session["last_confronto"]["bill_start"], "2026-01-01")
        self.assertEqual(self.client.session["last_confronto"]["bill_end"], "2026-03-31")
        self.assertEqual(self.client.session["last_confronto"]["bill_offer_expiry"], "2026-12-31")
        self.assertEqual(self.client.session["last_confronto"]["tariff_selection_mode"], "LATEST")

    def test_calculate_keeps_multi_month_fixed_bill_fields_as_entered_in_form(self):
        self.login()
        response = self.client.post(
            "/",
            valid_payload(
                bill_start="2026-01",
                bill_end="2026-02",
                b_vendita_fissa="10",
                b_rete_fissa="4",
            ),
        )
        self.assertEqual(response.status_code, 200)
        form = response.context["form"]
        self.assertEqual(form["b_vendita_fissa"].value(), 10)
        self.assertEqual(form["b_rete_fissa"].value(), 4)

    def test_calculate_renders_tax_cap_summary_for_capped_offer(self):
        self.login()
        response = self.client.post(
            "/",
            valid_payload(
                commodity="GAS",
                bill_tariff_type="VARIABILE",
                tax_power_kw="0",
                tax_annual_consumption="1200",
                consumo="100",
                b_vendita_consumo="40",
                b_vendita_fissa="10",
                b_rete_consumi="8",
                b_rete_fissa="1",
                b_quota_potenza="0",
                b_sconti="0",
                b_ricalcoli="0",
                b_bonus_sociale="0",
                b_arrotondamenti="0",
                b_servizi_accessori="0",
                b_accise_iva="5",
            ),
        )
        self.assertEqual(response.status_code, 200)
        html = response.content.decode()
        self.assertIn("Incidenza fiscale bolletta:</strong> 8,47%", html)
        self.assertIn("Cap fiscale applicato:</strong> Si", html)
        self.assertIn("Incidenza teorica senza cap:</strong>", html)
        self.assertNotIn("Quota Potenza", [row["voce"] for row in response.context["rows"]])

    def test_calculate_can_compare_selected_eon_offer(self):
        self.login()
        response = self.client.post(
            "/",
            valid_payload(
                provider="EON",
                offer_var_choice="E.ON Flex Luce Casa",
                offer_fix_choice="E.ON Luce Tua",
            ),
        )
        self.assertEqual(response.status_code, 200)
        html = response.content.decode()
        self.assertIn("Fornitori confronto:</strong> E.ON", html)
        self.assertIn("E.ON Flex Luce Casa", html)
        self.assertIn("E.ON Luce Tua", html)
        self.assertIn("E.ON Variabile", html)
        self.assertEqual(self.client.session["last_confronto"]["provider"], "EON")

    def test_calculate_can_compare_illumia_and_eon_together(self):
        self.login()
        response = self.client.post(
            "/",
            valid_payload(
                providers=["ILLUMIA", "EON"],
                offer_var_choice_eon="E.ON Flex Luce Casa",
                offer_fix_choice_eon="E.ON Luce Tua",
            ),
        )
        self.assertEqual(response.status_code, 200)
        html = response.content.decode()
        self.assertIn("Fornitori confronto:</strong> Illumia + E.ON", html)
        self.assertIn("Illumia Variabile", html)
        self.assertIn("Illumia Fissa", html)
        self.assertIn("E.ON Variabile", html)
        self.assertIn("E.ON Fissa", html)
        self.assertEqual(response.context["prepared"]["columns"][0]["label"], "Bolletta")
        self.assertEqual(len(response.context["prepared"]["columns"]), 5)
        self.assertEqual(self.client.session["last_confronto"]["providers"], "ILLUMIA,EON")

    def test_calculate_can_compare_cve_with_annual_band(self):
        self.login()
        response = self.client.post("/", valid_payload(provider="CVE", providers=["CVE"], tax_annual_consumption="500"))
        self.assertEqual(response.status_code, 200)
        html = response.content.decode()
        self.assertIn("Fornitori confronto:</strong> CVE", html)
        self.assertIn("CVE Over 70:</strong> No", html)
        self.assertIn("CVE 1Casa Small Luce", html)
        self.assertIn("CVE Variabile", html)
        self.assertIn("CVE Fissa", html)
        self.assertEqual(self.client.session["last_confronto"]["providers"], "CVE")

    def test_customer_area_always_calculates_with_illumia_only(self):
        self.login()
        response = self.client.post(
            "/area-clienti/",
            valid_payload(
                email_cliente="mario@example.com",
                telefono_cliente="3331234567",
                pod_pdr="IT001E99999999",
                bill_offer_expiry="",
                providers=["CVE"],
                tariff_selection_mode="PERIOD",
                offer_var_choice_eon="E.ON Flex Luce Casa",
                offer_fix_choice_eon="E.ON Luce Tua",
                offer_var_choice_cve="CVE 1Casa Over 70 Luce",
                cve_over70="on",
            ),
        )
        self.assertEqual(response.status_code, 200)
        html = response.content.decode()
        self.assertIn("Fornitori confronto:</strong> Illumia", html)
        self.assertIn("Illumia Variabile", html)
        self.assertIn('value="mario@example.com"', html)
        self.assertIn('value="3331234567"', html)
        self.assertIn("Scadenza offerta bolletta:</strong> N.D.", html)
        self.assertIn("customer-optional-panel", html)
        self.assertNotIn("Build locale CAP-FISCALE 2026-06-06", html)
        self.assertNotIn("Incidenza fiscale bolletta:</strong>", html)
        self.assertNotIn("Cap fiscale applicato:</strong>", html)
        self.assertNotIn("Incidenza teorica senza cap:</strong>", html)
        self.assertNotIn("Codice POD/PDR", html)
        self.assertNotIn("E.ON Variabile", html)
        self.assertNotIn("CVE Variabile", html)
        self.assertEqual(self.client.session["last_confronto_cliente_illumia"]["provider"], "ILLUMIA")
        self.assertEqual(self.client.session["last_confronto_cliente_illumia"]["providers"], ["ILLUMIA"])
        self.assertEqual(self.client.session["last_confronto_cliente_illumia"]["email_cliente"], "mario@example.com")
        self.assertEqual(self.client.session["last_confronto_cliente_illumia"]["telefono_cliente"], "3331234567")
        self.assertEqual(self.client.session["last_confronto_cliente_illumia"]["pod_pdr"], "")
        self.assertEqual(self.client.session["last_confronto_cliente_illumia"]["tariff_selection_mode"], "LATEST")

    def test_customer_page_hides_technical_fields_inside_optional_panel(self):
        self.login()
        response = self.client.get("/area-clienti/")
        html = response.content.decode()
        self.assertIn("Servizi extra", html)
        self.assertIn("Accise e IVA totali", html)
        self.assertIn("Quota potenza (solo luce)", html)
        self.assertIn("<details class=\"customer-optional-panel\"", html)
        self.assertNotIn("<details class=\"customer-optional-panel\" open", html)
    def test_change_bill_resets_bill_values_and_download_session(self):
        self.login()
        self.client.post("/", valid_payload())
        self.assertIn("last_confronto", self.client.session)
        session = self.client.session
        session["last_uploaded_bill_name"] = "bolletta.pdf"
        session.save()
        response = self.client.post(
            "/",
            valid_payload(action="reset_bill", b_vendita_consumo="999", tax_annual_consumption="9999"),
        )
        html = response.content.decode()
        self.assertEqual(response.status_code, 200)
        self.assertNotIn("last_confronto", self.client.session)
        self.assertNotIn("Scarica Excel", html)
        self.assertIn("Mario Rossi", html)
        self.assertNotIn("999", html)
        self.assertEqual(response.context["form"].initial["tax_annual_consumption"], "")
        self.assertNotIn("last_uploaded_bill_name", self.client.session)
        self.assertEqual(response.context["uploaded_bill_name"], "")

    @patch("confronti.views.parse_uploaded_bill")
    def test_pdf_upload_prefills_recognized_bill_values(self, mock_parse_uploaded_bill):
        mock_parse_uploaded_bill.return_value = ParsedBill(
            values={
                "nome_cliente": "Federico Boetto",
                "pod_pdr": "00881906523889",
                "commodity": "GAS",
                "consumo": 83,
                "tax_annual_consumption": 83,
                "b_vendita_consumo": 37.81,
            },
            warnings=["Data fine offerta bolletta non riconosciuta: compilala manualmente."],
        )
        self.login()
        response = self.client.post(
            "/",
            {
                "action": "extract_bill",
                "bill_pdf": SimpleUploadedFile("bolletta.pdf", b"%PDF-test", content_type="application/pdf"),
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Ho riportato nella dashboard 6 valori riconosciuti")
        self.assertContains(response, "Federico Boetto")
        self.assertContains(response, "00881906523889")
        self.assertContains(response, "Data fine offerta bolletta non riconosciuta")
        self.assertContains(response, "File caricato: bolletta.pdf")
        self.assertEqual(response.context["uploaded_bill_name"], "bolletta.pdf")
        self.assertEqual(self.client.session["last_uploaded_bill_name"], "bolletta.pdf")

    def test_download_requires_a_previous_comparison(self):
        self.login()
        response = self.client.get("/scarica-excel/")
        self.assertEqual(response.status_code, 400)

    def test_customer_download_requires_a_previous_comparison(self):
        self.login()
        response = self.client.get("/area-clienti/scarica-excel/")
        self.assertEqual(response.status_code, 400)

    def test_excel_download_contains_expected_labels_values_and_formulas(self):
        self.login()
        self.client.post("/", valid_payload(b_servizi_accessori="5", servizi_accessori_iva="10%"))
        response = self.client.get("/scarica-excel/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("confronto_illumia_Mario_Rossi_EE.xlsx", response["Content-Disposition"])
        self.assertEqual(response.content[:2], b"PK")

        wb = load_workbook(BytesIO(response.content), data_only=False)
        ws = wb["Confronto"]
        self.assertEqual(ws["A1"].value, "Mario Rossi")
        self.assertEqual(ws["F3"].value, "Scadenza offerta bolletta: 31/12/2026")
        self.assertTrue(ws["F3"].font.bold)
        self.assertEqual(ws["F3"].font.color.rgb[-6:], "B3261E")
        self.assertNotIn("scadenza", ws["F2"].value.lower())
        self.assertEqual(ws["F6"].value, "Tipo tariffa bolletta: Variabile")
        self.assertTrue(str(ws["F7"].value).startswith("Confronto eseguito: "))
        self.assertEqual(ws["F8"].value, "IVA servizi accessori: 10%")
        self.assertEqual(ws["F9"].value, "Consumo annuo stimato: 1200 kWh/anno")
        self.assertTrue(str(ws["F10"].value).startswith("Parametri Accise/IVA: "))
        self.assertEqual(ws["F11"].value, "Logica tariffe: Ultime tariffe disponibili")
        self.assertEqual(ws["F12"].value, "Codice POD/PDR: IT001E12345678")
        self.assertEqual(ws["F13"].value, "Fornitura: Luce")
        self.assertEqual(ws["F14"].value, "Periodo bolletta: Gennaio 2026 - Marzo 2026")
        self.assertEqual(ws["A12"].value, "Bonus Sociale")
        self.assertEqual(ws["A13"].value, "Arrotondamenti")
        self.assertEqual(ws["A14"].value, "Servizi accessori (IVA 10%)")
        self.assertEqual(ws["A15"].value, "Accise")
        self.assertEqual(ws["A16"].value, "IVA")
        self.assertEqual(ws["A17"].value, "Accise e Iva")
        self.assertEqual(ws["A18"].value, "Totale")
        self.assertEqual(ws["B12"].value, -21.6)
        self.assertEqual(ws["C12"].value, -21.6)
        self.assertEqual(ws["D12"].value, -21.6)
        self.assertEqual(ws["B14"].value, 5)
        self.assertEqual(ws["B15"].value, "N.D.")
        self.assertEqual(ws["B16"].value, "N.D.")
        self.assertEqual(ws["C17"].value, ws["C15"].value + ws["C16"].value)
        self.assertEqual(ws["C14"].value, 0)
        self.assertEqual(ws["D14"].value, 0)
        self.assertAlmostEqual(ws["B6"].value, 3.7, places=4)
        self.assertAlmostEqual(ws["B8"].value, 0.3967, places=4)
        self.assertIsInstance(ws["C15"].value, (int, float))
        self.assertIsInstance(ws["D15"].value, (int, float))
        self.assertEqual(ws["B17"].value, 12.12)
        self.assertEqual(ws["B18"].value, "=SUM(B4:B14)+B17")

    def test_gas_excel_download_removes_luce_only_rows(self):
        self.login()
        self.client.post(
            "/",
            valid_payload(
                commodity="GAS",
                tax_power_kw="0",
                b_quota_potenza="0",
                tax_annual_consumption="1200",
                consumo="100",
            ),
        )
        response = self.client.get("/scarica-excel/")
        self.assertEqual(response.status_code, 200)
        self.assertIn("confronto_illumia_Mario_Rossi_GAS.xlsx", response["Content-Disposition"])

        wb = load_workbook(BytesIO(response.content), data_only=False)
        ws = wb["Confronto"]
        labels = [ws[f"A{row}"].value for row in range(1, ws.max_row + 1)]
        self.assertNotIn("Vendita Fissa Luce", labels)
        self.assertNotIn("Quota Potenza", labels)
        self.assertIn("Vendita Fissa Gas", labels)
        self.assertEqual(ws["F13"].value, "Fornitura: Gas")

    def test_excel_download_contains_both_provider_columns(self):
        self.login()
        self.client.post(
            "/",
            valid_payload(
                providers=["ILLUMIA", "EON"],
                offer_var_choice_eon="E.ON Flex Luce Casa",
                offer_fix_choice_eon="E.ON Luce Tua",
            ),
        )
        response = self.client.get("/scarica-excel/")
        self.assertEqual(response.status_code, 200)
        self.assertIn("confronto_illumia_eon_Mario_Rossi_EE.xlsx", response["Content-Disposition"])

        wb = load_workbook(BytesIO(response.content), data_only=False)
        ws = wb["Confronto"]
        self.assertEqual(ws["B3"].value, "Bolletta")
        self.assertEqual(ws["C3"].value, "Illumia Variabile")
        self.assertEqual(ws["D3"].value, "Illumia Fissa")
        self.assertEqual(ws["E3"].value, "E.ON Variabile")
        self.assertEqual(ws["F3"].value, "E.ON Fissa")
        self.assertEqual(ws["H1"].value, "Fornitori confronto: Illumia + E.ON")
        self.assertIn("Illumia:", ws["H2"].value)
        self.assertIn("E.ON:", ws["H2"].value)
        self.assertEqual(ws["F18"].value, "=SUM(F4:F14)+F17")

    @patch("confronti.services.load_tariffe_file_for_segment_with_effective_segment", return_value=(None, "RESIDENZIALE"))
    def test_missing_illumia_offer_keeps_bill_and_marks_offers_nd(self, _mock_load_file):
        self.login()
        response = self.client.post("/", valid_payload())
        rows = response.context["rows"]
        self.assertTrue(all(row["variabile"] == "N.D." for row in rows))
        self.assertTrue(all(row["fissa"] == "N.D." for row in rows))
        self.assertNotEqual(next(row for row in rows if row["voce"] == "Totale")["bolletta"], "N.D.")
        self.assertNotEqual(next(row for row in rows if row["voce"] == "Accise e Iva")["bolletta"], "N.D.")
