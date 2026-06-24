import io
import re
from copy import copy
from datetime import date, datetime
from pathlib import Path

import openpyxl
from django.conf import settings
from django.utils import timezone
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


BASE_DIR = Path(settings.BASE_DIR)
TEMPLATE_XLSX = BASE_DIR / "esempio_confronto_corretto.xlsx"
TARIFFE_BASE = BASE_DIR / "tariffe"
EON_TARIFFE_DIR = BASE_DIR / "estrazioni_tariffe"
INDICI_XLSX = BASE_DIR / "indici_pun_psv_2025_2026.xlsx"
PROVIDERS = {
    "ILLUMIA": "Illumia",
    "EON": "E.ON",
    "CVE": "CVE",
}
SEGMENTS = ("RESIDENZIALE", "MICROBUSINESS", "BUSINESS")
EE_EXCISE_RATE = 0.0227
EE_RESIDENTIAL_EXEMPT_KWH_PER_MONTH = 150.0
GAS_VAT_REDUCED_ANNUAL_THRESHOLD = 480.0
GAS_EXCISE_BRACKETS = (
    (120.0, 0.044),
    (480.0, 0.175),
    (1560.0, 0.170),
    (None, 0.186),
)
GAS_REGIONAL_ADDITIONAL_MIN_RATES = {
    "Abruzzo": 0.02,
    "Basilicata": 0.02,
    "Calabria": 0.02,
    "Campania": 0.02,
    "Emilia-Romagna": 0.02,
    "Friuli-Venezia Giulia": 0.0,
    "Lazio": 0.02,
    "Liguria": 0.02,
    "Lombardia": 0.0,
    "Marche": 0.02,
    "Molise": 0.02,
    "Piemonte": 0.022,
    "Puglia": 0.02,
    "Sardegna": 0.0,
    "Sicilia": 0.0,
    "Toscana": 0.02,
    "Trentino-Alto Adige": 0.0,
    "Umbria": 0.02,
    "Valle d'Aosta": 0.0,
    "Veneto": 0.007747,
}
BILL_TARIFF_TYPE_LABELS = {
    "VARIABILE": "Variabile",
    "FISSA": "Fissa",
}
TARIFF_SELECTION_MODE_LABELS = {
    "LATEST": "Ultime tariffe disponibili",
    "PERIOD": "Tariffe del periodo bolletta",
}
ACCESSORY_SERVICES_VAT_RATES = {
    "22%": 0.22,
    "10%": 0.10,
    "Esente": 0.0,
}
ACCESSORY_SERVICES_VAT_OPTIONS = tuple(ACCESSORY_SERVICES_VAT_RATES)

KEYS = [
    "vendita_consumo",
    "vendita_fissa",
    "rete_consumi",
    "rete_fissa",
    "quota_potenza",
    "sconti",
    "ricalcoli",
    "bonus_sociale",
    "arrotondamenti",
    "servizi_accessori",
    "accise_iva",
]

MONTH_NAMES_IT = {
    1: "Gennaio",
    2: "Febbraio",
    3: "Marzo",
    4: "Aprile",
    5: "Maggio",
    6: "Giugno",
    7: "Luglio",
    8: "Agosto",
    9: "Settembre",
    10: "Ottobre",
    11: "Novembre",
    12: "Dicembre",
}


def clean_text(v) -> str:
    return "" if v is None else str(v).strip()


def normalize_provider(value) -> str:
    cleaned = clean_text(value or "ILLUMIA").upper().replace(".", "").replace("-", "").replace(" ", "")
    if cleaned == "EON":
        return "EON"
    if cleaned in {"CVE", "CENTROVENETOENERGIE", "CENTROVENETO"}:
        return "CVE"
    return "ILLUMIA"


def normalize_providers(value) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        parts = re.split(r"[,;|]+", value.replace("[", "").replace("]", "").replace("'", "").replace('"', ""))
    elif isinstance(value, (list, tuple, set)):
        parts = value
    else:
        parts = [value]
    out = []
    for part in parts:
        if not clean_text(part):
            continue
        provider = normalize_provider(part)
        if provider not in out:
            out.append(provider)
    return out


def provider_list_label(providers) -> str:
    normalized = normalize_providers(providers)
    return " + ".join(provider_label(provider) for provider in normalized) if normalized else "N.D."


def provider_label(value) -> str:
    return PROVIDERS.get(normalize_provider(value), "Illumia")


def bool_from_data(value) -> bool:
    if isinstance(value, bool):
        return value
    return clean_text(value).upper() in {"1", "TRUE", "SI", "SÌ", "YES", "ON"}


def normalize_bill_tariff_type(value) -> str:
    cleaned = clean_text(value).upper()
    return cleaned if cleaned in BILL_TARIFF_TYPE_LABELS else "VARIABILE"


def bill_tariff_type_label(value) -> str:
    return BILL_TARIFF_TYPE_LABELS.get(normalize_bill_tariff_type(value), "Variabile")


def normalize_tariff_selection_mode(value) -> str:
    cleaned = clean_text(value).upper()
    return cleaned if cleaned in TARIFF_SELECTION_MODE_LABELS else "LATEST"


def tariff_selection_mode_label(value) -> str:
    return TARIFF_SELECTION_MODE_LABELS[normalize_tariff_selection_mode(value)]


def normalize_accessory_services_vat_label(value) -> str:
    label = clean_text(value) or "22%"
    return label if label in ACCESSORY_SERVICES_VAT_RATES else "22%"


def normalize_primary_home(value) -> str:
    cleaned = clean_text(value).upper()
    return "NO" if cleaned in {"NO", "N", "0", "FALSE"} else "SI"


def primary_home_label(value) -> str:
    return "Sì" if normalize_primary_home(value) == "SI" else "No"


def normalize_region(value) -> str:
    cleaned = clean_text(value)
    return cleaned if cleaned in GAS_REGIONAL_ADDITIONAL_MIN_RATES else "Veneto"


def accessory_services_vat_rate(value) -> float:
    return float(ACCESSORY_SERVICES_VAT_RATES[normalize_accessory_services_vat_label(value)])


def progressive_tax_amount(consumption, brackets):
    consumption = max(0.0, float(consumption))
    total = 0.0
    lower = 0.0
    for upper, rate in brackets:
        if upper is None:
            taxable = max(0.0, consumption - lower)
        else:
            taxable = max(0.0, min(consumption, upper) - lower)
        total += taxable * float(rate)
        if upper is None or consumption <= upper:
            break
        lower = float(upper)
    return total


def annual_progressive_tax_for_period(period_consumption, annual_consumption, brackets):
    period_consumption = max(0.0, float(period_consumption))
    annual_consumption = max(0.0, float(annual_consumption))
    if period_consumption <= 0 or annual_consumption <= 0:
        return 0.0
    annual_tax = progressive_tax_amount(annual_consumption, brackets)
    return annual_tax * (period_consumption / annual_consumption)


def comparison_datetime_from_data(value=None):
    if isinstance(value, datetime):
        dt = value
    elif value:
        try:
            dt = datetime.fromisoformat(str(value))
        except ValueError:
            dt = timezone.localtime()
    else:
        dt = timezone.localtime()
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())
    return timezone.localtime(dt)


def comparison_datetime_label(value=None) -> str:
    return comparison_datetime_from_data(value).strftime("%d/%m/%Y %H:%M:%S")


def date_label_it(value) -> str:
    return value.strftime("%d/%m/%Y") if isinstance(value, date) else "N.D."


def safe_download_filename(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", str(name).strip())
    return cleaned or "confronto_bollette.xlsx"


def parse_number(x) -> float:
    if x is None:
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace("€", "").replace("\u20ac", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def parse_date_any(x):
    if x is None:
        return None
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x
    s = str(x).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def month_key_from_date(d: date) -> str:
    return f"{d.year:04d}-{d.month:02d}"


def billing_months_from_dates(start: date, end: date) -> int:
    if end < start:
        start, end = end, start
    months = (end.year - start.year) * 12 + (end.month - start.month) + 1
    return max(1, months)


def billing_divisor_from_months(months) -> float:
    months = max(1, int(months))
    return 12.0 / float(months)


def billing_label_from_months(months) -> str:
    labels = {1: "MENSILE", 2: "BIMESTRALE", 3: "TRIMESTRALE"}
    return labels.get(int(months), f"{int(months)} MESI")


def month_year_label(d: date) -> str:
    return f"{MONTH_NAMES_IT[d.month]} {d.year}"


def bill_period_label(start: date, end: date) -> str:
    if end < start:
        start, end = end, start
    if start.year == end.year and start.month == end.month:
        return month_year_label(start)
    return f"{month_year_label(start)} - {month_year_label(end)}"


def bill_period_outside_offer_validity(bill_start, bill_end, offer_valid_from, offer_valid_to) -> bool:
    if not offer_valid_from or not offer_valid_to:
        return False
    if bill_end < bill_start:
        bill_start, bill_end = bill_end, bill_start
    return bill_start < offer_valid_from or bill_end > offer_valid_to


def parse_index_month(x):
    if x is None:
        return ""
    if isinstance(x, datetime):
        return month_key_from_date(x.date())
    if isinstance(x, date):
        return month_key_from_date(x)
    s = str(x).strip()
    m = re.search(r"(\d{4})[-/](\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    parsed = parse_date_any(s)
    return month_key_from_date(parsed) if parsed else ""


def load_indici_rows(path: Path = INDICI_XLSX):
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "indici" not in wb.sheetnames:
        return []
    ws = wb["indici"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [clean_text(c).lower() for c in rows[0]]
    if not {"mese", "pun", "psv"}.issubset(set(header)):
        return []
    i_mese = header.index("mese")
    i_pun = header.index("pun")
    i_psv = header.index("psv")
    out = []
    for r in rows[1:]:
        if not r or not any(r):
            continue
        mese = parse_index_month(r[i_mese] if i_mese < len(r) else None)
        if not mese:
            continue
        out.append(
            {
                "mese": mese,
                "pun": parse_number(r[i_pun] if i_pun < len(r) else None),
                "psv": parse_number(r[i_psv] if i_psv < len(r) else None),
            }
        )
    return sorted(out, key=lambda row: row["mese"])


def select_indice_for_bill_period(rows, bill_start: date, bill_end: date):
    if not rows:
        return None, "missing"
    if bill_end < bill_start:
        bill_start, bill_end = bill_end, bill_start
    start_month = month_key_from_date(bill_start)
    end_month = month_key_from_date(bill_end)
    in_period = [r for r in rows if start_month <= r["mese"] <= end_month]
    if in_period:
        return in_period[-1], "period"
    before_end = [r for r in rows if r["mese"] <= end_month]
    if before_end:
        return before_end[-1], "before_end"
    return rows[-1], "latest_available"


def tariff_month_key(path: Path) -> str:
    try:
        parts = path.relative_to(TARIFFE_BASE).parts
    except ValueError:
        parts = path.parts
    for i, part in enumerate(parts):
        if re.match(r"^\d{4}-\d{2}$", part):
            return part
        if re.match(r"^\d{4}$", part) and i + 1 < len(parts) and re.match(r"^\d{2}$", parts[i + 1]):
            return f"{part}-{parts[i + 1]}"
    match = re.search(r"(\d{4})-(\d{2})", path.name)
    if match:
        return f"{match.group(1)}-{match.group(2)}"
    return ""



def tariff_matches_segment(path: Path, segmento: str) -> bool:
    segment_norm = clean_text(segmento).upper()
    name = path.name.lower()
    parts = {p.lower() for p in path.parts}
    if path.suffix.lower() != ".xlsx" or name.startswith("~$"):
        return False
    if segment_norm == "BUSINESS":
        is_micro = "microbusiness" in parts or "microbusiness" in name or "micro business" in str(path).lower()
        return not is_micro and ("business" in parts or "business" in name)
    if segment_norm == "MICROBUSINESS":
        return "microbusiness" in parts or "microbusiness" in name or "micro business" in str(path).lower()
    return ("residenziale" in parts or "template" in name) and "business" not in name


def tariff_segment_fallbacks(segmento: str) -> tuple[str, ...]:
    segment_norm = clean_text(segmento).upper()
    if segment_norm == "MICROBUSINESS":
        return ("MICROBUSINESS", "BUSINESS")
    return (segment_norm,)


def _tariff_file_candidates_for_segment_exact(segmento: str, provider: str = "ILLUMIA"):
    provider_norm = normalize_provider(provider)
    if provider_norm in {"EON", "CVE"}:
        if not EON_TARIFFE_DIR.exists():
            return []
        prefix = provider_norm.lower()
        candidates = [
            p for p in EON_TARIFFE_DIR.glob(f"{prefix}_tariffe_*.xlsx") if p.is_file() and not p.name.startswith("~$")
        ]
    else:
        if not TARIFFE_BASE.exists():
            return []
        candidates = [p for p in TARIFFE_BASE.rglob("*.xlsx") if tariff_matches_segment(p, segmento)]
    candidates.sort(key=lambda p: (tariff_month_key(p), p.stat().st_mtime, p.name))
    return candidates


def tariff_file_candidates_for_segment(segmento: str, provider: str = "ILLUMIA"):
    for effective_segment in tariff_segment_fallbacks(segmento):
        candidates = _tariff_file_candidates_for_segment_exact(effective_segment, provider)
        if candidates:
            return candidates
    return []


def select_tariffe_file_from_candidates(candidates, target_month: str = ""):
    candidates = list(candidates)
    if not candidates:
        return None
    target_month = clean_text(target_month)
    if not target_month:
        return candidates[-1]
    with_month = [p for p in candidates if tariff_month_key(p)]
    if not with_month:
        return candidates[-1]
    exact = [p for p in with_month if tariff_month_key(p) == target_month]
    return exact[-1] if exact else None


def load_latest_eon_tariffe_file():
    candidates = tariff_file_candidates_for_segment("RESIDENZIALE", "EON")
    return candidates[-1] if candidates else None


def load_tariffe_file_for_segment(
    segmento: str,
    provider: str = "ILLUMIA",
    selection_mode: str = "LATEST",
    target_month: str = "",
    commodity: str = "",
):
    offer_file, _effective_segment = load_tariffe_file_for_segment_with_effective_segment(
        segmento, provider, selection_mode, target_month, commodity
    )
    return offer_file


def load_tariffe_file_for_segment_with_effective_segment(
    segmento: str,
    provider: str = "ILLUMIA",
    selection_mode: str = "LATEST",
    target_month: str = "",
    commodity: str = "",
):
    requested_segment = clean_text(segmento).upper()
    for effective_segment in tariff_segment_fallbacks(requested_segment):
        candidates = _tariff_file_candidates_for_segment_exact(effective_segment, provider)
        if not candidates:
            continue
        if normalize_tariff_selection_mode(selection_mode) == "PERIOD":
            selected = select_tariffe_file_from_candidates(candidates, target_month)
            if selected and tariff_file_has_context_options(selected, provider, effective_segment, commodity):
                return selected, effective_segment
            continue
        for selected in reversed(candidates):
            if tariff_file_has_context_options(selected, provider, effective_segment, commodity):
                return selected, effective_segment
        return candidates[-1], effective_segment
    return None, requested_segment


def tariff_file_has_context_options(path: Path, provider: str, segmento: str, commodity: str = "") -> bool:
    rows = load_tariffe_from_path(path)
    filtered_rows = filter_rows_by_context(rows, provider, segmento)
    return rows_have_tariff_options(filtered_rows, commodity)


def get_file_valid_range(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if "tariffe" not in wb.sheetnames:
        return None, None
    ws = wb["tariffe"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None, None
    header = [clean_text(c).lower() for c in rows[0]]
    if "valid_from" not in header or "valid_to" not in header:
        return None, None
    i_from = header.index("valid_from")
    i_to = header.index("valid_to")
    dfs, dts = [], []
    for r in rows[1:]:
        if not r:
            continue
        vf = parse_date_any(r[i_from] if i_from < len(r) else None)
        vt = parse_date_any(r[i_to] if i_to < len(r) else None)
        if vf:
            dfs.append(vf)
        if vt:
            dts.append(vt)
    if not dfs or not dts:
        return None, None
    return min(dfs), max(dts)


def valid_range_from_rows(rows):
    dfs, dts = [], []
    for r in rows:
        vf = parse_date_any(r.get("valid_from"))
        vt = parse_date_any(r.get("valid_to"))
        if vf:
            dfs.append(vf)
        if vt:
            dts.append(vt)
    return (min(dfs) if dfs else None, max(dts) if dts else None)


def load_tariffe_from_path(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "tariffe" not in wb.sheetnames:
        return []
    ws = wb["tariffe"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [clean_text(c).lower() for c in rows[0]]
    out = []
    for r in rows[1:]:
        if not any(r):
            continue
        d = {key: (r[i] if i < len(r) else None) for i, key in enumerate(header)}
        d["commodity"] = clean_text(d.get("commodity")).upper()
        d["offer_type"] = clean_text(d.get("offer_type")).upper()
        d["offer_name"] = clean_text(d.get("offer_name"))
        d["component"] = clean_text(d.get("component")).lower()
        d["provider"] = normalize_provider(d.get("provider"))
        d["segmento"] = clean_text(d.get("segmento")).upper()
        d["is_sellable"] = clean_text(d.get("is_sellable")).upper()
        d["value_num"] = parse_number(d.get("value"))
        out.append(d)
    return out


def filter_rows_by_context(rows, provider, segmento):
    provider_norm = normalize_provider(provider)
    segment_norm = clean_text(segmento).upper()
    if provider_norm == "ILLUMIA":
        return [r for r in rows if normalize_provider(r.get("provider")) == "ILLUMIA"]
    return [
        r
        for r in rows
        if normalize_provider(r.get("provider")) == provider_norm
        and clean_text(r.get("segmento")).upper() == segment_norm
    ]


def rows_have_tariff_options(rows, commodity: str = "") -> bool:
    commodity_norm = clean_text(commodity).upper()
    if commodity_norm:
        return bool(offer_names(rows, commodity_norm, "VARIABILE") or offer_names(rows, commodity_norm, "FISSA"))
    return any(row_is_sellable(r) and clean_text(r.get("offer_name")) for r in rows)


def filter_rows_by_context_with_fallback(rows, provider, segmento, commodity: str = ""):
    segment_norm = clean_text(segmento).upper()
    primary_rows = filter_rows_by_context(rows, provider, segment_norm)
    if segment_norm == "MICROBUSINESS" and not rows_have_tariff_options(primary_rows, commodity):
        business_rows = filter_rows_by_context(rows, provider, "BUSINESS")
        if rows_have_tariff_options(business_rows, commodity):
            return business_rows, "BUSINESS"
    return primary_rows, segment_norm


def cve_row_matches_annual_context(row, annual_consumption: float, over70: bool) -> bool:
    requires_over70 = bool_from_data(row.get("requires_over70"))
    if over70:
        return requires_over70
    if requires_over70:
        return False
    annual = max(0.0, float(annual_consumption or 0.0))
    annual_min = clean_text(row.get("annual_min"))
    annual_max = clean_text(row.get("annual_max"))
    min_value = parse_number(annual_min) if annual_min else None
    max_value = parse_number(annual_max) if annual_max else None
    if min_value is not None and annual <= min_value:
        return False
    if max_value is not None and annual > max_value:
        return False
    return True


def filter_cve_rows_by_annual_context(rows, commodity, annual_consumption, over70=False):
    commodity_norm = clean_text(commodity).upper()
    annual = max(0.0, float(annual_consumption or 0.0))
    return [
        row
        for row in rows
        if clean_text(row.get("commodity")).upper() == commodity_norm
        and cve_row_matches_annual_context(row, annual, bool_from_data(over70))
    ]


def filter_rows_by_offer(rows, commodity, offer_type, offer_name):
    c = commodity.upper()
    o = offer_type.upper()
    return [r for r in rows if r.get("commodity") == c and r.get("offer_type") == o and clean_text(r.get("offer_name")) == offer_name]


def row_is_sellable(row):
    marker = clean_text(row.get("is_sellable")).upper()
    return marker not in {"FALSE", "NO", "0"}


def offer_names(rows, commodity, offer_type):
    c = commodity.upper()
    o = offer_type.upper()
    subset = [r for r in rows if r.get("commodity") == c and r.get("offer_type") == o and row_is_sellable(r)]
    counts = {}
    for r in subset:
        name = clean_text(r.get("offer_name"))
        if name:
            counts[name] = counts.get(name, 0) + 1
    return [name for name, _count in sorted(counts.items(), key=lambda x: (-x[1], x[0]))]


def select_offer_name(rows, commodity, offer_type, segmento, preferred=""):
    names = offer_names(rows, commodity, offer_type)
    preferred = clean_text(preferred)
    if preferred and preferred in names:
        return preferred
    if segmento == "BUSINESS" and offer_type.upper() == "FISSA" and not names:
        return ""
    if not names:
        return ""
    return names[0]


def offer_options_payload():
    payload = {}
    for provider in PROVIDERS:
        for segmento in SEGMENTS:
            for commodity in ("GAS", "EE"):
                rows = []
                for effective_segment in tariff_segment_fallbacks(segmento):
                    segment_rows = []
                    for offer_file in _tariff_file_candidates_for_segment_exact(effective_segment, provider):
                        segment_rows.extend(
                            filter_rows_by_context(load_tariffe_from_path(offer_file), provider, effective_segment)
                        )
                    if rows_have_tariff_options(segment_rows, commodity):
                        rows = segment_rows
                        break
                key = f"{provider}|{segmento}|{commodity}"
                payload[key] = {
                    "VARIABILE": offer_names(rows, commodity, "VARIABILE"),
                    "FISSA": offer_names(rows, commodity, "FISSA"),
                }
    return payload


def comp_sum(rows, commodity, offer, component):
    c = commodity.upper()
    o = offer.upper()
    comp = component.lower()
    return float(sum(r.get("value_num", 0.0) for r in rows if r.get("commodity") == c and r.get("offer_type") == o and r.get("component") == comp))


def comp_first(rows, commodity, offer, component):
    c = commodity.upper()
    o = offer.upper()
    comp = component.lower()
    for r in rows:
        if r.get("commodity") == c and r.get("offer_type") == o and r.get("component") == comp:
            return float(r.get("value_num", 0.0))
    return 0.0


def calc_illumia_vendite(rows_offer, commodity, offer, consumo, pun, psv, billing_divisor, include_dispbt=True):
    comm = commodity.upper()
    off = offer.upper()

    def fixed_annua(c, o):
        base = comp_sum(rows_offer, c, o, "ccv_quota_fissa")
        disp = comp_sum(rows_offer, c, o, "dispbt") if include_dispbt else 0.0
        return base + disp

    if comm == "EE":
        if off == "VARIABILE":
            fee = comp_sum(rows_offer, "EE", "VARIABILE", "fee_energia")
            ccv_var = comp_sum(rows_offer, "EE", "VARIABILE", "ccv_quota_variabile")
            sbil = comp_sum(rows_offer, "EE", "VARIABILE", "sbilanciamento")
            prezzo = (pun * 1.10) + fee + ccv_var + sbil
            return float(consumo * prezzo), float(fixed_annua("EE", "VARIABILE") / float(billing_divisor))
        prezzo = comp_first(rows_offer, "EE", "FISSA", "prezzo_mono")
        if prezzo == 0.0:
            prezzo = max(comp_first(rows_offer, "EE", "FISSA", "prezzo_f1"), comp_first(rows_offer, "EE", "FISSA", "prezzo_f23"))
        return float(consumo * prezzo), float(fixed_annua("EE", "FISSA") / float(billing_divisor))

    if off == "VARIABILE":
        fee = comp_sum(rows_offer, "GAS", "VARIABILE", "fee_energia")
        ccv_var = comp_sum(rows_offer, "GAS", "VARIABILE", "ccv_quota_variabile")
        bil = comp_sum(rows_offer, "GAS", "VARIABILE", "bilanciamento")
        prezzo = psv + fee + ccv_var + bil
        return float(consumo * prezzo), float(fixed_annua("GAS", "VARIABILE") / float(billing_divisor))
    prezzo = comp_first(rows_offer, "GAS", "FISSA", "prezzo_fisso")
    prezzo += comp_sum(rows_offer, "GAS", "FISSA", "ccv_quota_variabile")
    return float(consumo * prezzo), float(fixed_annua("GAS", "FISSA") / float(billing_divisor))


def format_eur(value):
    if isinstance(value, str):
        return value
    amount = float(value)
    sign = "-" if amount < 0 else ""
    formatted = f"{abs(amount):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{sign}€ {formatted}"


def format_percent(value):
    if value is None:
        return "N.D."
    return f"{float(value) * 100:.2f}".replace(".", ",") + "%"


def tax_incidence_ratio(total_tax, subtotal):
    if subtotal <= 0:
        return None
    return max(0.0, float(total_tax)) / float(subtotal)


def comparison_subtotal(vals, commodity):
    subtotal = (
        float(vals["vendita_consumo"])
        + float(vals["rete_consumi"])
        + float(vals["vendita_fissa"])
        + float(vals["rete_fissa"])
        + float(vals["sconti"])
        + float(vals["ricalcoli"])
        + float(vals.get("bonus_sociale", 0.0))
        + float(vals["arrotondamenti"])
        + float(vals.get("servizi_accessori", 0.0))
    )
    if commodity == "EE":
        subtotal += float(vals["quota_potenza"])
    return subtotal


def comparison_total(vals, commodity):
    return comparison_subtotal(vals, commodity) + float(vals["accise_iva"])


def comparison_value(vals, key, commodity):
    if key == "vendita_fissa_luce":
        return float(vals["vendita_fissa"]) if commodity == "EE" else "N.A."
    if key == "vendita_fissa_gas":
        return float(vals["vendita_fissa"]) if commodity == "GAS" else "N.A."
    if key == "quota_potenza":
        return float(vals["quota_potenza"]) if commodity == "EE" else "N.A."
    if key in {"accise", "iva"}:
        value = vals.get(key)
        return "N.D." if value is None else float(value)
    if key == "totale":
        return comparison_total(vals, commodity)
    return float(vals[key])


def build_comparison_table_rows(values):
    comm = values["commodity"]
    servizi_accessori_label = f"Servizi accessori (IVA {values['servizi_accessori_iva_label']})"
    accise_label = "Accise + Addizionale regionale" if comm == "GAS" else "Accise"
    rows_config = [
        ("vendita_consumo", "Vendita Consumo"),
        ("rete_consumi", "Rete e oneri di sistema Consumi"),
        ("vendita_fissa_luce", "Vendita Fissa Luce"),
        ("vendita_fissa_gas", "Vendita Fissa Gas"),
        ("rete_fissa", "Rete e oneri di sistema Fissa"),
        ("quota_potenza", "Quota Potenza"),
        ("sconti", "Sconti"),
        ("ricalcoli", "Ricalcoli/Partite pregresse"),
        ("bonus_sociale", "Bonus Sociale"),
        ("arrotondamenti", "Arrotondamenti"),
        ("servizi_accessori", servizi_accessori_label),
        ("accise", accise_label),
        ("iva", "IVA"),
        ("accise_iva", "Accise e Iva"),
        ("totale", "Totale"),
    ]
    out = []
    for key, label in rows_config:
        if comm == "GAS" and key in {"vendita_fissa_luce", "quota_potenza"}:
            continue
        cells = [format_eur(comparison_value(values["bolletta"], key, comm))]
        for column in values["offer_columns"]:
            value = comparison_value(column["vals"], key, comm) if column["has_offer"] else "N.D."
            cells.append(format_eur(value))
        row = {
            "voce": label,
            "cells": cells,
            "bolletta": cells[0],
        }
        if len(cells) >= 3:
            row["variabile"] = cells[1]
            row["fissa"] = cells[2]
        out.append(row)
    return out


def fiscal_unit(commodity: str) -> str:
    return "Smc" if commodity == "GAS" else "kWh"


def format_fiscal_parameters(data, calc=None) -> str:
    comm = data.get("commodity", "EE")
    return (
        f"Prima casa/residente: {primary_home_label(data.get('tax_primary_home'))} | "
        f"Potenza: {float(data.get('tax_power_kw', 0.0)):g} kW | "
        f"Consumo annuo: {float(data.get('tax_annual_consumption', 0.0)):g} {fiscal_unit(comm)}/anno | "
        f"Regione: {normalize_region(data.get('tax_region'))} | "
        f"IVA servizi accessori: {normalize_accessory_services_vat_label(data.get('servizi_accessori_iva'))}"
    )


def calculate_tax_breakdown(data, calc, vals, commodity):
    period_consumption = max(0.0, float(data.get("consumo", 0.0)))
    accessory_vat_rate_value = accessory_services_vat_rate(data.get("servizi_accessori_iva"))
    accessory_services = float(vals.get("servizi_accessori", 0.0))
    social_bonus = float(vals.get("bonus_sociale", 0.0))
    taxable_supply_subtotal = comparison_subtotal(vals, commodity) - accessory_services - social_bonus
    accessory_services_vat = accessory_services * accessory_vat_rate_value

    if commodity == "EE":
        power_kw = max(0.0, float(data.get("tax_power_kw", 0.0)))
        residential = data.get("segmento") == "RESIDENZIALE"
        primary_home = normalize_primary_home(data.get("tax_primary_home")) == "SI"
        exempt_kwh = EE_RESIDENTIAL_EXEMPT_KWH_PER_MONTH * float(calc.get("billing_months", 1))
        if residential and primary_home and power_kw <= 3.0:
            taxable_kwh = max(0.0, period_consumption - exempt_kwh)
        else:
            taxable_kwh = period_consumption
        excise = taxable_kwh * EE_EXCISE_RATE
        vat_rate = 0.10 if residential and primary_home else 0.22
        vat = ((taxable_supply_subtotal + excise) * vat_rate) + accessory_services_vat
        return {
            "accise": excise,
            "iva": vat,
            "accise_iva": excise + vat,
            "addizionale_regionale": 0.0,
        }

    annual_consumption = max(0.0, float(data.get("tax_annual_consumption", 0.0)))
    excise = annual_progressive_tax_for_period(period_consumption, annual_consumption, GAS_EXCISE_BRACKETS)
    regional = period_consumption * float(GAS_REGIONAL_ADDITIONAL_MIN_RATES[normalize_region(data.get("tax_region"))])
    variable_base = (
        float(vals["vendita_consumo"])
        + float(vals["rete_consumi"])
        + float(vals["sconti"])
        + excise
        + regional
    )
    fixed_base = (
        float(vals["vendita_fissa"])
        + float(vals["rete_fissa"])
        + float(vals["ricalcoli"])
        + float(vals["arrotondamenti"])
    )
    reduced_ratio = min(annual_consumption, GAS_VAT_REDUCED_ANNUAL_THRESHOLD) / annual_consumption if annual_consumption else 0.0
    reduced_ratio = min(max(reduced_ratio, 0.0), 1.0)
    vat = (
        (variable_base * reduced_ratio * 0.10)
        + (variable_base * (1.0 - reduced_ratio) * 0.22)
        + (fixed_base * 0.22)
        + accessory_services_vat
    )
    return {
        "accise": excise + regional,
        "iva": vat,
        "accise_iva": excise + regional + vat,
        "addizionale_regionale": regional,
    }


def calculate_accise_iva(data, calc, vals, commodity):
    return calculate_tax_breakdown(data, calc, vals, commodity)["accise_iva"]


def bill_tax_incidence_ratio(vals, commodity):
    subtotal = comparison_subtotal(vals, commodity)
    return tax_incidence_ratio(vals.get("accise_iva", 0.0), subtotal)


def cap_tax_breakdown_to_bill_incidence(tax, vals, commodity, bill_ratio):
    if bill_ratio is None:
        return tax
    subtotal = comparison_subtotal(vals, commodity)
    current_total = max(0.0, float(tax.get("accise_iva", 0.0)))
    capped_total = max(0.0, float(bill_ratio)) * max(0.0, subtotal)
    if subtotal <= 0 or current_total <= 0 or current_total <= capped_total:
        return tax
    scale = capped_total / current_total
    accise = float(tax.get("accise", 0.0)) * scale
    iva = float(tax.get("iva", 0.0)) * scale
    return {
        "accise": accise,
        "iva": iva,
        "accise_iva": accise + iva,
        "addizionale_regionale": float(tax.get("addizionale_regionale", 0.0)) * scale,
    }


def offer_choice_from_data(data, provider, offer_type):
    provider_norm = normalize_provider(provider)
    suffix = provider_norm.lower()
    choice_type = "var" if offer_type.upper() == "VARIABILE" else "fix"
    choice = clean_text(data.get(f"offer_{choice_type}_choice_{suffix}", ""))
    if not choice and normalize_provider(data.get("provider", provider_norm)) == provider_norm:
        choice = clean_text(data.get(f"offer_{choice_type}_choice", ""))
    return choice


def calculate_provider_result(data, base_calc, provider):
    bill_start = data["bill_start"]
    bill_end = data["bill_end"]
    segmento = data["segmento"]
    commodity = data["commodity"]
    provider_norm = normalize_provider(provider)
    consumo = float(data["consumo"])
    billing_divisor = base_calc["billing_divisor"]
    pun = base_calc["pun"]
    psv = base_calc["psv"]

    offer_file, effective_segment = load_tariffe_file_for_segment_with_effective_segment(
        segmento,
        provider_norm,
        base_calc.get("tariff_selection_mode", "LATEST"),
        base_calc.get("tariff_target_month", ""),
        commodity,
    )
    if offer_file:
        raw_tariffe_rows = load_tariffe_from_path(offer_file)
        tariffe_rows, effective_segment = filter_rows_by_context_with_fallback(
            raw_tariffe_rows, provider_norm, effective_segment, commodity
        )
        if provider_norm == "CVE":
            tariffe_rows = filter_cve_rows_by_annual_context(
                tariffe_rows, commodity, data.get("tax_annual_consumption", 0.0), data.get("cve_over70")
            )
    else:
        tariffe_rows = []
    offer_valid_from, offer_valid_to = get_file_valid_range(offer_file) if offer_file else (None, None)
    offer_var = select_offer_name(
        tariffe_rows,
        commodity,
        "VARIABILE",
        effective_segment,
        offer_choice_from_data(data, provider_norm, "VARIABILE"),
    )
    offer_fix = select_offer_name(
        tariffe_rows,
        commodity,
        "FISSA",
        effective_segment,
        offer_choice_from_data(data, provider_norm, "FISSA"),
    )

    sconto_var = float(data.get("ill_sconto_var", -3.0)) if provider_norm == "ILLUMIA" else 0.0
    sconto_fix = float(data.get("ill_sconto_fix", -3.0)) if provider_norm == "ILLUMIA" else 0.0
    rows_var = []
    rows_fix = []

    if offer_var:
        rows_var = filter_rows_by_offer(tariffe_rows, commodity, "VARIABILE", offer_var)
        v_cons, v_fix = calc_illumia_vendite(rows_var, commodity, "VARIABILE", consumo, pun, psv, billing_divisor)
        if provider_norm != "ILLUMIA":
            sconto_var = comp_sum(rows_var, commodity, "VARIABILE", "sconto_bonus")
    else:
        v_cons, v_fix = 0.0, 0.0

    if offer_fix:
        rows_fix = filter_rows_by_offer(tariffe_rows, commodity, "FISSA", offer_fix)
        f_cons, f_fix = calc_illumia_vendite(rows_fix, commodity, "FISSA", consumo, pun, psv, billing_divisor)
        if provider_norm != "ILLUMIA":
            sconto_fix = comp_sum(rows_fix, commodity, "FISSA", "sconto_bonus")
    else:
        f_cons, f_fix = 0.0, 0.0

    selected_valid_from, selected_valid_to = valid_range_from_rows(rows_var + rows_fix)
    if selected_valid_from or selected_valid_to:
        offer_valid_from, offer_valid_to = selected_valid_from, selected_valid_to
    offer_period_warning = ""
    if bill_period_outside_offer_validity(bill_start, bill_end, offer_valid_from, offer_valid_to):
        offer_period_warning = "Il periodo bolletta NON rientra nella validità dell'offerta selezionata."

    return {
        "provider": provider_norm,
        "provider_label": provider_label(provider_norm),
        "offer_file": str(offer_file) if offer_file else "",
        "requested_segment": segmento,
        "tariff_segment": effective_segment,
        "offer_valid_from": offer_valid_from,
        "offer_valid_to": offer_valid_to,
        "offer_expiry_label": date_label_it(offer_valid_to),
        "offer_period_warning": offer_period_warning,
        "offer_var": offer_var,
        "offer_fix": offer_fix,
        "v_cons": v_cons,
        "v_fix": v_fix,
        "f_cons": f_cons,
        "f_fix": f_fix,
        "sconto_var": sconto_var,
        "sconto_fix": sconto_fix,
    }


def build_offer_column_values(data, calc, base_values, provider_result, offer_type, bill_ratio=None):
    comm = data["commodity"]
    vals = base_values.copy()
    if offer_type == "VARIABILE":
        vals["vendita_consumo"] = provider_result["v_cons"]
        vals["vendita_fissa"] = provider_result["v_fix"]
        vals["sconti"] = float(provider_result.get("sconto_var", data.get("ill_sconto_var", -3.0)))
        offer_name = provider_result["offer_var"]
    else:
        vals["vendita_consumo"] = provider_result["f_cons"]
        vals["vendita_fissa"] = provider_result["f_fix"]
        vals["sconti"] = float(provider_result.get("sconto_fix", data.get("ill_sconto_fix", -3.0)))
        offer_name = provider_result["offer_fix"]
    vals["ricalcoli"] = 0.0
    vals["arrotondamenti"] = 0.0
    vals["servizi_accessori"] = 0.0
    subtotal = comparison_subtotal(vals, comm)
    raw_tax = calculate_tax_breakdown(data, calc, vals, comm)
    raw_tax_ratio = tax_incidence_ratio(raw_tax["accise_iva"], subtotal)
    tax = cap_tax_breakdown_to_bill_incidence(raw_tax, vals, comm, bill_ratio)
    final_tax_ratio = tax_incidence_ratio(tax["accise_iva"], subtotal)
    tax_cap_applied = bool(
        bill_ratio is not None and raw_tax_ratio is not None and raw_tax_ratio > float(bill_ratio) + 1e-9
    )
    vals["accise"] = tax["accise"]
    vals["iva"] = tax["iva"]
    vals["accise_iva"] = tax["accise_iva"]
    type_label = "Variabile" if offer_type == "VARIABILE" else "Fissa"
    return {
        "provider": provider_result["provider"],
        "provider_label": provider_result["provider_label"],
        "offer_type": offer_type,
        "offer_type_label": type_label,
        "offer_name": offer_name,
        "label": f"{provider_result['provider_label']} {type_label}",
        "vals": vals,
        "has_offer": bool(offer_name),
        "tax_cap_applied": tax_cap_applied,
        "tax_cap_status_label": "Si" if tax_cap_applied else "No",
        "bill_tax_ratio": bill_ratio,
        "bill_tax_ratio_label": format_percent(bill_ratio),
        "raw_tax_ratio": raw_tax_ratio,
        "raw_tax_ratio_label": format_percent(raw_tax_ratio),
        "tax_ratio": final_tax_ratio,
        "tax_ratio_label": format_percent(final_tax_ratio),
    }


def build_comparison_values(data, calc):
    comm = data["commodity"]
    accessory_vat_label = normalize_accessory_services_vat_label(data.get("servizi_accessori_iva"))
    accessory_vat_rate = accessory_services_vat_rate(accessory_vat_label)
    b_vals = {k: float(data.get(f"b_{k}", 0.0)) for k in KEYS}
    b_vals["bonus_sociale"] = -abs(float(b_vals.get("bonus_sociale", 0.0)))
    if comm == "GAS":
        b_vals["quota_potenza"] = 0.0
    b_vals["accise"] = None
    b_vals["iva"] = None
    bill_ratio = bill_tax_incidence_ratio(b_vals, comm)

    provider_results = calc.get("provider_results") or [
        {
            "provider": normalize_provider(calc.get("provider", data.get("provider", "ILLUMIA"))),
            "provider_label": provider_label(calc.get("provider", data.get("provider", "ILLUMIA"))),
            "offer_var": calc.get("offer_var", ""),
            "offer_fix": calc.get("offer_fix", ""),
            "v_cons": calc.get("v_cons", 0.0),
            "v_fix": calc.get("v_fix", 0.0),
            "f_cons": calc.get("f_cons", 0.0),
            "f_fix": calc.get("f_fix", 0.0),
            "sconto_var": calc.get("sconto_var", data.get("ill_sconto_var", -3.0)),
            "sconto_fix": calc.get("sconto_fix", data.get("ill_sconto_fix", -3.0)),
        }
    ]
    offer_columns = []
    for provider_result in provider_results:
        offer_columns.append(build_offer_column_values(data, calc, b_vals, provider_result, "VARIABILE", bill_ratio))
        offer_columns.append(build_offer_column_values(data, calc, b_vals, provider_result, "FISSA", bill_ratio))

    first_var = next((column for column in offer_columns if column["offer_type"] == "VARIABILE"), None)
    first_fix = next((column for column in offer_columns if column["offer_type"] == "FISSA"), None)
    return {
        "commodity": comm,
        "bolletta": b_vals,
        "offer_columns": offer_columns,
        "variabile": first_var["vals"] if first_var else b_vals.copy(),
        "fissa": first_fix["vals"] if first_fix else b_vals.copy(),
        "bill_tax_ratio": bill_ratio,
        "bill_tax_ratio_label": format_percent(bill_ratio),
        "servizi_accessori_iva_label": accessory_vat_label,
        "servizi_accessori_iva_rate": accessory_vat_rate,
        "has_offer_var": bool(first_var and first_var["has_offer"]),
        "has_offer_fix": bool(first_fix and first_fix["has_offer"]),
    }


def prepare_comparison(data):
    bill_start = data["bill_start"]
    bill_end = data["bill_end"]
    bill_offer_expiry = data.get("bill_offer_expiry")
    segmento = data["segmento"]
    commodity = data["commodity"]
    providers = normalize_providers(data.get("providers") or data.get("provider", "ILLUMIA"))
    if not providers:
        providers = [normalize_provider(data.get("provider", "ILLUMIA"))]
    tariff_selection_mode = normalize_tariff_selection_mode(data.get("tariff_selection_mode"))
    billing_months = billing_months_from_dates(bill_start, bill_end)
    billing_divisor = billing_divisor_from_months(billing_months)
    tariff_target_month = month_key_from_date(bill_end)

    indici_rows = load_indici_rows(INDICI_XLSX)
    indice, indice_reason = select_indice_for_bill_period(indici_rows, bill_start, bill_end)
    pun = float(indice["pun"]) if indice else 0.0
    psv = float(indice["psv"]) if indice else 0.0

    calc = {
        "nome_cliente": clean_text(data.get("nome_cliente")) or "Cliente",
        "indirizzo_fornitura": clean_text(data.get("indirizzo_fornitura")),
        "pod_pdr": clean_text(data.get("pod_pdr")),
        "commodity": commodity,
        "bill_tariff_type": normalize_bill_tariff_type(data.get("bill_tariff_type")),
        "bill_tariff_type_label": bill_tariff_type_label(data.get("bill_tariff_type")),
        "tax_primary_home": normalize_primary_home(data.get("tax_primary_home")),
        "tax_primary_home_label": primary_home_label(data.get("tax_primary_home")),
        "tax_power_kw": float(data.get("tax_power_kw", 0.0)),
        "tax_annual_consumption": float(data.get("tax_annual_consumption", 0.0)),
        "tax_annual_consumption_label": (
            f"{float(data.get('tax_annual_consumption', 0.0)):g} {fiscal_unit(commodity)}/anno"
        ),
        "tax_region": normalize_region(data.get("tax_region")),
        "fiscal_parameters_label": format_fiscal_parameters(data),
        "comparison_datetime": comparison_datetime_from_data(data.get("comparison_datetime")),
        "comparison_datetime_label": comparison_datetime_label(data.get("comparison_datetime")),
        "providers": providers,
        "providers_label": provider_list_label(providers),
        "tariff_selection_mode": tariff_selection_mode,
        "tariff_selection_mode_label": tariff_selection_mode_label(tariff_selection_mode),
        "tariff_target_month": tariff_target_month,
        "provider": providers[0],
        "provider_label": provider_label(providers[0]),
        "billing_months": billing_months,
        "billing_divisor": billing_divisor,
        "billing_label": billing_label_from_months(billing_months),
        "period_label": bill_period_label(bill_start, bill_end),
        "bill_offer_expiry": bill_offer_expiry,
        "bill_offer_expiry_label": date_label_it(bill_offer_expiry),
        "indice": indice,
        "indice_reason": indice_reason,
        "pun": pun,
        "psv": psv,
        "servizi_accessori_iva_label": normalize_accessory_services_vat_label(data.get("servizi_accessori_iva")),
        "cve_over70": bool_from_data(data.get("cve_over70")),
        "cve_over70_label": "Si" if bool_from_data(data.get("cve_over70")) else "No",
        "cve_selected": "CVE" in providers,
    }
    provider_results = [calculate_provider_result(data, calc, provider) for provider in providers]
    calc["provider_results"] = provider_results
    primary = provider_results[0] if provider_results else {}
    calc.update(
        {
            "offer_file": primary.get("offer_file", ""),
            "offer_valid_from": primary.get("offer_valid_from"),
            "offer_valid_to": primary.get("offer_valid_to"),
            "offer_expiry_label": calc["bill_offer_expiry_label"],
            "offer_period_warning": " ".join(
                f"{result['provider_label']}: {result['offer_period_warning']}"
                for result in provider_results
                if result.get("offer_period_warning")
            ),
            "offer_var": primary.get("offer_var", ""),
            "offer_fix": primary.get("offer_fix", ""),
            "v_cons": primary.get("v_cons", 0.0),
            "v_fix": primary.get("v_fix", 0.0),
            "f_cons": primary.get("f_cons", 0.0),
            "f_fix": primary.get("f_fix", 0.0),
            "sconto_var": primary.get("sconto_var", 0.0),
            "sconto_fix": primary.get("sconto_fix", 0.0),
        }
    )
    values = build_comparison_values(data, calc)
    columns = [{"label": "Bolletta"}] + [{"label": column["label"]} for column in values["offer_columns"]]
    return {"calc": calc, "values": values, "rows": build_comparison_table_rows(values), "columns": columns}


def find_row_map(ws):
    rm = {}
    for r in range(1, ws.max_row + 1):
        t = clean_text(ws[f"A{r}"].value).lower()
        if not t:
            continue
        if "vendita consumo" in t:
            rm["vendita_consumo"] = r
        elif "vendita fissa luce" in t:
            rm["vendita_fissa_luce"] = r
        elif "vendita fissa" in t and "gas" in t:
            rm["vendita_fissa_gas"] = r
        elif "rete" in t and "consumi" in t:
            rm["rete_consumi"] = r
        elif "rete" in t and "fissa" in t:
            rm["rete_fissa"] = r
        elif "quota potenza" in t:
            rm["quota_potenza"] = r
        elif "sconti" in t:
            rm["sconti"] = r
        elif "ricalcoli" in t or "partite" in t:
            rm["ricalcoli"] = r
        elif "bonus social" in t:
            rm["bonus_sociale"] = r
        elif "arrotondamenti" in t:
            rm["arrotondamenti"] = r
        elif "servizi" in t and "access" in t:
            rm["servizi_accessori"] = r
        elif "accise" in t and "iva" in t:
            rm["accise_iva"] = r
        elif t.startswith("accise"):
            rm["accise"] = r
        elif t == "iva":
            rm["iva"] = r
        elif t == "totale":
            rm["totale"] = r
    if "arrotondamenti" not in rm and "bonus_sociale" not in rm and "accise_iva" in rm:
        candidate = rm["accise_iva"] - 1
        if candidate > 0:
            rm["arrotondamenti"] = candidate
    return rm


def copy_row_format(ws, source_row: int, target_row: int):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        source = ws.cell(row=source_row, column=col)
        target = ws.cell(row=target_row, column=col)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def ensure_export_rows(ws):
    rm = find_row_map(ws)
    missing_count = 0
    if "arrotondamenti" not in rm:
        missing_count += 1
    if "servizi_accessori" not in rm:
        missing_count += 1
    if "accise" not in rm:
        missing_count += 1
    if "iva" not in rm:
        missing_count += 1
    if missing_count == 0:
        return

    insert_at = rm.get("accise_iva", rm.get("totale", ws.max_row))
    ws.insert_rows(insert_at, amount=missing_count)
    source_row = insert_at + missing_count
    for row in range(insert_at, insert_at + missing_count):
        copy_row_format(ws, source_row, row)


def validate_row_map(rm, commodity: str = "EE"):
    required = [
        "vendita_consumo",
        "vendita_fissa_gas",
        "rete_consumi",
        "rete_fissa",
        "sconti",
        "ricalcoli",
        "bonus_sociale",
        "arrotondamenti",
        "servizi_accessori",
        "accise",
        "iva",
        "accise_iva",
        "totale",
    ]
    if commodity == "EE":
        required.extend(["vendita_fissa_luce", "quota_potenza"])
    missing = [key for key in required if key not in rm]
    if missing:
        raise ValueError("Template Excel incompleto: mancano le righe " + ", ".join(missing))


def remove_export_rows_for_commodity(ws, rm, commodity: str):
    if commodity != "GAS":
        return
    rows_to_delete = [rm[key] for key in ("quota_potenza", "vendita_fissa_luce") if key in rm]
    for row in sorted(set(rows_to_delete), reverse=True):
        ws.delete_rows(row, amount=1)


def apply_export_labels(ws, nome_cliente: str, servizi_accessori_iva_label: str = "22%", commodity: str = "EE"):
    accise_label = "Accise + Addizionale regionale" if commodity == "GAS" else "Accise"
    labels = {
        1: nome_cliente or "Cliente",
        3: "VOCE",
        4: "Vendita Consumo",
        5: "Rete e oneri di sistema Consumi",
        6: "Vendita Fissa Luce",
        7: "Vendita Fissa Gas",
        8: "Rete e oneri di sistema Fissa",
        9: "Quota Potenza",
        10: "Sconti",
        11: "Ricalcoli/Partite pregresse",
        12: "Bonus Sociale",
        13: "Arrotondamenti",
        14: f"Servizi accessori (IVA {servizi_accessori_iva_label})",
        15: accise_label,
        16: "IVA",
        17: "Accise e Iva",
        18: "Totale",
    }
    for row, value in labels.items():
        ws[f"A{row}"] = value


def write_export_metadata(ws, prepared, start_col="F"):
    calc = prepared["calc"]
    provider_lines = []
    file_lines = []
    for result in calc.get("provider_results", []):
        provider_lines.append(
            f"{result['provider_label']}: variabile {result['offer_var'] or 'N.D.'}; "
            f"fissa {result['offer_fix'] or 'N.D.'}"
        )
        file_lines.append(f"{result['provider_label']}: {result.get('offer_file') or 'N.D.'}")
    ws[f"{start_col}1"] = f"Fornitori confronto: {calc.get('providers_label', calc.get('provider_label', 'Illumia'))}"
    ws[f"{start_col}2"] = " | ".join(provider_lines) if provider_lines else "Offerte: N.D."
    ws[f"{start_col}3"] = f"Scadenza offerta bolletta: {calc.get('bill_offer_expiry_label', 'N.D.')}"
    ws[f"{start_col}3"].font = Font(bold=True, color="B3261E")
    ws[f"{start_col}4"] = f"File tariffe: {' | '.join(file_lines) if file_lines else 'N.D.'}"
    indice = calc["indice"] or {}
    ws[f"{start_col}5"] = f"Indice PUN/PSV: {indice.get('mese', 'N.D.')} ({INDICI_XLSX.name})"
    ws[f"{start_col}6"] = f"Tipo tariffa bolletta: {calc.get('bill_tariff_type_label', 'Variabile')}"
    ws[f"{start_col}7"] = f"Confronto eseguito: {calc.get('comparison_datetime_label', '')}"
    ws[f"{start_col}8"] = f"IVA servizi accessori: {calc.get('servizi_accessori_iva_label', '22%')}"
    ws[f"{start_col}9"] = f"Consumo annuo stimato: {calc.get('tax_annual_consumption_label', 'N.D.')}"
    ws[f"{start_col}10"] = f"Parametri Accise/IVA: {calc.get('fiscal_parameters_label', '')}"
    ws[f"{start_col}11"] = (
        f"Logica tariffe: {calc.get('tariff_selection_mode_label', 'Ultime tariffe disponibili')}"
    )
    ws[f"{start_col}12"] = f"Indirizzo fornitura: {calc.get('indirizzo_fornitura') or 'N.D.'}"
    ws[f"{start_col}13"] = f"Codice POD/PDR: {calc.get('pod_pdr') or 'N.D.'}"
    ws[f"{start_col}14"] = f"Fornitura: {'Gas' if calc.get('commodity') == 'GAS' else 'Luce'}"
    ws[f"{start_col}15"] = f"Periodo bolletta: {calc.get('period_label', 'N.D.')}"
    if calc.get("cve_selected"):
        ws[f"{start_col}16"] = f"CVE Over 70: {calc.get('cve_over70_label', 'No')}"


def _excel_decimal(value: float) -> str:
    text = f"{float(value):.6f}".rstrip("0").rstrip(".")
    return text or "0"


def apply_accise_formula_conforme(ws, rm, col_letter, servizi_accessori_iva_rate=0.0):
    acc = rm["accise_iva"]
    start = rm["vendita_consumo"]
    end = rm.get("accise", acc) - 1
    if "servizi_accessori" in rm:
        svc = rm["servizi_accessori"]
        rate = _excel_decimal(servizi_accessori_iva_rate)
        ws[f"{col_letter}{acc}"] = (
            f"=IFERROR((SUM({col_letter}{start}:{col_letter}{end})-{col_letter}{svc})"
            f"*(B{acc}-B{svc}*{rate})/(SUM(B{start}:B{end})-B{svc})"
            f"+{col_letter}{svc}*{rate},0)"
        )
    else:
        ws[f"{col_letter}{acc}"] = f"=SUM({col_letter}{start}:{col_letter}{end})*B{acc}/SUM(B{start}:B{end})"


def apply_total_formula(ws, rm, col_letter):
    acc = rm["accise_iva"]
    start = rm["vendita_consumo"]
    end = rm.get("accise", acc) - 1
    ws[f"{col_letter}{rm['totale']}"] = f"=SUM({col_letter}{start}:{col_letter}{end})+{col_letter}{acc}"


def write_column(ws, rm, col, vals, commodity):
    ws[f"{col}{rm['vendita_consumo']}"] = float(vals["vendita_consumo"])
    if commodity == "GAS":
        ws[f"{col}{rm['vendita_fissa_gas']}"] = float(vals["vendita_fissa"])
        if "vendita_fissa_luce" in rm:
            ws[f"{col}{rm['vendita_fissa_luce']}"] = "N.A."
        if "quota_potenza" in rm:
            ws[f"{col}{rm['quota_potenza']}"] = "N.A."
    else:
        ws[f"{col}{rm['vendita_fissa_luce']}"] = float(vals["vendita_fissa"])
        ws[f"{col}{rm['vendita_fissa_gas']}"] = "N.A."
        ws[f"{col}{rm['quota_potenza']}"] = float(vals["quota_potenza"])
    ws[f"{col}{rm['rete_consumi']}"] = float(vals["rete_consumi"])
    ws[f"{col}{rm['rete_fissa']}"] = float(vals["rete_fissa"])
    ws[f"{col}{rm['sconti']}"] = float(vals["sconti"])
    ws[f"{col}{rm['ricalcoli']}"] = float(vals["ricalcoli"])
    if "bonus_sociale" in rm:
        ws[f"{col}{rm['bonus_sociale']}"] = float(vals.get("bonus_sociale", 0.0))
    if "arrotondamenti" in rm:
        ws[f"{col}{rm['arrotondamenti']}"] = float(vals["arrotondamenti"])
    if "servizi_accessori" in rm:
        ws[f"{col}{rm['servizi_accessori']}"] = float(vals.get("servizi_accessori", 0.0))
    for key in ("accise", "iva"):
        if key in rm:
            value = vals.get(key)
            ws[f"{col}{rm[key]}"] = "N.D." if value is None else float(value)


def fill_column_text(ws, rm, col, text):
    for key, r in rm.items():
        if key == "totale" and text == "N.A.":
            continue
        ws[f"{col}{r}"] = text


def build_excel_bytes(data, prepared=None):
    prepared = prepared or prepare_comparison(data)
    wb = openpyxl.load_workbook(TEMPLATE_XLSX)
    ws = wb["Confronto"]
    ensure_export_rows(ws)
    servizi_accessori_iva_label = normalize_accessory_services_vat_label(data.get("servizi_accessori_iva"))
    servizi_accessori_iva_rate = accessory_services_vat_rate(servizi_accessori_iva_label)
    values = prepared["values"]
    comm = values["commodity"]
    apply_export_labels(ws, data.get("nome_cliente", "Cliente"), servizi_accessori_iva_label, comm)
    rm = find_row_map(ws)
    remove_export_rows_for_commodity(ws, rm, comm)
    rm = find_row_map(ws)
    validate_row_map(rm, comm)
    ws["B1"] = None
    ws["C1"] = float(data["consumo"])
    ws["B3"] = "Bolletta"

    b_vals = values["bolletta"]

    write_column(ws, rm, "B", b_vals, comm)
    ws[f"B{rm['accise_iva']}"] = b_vals["accise_iva"]
    apply_total_formula(ws, rm, "B")

    for offset, column in enumerate(values["offer_columns"], start=3):
        col_letter = get_column_letter(offset)
        ws[f"{col_letter}3"] = column["label"]
        if column["has_offer"]:
            write_column(ws, rm, col_letter, column["vals"], comm)
            ws[f"{col_letter}{rm['accise_iva']}"] = column["vals"]["accise_iva"]
            apply_total_formula(ws, rm, col_letter)
        else:
            fill_column_text(ws, rm, col_letter, "N.D.")

    metadata_column = get_column_letter(3 + len(values["offer_columns"]) + 1)
    write_export_metadata(ws, prepared, metadata_column)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
